from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
import shutil
import subprocess
import tempfile
import time
import textwrap
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from bs4 import BeautifulSoup


DEFAULT_SCALE = 2.0
DEFAULT_MAX_VIEWPORT_WIDTH = 16000
DEFAULT_MAX_VIEWPORT_HEIGHT = 12000


@dataclass(frozen=True)
class TableDocument:
    html: str
    title: str = ""
    source_format: str = "unknown"
    estimated_width: int = 1200
    estimated_height: int = 800


@dataclass(frozen=True)
class RenderResult:
    image_path: Path
    html_path: Path | None
    width: int | None
    height: int | None
    browser_path: Path


@dataclass(frozen=True)
class _FontFace:
    font: Any
    codepoints: frozenset[int] | None
    missing_glyph: tuple[tuple[int, int], bytes] | None = None


def table_to_image(
    input_path: str | Path,
    output_path: str | Path | None = None,
    *,
    input_format: str = "auto",
    json_index: int | None = None,
    table_index: int = 0,
    sheet: str | int | None = None,
    scale: float = DEFAULT_SCALE,
    browser_path: str | Path | None = None,
    keep_html: bool = True,
    timeout_seconds: float = 60.0,
) -> RenderResult:
    path = Path(input_path)
    if output_path is None:
        output = path.with_suffix(".png")
    else:
        output = Path(output_path)

    document = document_from_file(
        path,
        input_format=input_format,
        json_index=json_index,
        table_index=table_index,
        sheet=sheet,
    )
    return render_document(
        document,
        output,
        scale=scale,
        browser_path=browser_path,
        keep_html=keep_html,
        timeout_seconds=timeout_seconds,
    )


def document_from_file(
    input_path: str | Path,
    *,
    input_format: str = "auto",
    json_index: int | None = None,
    table_index: int = 0,
    sheet: str | int | None = None,
) -> TableDocument:
    path = Path(input_path)
    fmt = _detect_format(path, input_format)

    if fmt == "html":
        return document_from_html(path.read_text(encoding="utf-8"), source_format="html")
    if fmt == "md":
        rows = rows_from_markdown(path.read_text(encoding="utf-8"))
        return document_from_rows(rows, source_format="md")
    if fmt == "json":
        data = json.loads(path.read_text(encoding="utf-8"))
        return document_from_json(data, json_index=json_index, table_index=table_index)
    if fmt == "jsonl":
        data = _read_jsonl(path, json_index=json_index)
        return document_from_json(data, json_index=None, table_index=table_index)
    if fmt == "xlsx":
        return document_from_xlsx(path, sheet=sheet)
    if fmt == "csv":
        return document_from_rows(rows_from_delimited(path, delimiter=","), source_format="csv")
    if fmt == "tsv":
        return document_from_rows(rows_from_delimited(path, delimiter="\t"), source_format="tsv")

    text = path.read_text(encoding="utf-8")
    return document_from_text(text)


def document_from_text(text: str) -> TableDocument:
    stripped = text.lstrip()
    if "<table" in stripped.lower():
        return document_from_html(text, source_format="html")
    if stripped.startswith("{") or stripped.startswith("["):
        return document_from_json(json.loads(text))
    if "|" in text:
        return document_from_rows(rows_from_markdown(text), source_format="md")
    return document_from_rows(list(csv.reader(text.splitlines())), source_format="csv")


def document_from_html(content: str, *, source_format: str = "html") -> TableDocument:
    soup = BeautifulSoup(content, "html.parser")
    table = soup.find("table")
    if table is not None:
        title = _clean_text(table.find("caption").get_text(" ", strip=True)) if table.find("caption") else ""
        fragment = str(table)
        width, height = estimate_html_size(fragment)
        return TableDocument(
            html=wrap_html(fragment, title=title),
            title=title,
            source_format=source_format,
            estimated_width=width,
            estimated_height=height,
        )

    body = str(soup.body) if soup.body else content
    width, height = estimate_html_size(body)
    return TableDocument(
        html=wrap_html(body),
        source_format=source_format,
        estimated_width=width,
        estimated_height=height,
    )


def document_from_json(
    data: Any,
    *,
    json_index: int | None = None,
    table_index: int = 0,
) -> TableDocument:
    data = _select_json_record(data, json_index=json_index)

    if isinstance(data, dict) and "texts" in data:
        return document_from_hitab_json(data)

    if isinstance(data, dict) and isinstance(data.get("tables"), list):
        tables = data["tables"]
        if not tables:
            raise ValueError("JSON field 'tables' is empty.")
        selected = tables[_bounded_index(table_index, len(tables), "table_index")]
        if isinstance(selected, str):
            return document_from_html(selected, source_format="json:html-table")
        if isinstance(selected, list):
            return document_from_rows(selected, source_format="json:rows")
        if isinstance(selected, dict):
            return document_from_json(selected, json_index=None, table_index=table_index)
        raise ValueError(f"Unsupported table entry type in JSON 'tables': {type(selected).__name__}")

    if isinstance(data, dict):
        for key in ("rows", "data", "table", "values"):
            if key in data:
                value = data[key]
                if isinstance(value, str) and "<table" in value.lower():
                    return document_from_html(value, source_format=f"json:{key}")
                if isinstance(value, list):
                    return document_from_rows(value, source_format=f"json:{key}")
                if isinstance(value, dict):
                    return document_from_json(value, json_index=None, table_index=table_index)
        return document_from_rows([[key, value] for key, value in data.items()], source_format="json:object")

    if isinstance(data, list):
        return document_from_rows(data, source_format="json:rows")

    raise ValueError(f"Unsupported JSON table shape: {type(data).__name__}")


def document_from_hitab_json(data: dict[str, Any]) -> TableDocument:
    rows = _rectangularize(data.get("texts") or [])
    title = _clean_text(data.get("title", ""))
    merged_regions = data.get("merged_regions") or []
    fragment = rows_to_html(rows, title=title, merged_regions=merged_regions)
    width, height = estimate_rows_size(rows)
    return TableDocument(
        html=wrap_html(fragment, title=title),
        title=title,
        source_format="hitab-json",
        estimated_width=width,
        estimated_height=height,
    )


def document_from_xlsx(
    input_path: str | Path,
    *,
    sheet: str | int | None = None,
    add_coordinates: bool = True,
    cell_range: str | None = None,
) -> TableDocument:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("XLSX input requires openpyxl. Install it or run with an environment that has it.") from exc

    workbook = openpyxl.load_workbook(input_path, data_only=True)
    if sheet is None:
        worksheet = workbook.active
    elif isinstance(sheet, int) or str(sheet).isdigit():
        worksheet = workbook.worksheets[int(sheet)]
    else:
        worksheet = workbook[str(sheet)]

    if cell_range:
        from openpyxl.utils.cell import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        rows_iter = list(
            worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )
        )
    else:
        rows_iter = list(worksheet.iter_rows())
    cell_styles: list[list[str]] = []
    column_widths: list[int | None] = []
    row_heights: list[int | None] = []
    if not rows_iter:
        rows = []
        merged_regions = []
    elif add_coordinates:
        from openpyxl.utils import get_column_letter

        first_row = rows_iter[0]
        start_row = first_row[0].row
        start_col = first_row[0].column

        col_letters = [get_column_letter(cell.column) for cell in first_row]
        new_rows = [[""] + col_letters]
        cell_styles = [[""] * (len(first_row) + 1)]
        column_widths = [42] + [_xlsx_column_width(worksheet, cell.column) for cell in first_row]
        row_heights = [24]
        for row in rows_iter:
            row_num = str(row[0].row)
            row_values = [cell.value if cell.value is not None else "" for cell in row]
            new_rows.append([row_num] + row_values)
            cell_styles.append([""] + [_xlsx_cell_style(cell) for cell in row])
            row_heights.append(_xlsx_row_height(worksheet, row[0].row))
        rows = new_rows

        merged_regions = []
        end_row = rows_iter[-1][0].row
        end_col = rows_iter[0][-1].column
        for merged in worksheet.merged_cells.ranges:
            clipped_first_row = max(merged.min_row, start_row)
            clipped_last_row = min(merged.max_row, end_row)
            clipped_first_col = max(merged.min_col, start_col)
            clipped_last_col = min(merged.max_col, end_col)
            if clipped_first_row > clipped_last_row or clipped_first_col > clipped_last_col:
                continue
            anchor = worksheet.cell(merged.min_row, merged.min_col)
            new_rows[clipped_first_row - start_row + 1][clipped_first_col - start_col + 1] = (
                anchor.value if anchor.value is not None else ""
            )
            cell_styles[clipped_first_row - start_row + 1][clipped_first_col - start_col + 1] = _xlsx_cell_style(anchor)
            merged_regions.append({
                "first_row": clipped_first_row - start_row + 1,
                "last_row": clipped_last_row - start_row + 1,
                "first_column": clipped_first_col - start_col + 1,
                "last_column": clipped_last_col - start_col + 1,
            })
    else:
        rows = [
            [cell.value if cell.value is not None else "" for cell in row]
            for row in rows_iter
        ]
        cell_styles = [[_xlsx_cell_style(cell) for cell in row] for row in rows_iter]
        column_widths = [_xlsx_column_width(worksheet, cell.column) for cell in rows_iter[0]]
        row_heights = [_xlsx_row_height(worksheet, row[0].row) for row in rows_iter]
        start_row = rows_iter[0][0].row
        start_col = rows_iter[0][0].column
        end_row = rows_iter[-1][0].row
        end_col = rows_iter[0][-1].column
        merged_regions = []
        for merged in worksheet.merged_cells.ranges:
            clipped_first_row = max(merged.min_row, start_row)
            clipped_last_row = min(merged.max_row, end_row)
            clipped_first_col = max(merged.min_col, start_col)
            clipped_last_col = min(merged.max_col, end_col)
            if clipped_first_row > clipped_last_row or clipped_first_col > clipped_last_col:
                continue
            anchor = worksheet.cell(merged.min_row, merged.min_col)
            rows[clipped_first_row - start_row][clipped_first_col - start_col] = (
                anchor.value if anchor.value is not None else ""
            )
            cell_styles[clipped_first_row - start_row][clipped_first_col - start_col] = _xlsx_cell_style(anchor)
            merged_regions.append({
                "first_row": clipped_first_row - start_row,
                "last_row": clipped_last_row - start_row,
                "first_column": clipped_first_col - start_col,
                "last_column": clipped_last_col - start_col,
            })

    rows = _rectangularize(rows)
    fragment = rows_to_html(
        rows,
        title=worksheet.title,
        merged_regions=merged_regions,
        add_coordinates=add_coordinates,
        cell_styles=cell_styles,
        column_widths=column_widths,
        row_heights=row_heights,
    )
    width, height = estimate_rows_size(rows)
    workbook.close()
    return TableDocument(
        html=wrap_html(fragment, title=worksheet.title),
        title=worksheet.title,
        source_format="xlsx",
        estimated_width=width,
        estimated_height=height,
    )


def _xlsx_column_width(worksheet: Any, column: int) -> int | None:
    from openpyxl.utils import get_column_letter

    dimension = worksheet.column_dimensions.get(get_column_letter(column))
    if dimension is None or dimension.width is None:
        return None
    return max(1, round(float(dimension.width) * 7 + 5))


def _xlsx_row_height(worksheet: Any, row: int) -> int | None:
    dimension = worksheet.row_dimensions.get(row)
    if dimension is None or dimension.height is None:
        return None
    return max(1, round(float(dimension.height) * 96 / 72))


def _xlsx_color(color: Any) -> str | None:
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        return f"#{str(color.rgb)[-6:]}"
    if color.type == "indexed" and color.indexed is not None:
        from openpyxl.styles.colors import COLOR_INDEX

        index = int(color.indexed)
        if 0 <= index < len(COLOR_INDEX):
            return f"#{COLOR_INDEX[index][-6:]}"
    return None


def _xlsx_cell_style(cell: Any) -> str:
    declarations = []
    fill = cell.fill
    if fill and fill.patternType == "solid":
        color = _xlsx_color(fill.fgColor)
        if color:
            declarations.append(f"background-color: {color}")
    font = cell.font
    color = _xlsx_color(font.color)
    if color:
        declarations.append(f"color: {color}")
    if font.bold:
        declarations.append("font-weight: 700")
    if font.italic:
        declarations.append("font-style: italic")
    if font.sz:
        declarations.append(f"font-size: {float(font.sz):g}pt")
    alignment = cell.alignment
    if alignment.horizontal in {"left", "center", "right", "justify"}:
        declarations.append(f"text-align: {alignment.horizontal}")
    if alignment.vertical in {"top", "center", "bottom"}:
        vertical = "middle" if alignment.vertical == "center" else alignment.vertical
        declarations.append(f"vertical-align: {vertical}")
    if alignment.wrap_text is False:
        declarations.append("white-space: nowrap")
    return "; ".join(declarations)


def document_from_rows(rows: Iterable[Iterable[Any] | dict[str, Any]], *, source_format: str) -> TableDocument:
    normalized = _rows_from_any(rows)
    fragment = rows_to_html(normalized)
    width, height = estimate_rows_size(normalized)
    return TableDocument(
        html=wrap_html(fragment),
        source_format=source_format,
        estimated_width=width,
        estimated_height=height,
    )


def rows_from_delimited(input_path: str | Path, *, delimiter: str) -> list[list[str]]:
    with Path(input_path).open("r", encoding="utf-8-sig", newline="") as f:
        return [row for row in csv.reader(f, delimiter=delimiter)]


def rows_from_markdown(content: str) -> list[list[str]]:
    table_lines = []
    for line in content.splitlines():
        stripped = line.strip()
        if not stripped:
            if table_lines:
                break
            continue
        if "|" in stripped:
            table_lines.append(stripped)
        elif table_lines:
            break

    rows = []
    for line in table_lines:
        cells = _split_markdown_row(line)
        if not cells:
            continue
        if _is_markdown_separator(cells):
            continue
        rows.append(cells)

    if not rows:
        raise ValueError("No markdown table found.")
    return rows


def rows_to_html(
    rows: list[list[Any]],
    *,
    title: str = "",
    merged_regions: list[dict[str, Any]] | None = None,
    add_coordinates: bool = False,
    cell_styles: list[list[str]] | None = None,
    column_widths: list[int | None] | None = None,
    row_heights: list[int | None] | None = None,
) -> str:
    rows = _rectangularize(rows)
    merge_map: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()

    for region in merged_regions or []:
        first_row = _safe_int(region.get("first_row"), 0)
        last_row = _safe_int(region.get("last_row"), first_row)
        first_col = _safe_int(region.get("first_column"), 0)
        last_col = _safe_int(region.get("last_column"), first_col)
        if first_row < 0 or first_col < 0 or first_row >= len(rows):
            continue
        if first_col >= len(rows[first_row]):
            continue
        last_row = min(last_row, len(rows) - 1)
        last_col = min(last_col, len(rows[first_row]) - 1)
        rowspan = max(1, last_row - first_row + 1)
        colspan = max(1, last_col - first_col + 1)
        if rowspan == 1 and colspan == 1:
            continue
        merge_map[(first_row, first_col)] = (rowspan, colspan)
        for row_index in range(first_row, first_row + rowspan):
            for col_index in range(first_col, first_col + colspan):
                if (row_index, col_index) != (first_row, first_col):
                    covered.add((row_index, col_index))

    lines = ["<table>"]
    if column_widths:
        lines.append("<colgroup>")
        for width in column_widths:
            style = f' style="width: {width}px"' if width else ""
            lines.append(f"<col{style}>")
        lines.append("</colgroup>")
    if title:
        lines.append(f"<caption>{html.escape(title)}</caption>")

    for row_index, row in enumerate(rows):
        row_style = ""
        if row_heights and row_index < len(row_heights) and row_heights[row_index]:
            row_style = f' style="height: {row_heights[row_index]}px"'
        lines.append(f"<tr{row_style}>")
        for col_index, cell in enumerate(row):
            if (row_index, col_index) in covered:
                continue
            rowspan, colspan = merge_map.get((row_index, col_index), (1, 1))
            attrs = []
            if rowspan > 1:
                attrs.append(f'rowspan="{rowspan}"')
            if colspan > 1:
                attrs.append(f'colspan="{colspan}"')
            if add_coordinates and (row_index == 0 or col_index == 0):
                attrs.append('class="excel-coord"')
            if cell_styles and row_index < len(cell_styles) and col_index < len(cell_styles[row_index]):
                style = cell_styles[row_index][col_index]
                if style:
                    attrs.append(f'style="{html.escape(style, quote=True)}"')
            attr_text = " " + " ".join(attrs) if attrs else ""
            text = _nbsp_pad(_clean_text(cell))
            lines.append(f"<td{attr_text}>{html.escape(text)}</td>")
        lines.append("</tr>")

    lines.append("</table>")
    return "\n".join(lines)


def wrap_html(fragment: str, *, title: str = "") -> str:
    page_title = html.escape(title or "table2img")
    return f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>{page_title}</title>
  <style>
    html, body {{
      background: #ffffff;
      margin: 0;
      padding: 0;
    }}
    body {{
      color: #111827;
      display: inline-block;
      font-family: "Malgun Gothic", "Apple SD Gothic Neo", "Noto Sans CJK KR",
        "NanumGothic", "Meiryo", "Microsoft JhengHei", "Microsoft YaHei",
        "Nirmala UI", "Leelawadee UI", "Ebrima", "Noto Sans CJK SC",
        "Arial Unicode MS", "Segoe UI", "Noto Sans", system-ui, Arial, sans-serif;
      font-size: 14px;
      line-height: 1.35;
      padding: 24px;
    }}
    table {{
      border-collapse: collapse;
      border-spacing: 0;
      width: max-content;
    }}
    caption {{
      caption-side: top;
      color: #111827;
      font-size: 16px;
      font-weight: 700;
      padding: 0 0 10px;
      text-align: left;
      white-space: normal;
    }}
    td, th {{
      border: 1px solid #1f2937;
      max-width: 420px;
      min-width: 48px;
      padding: 6px 9px;
      text-align: left;
      vertical-align: middle;
      white-space: pre-wrap;
      word-break: normal;
      overflow-wrap: anywhere;
    }}
    th {{
      font-weight: 700;
    }}
    .excel-coord {{
      background-color: #f3f4f6;
      color: #374151;
      font-weight: 700;
      text-align: center;
      border: 1px solid #9ca3af;
      font-size: 12px;
    }}
  </style>
</head>
<body>
{fragment}
</body>
</html>
"""


def render_document(
    document: TableDocument,
    output_path: str | Path,
    *,
    scale: float = DEFAULT_SCALE,
    backend: str = "auto",
    browser_path: str | Path | None = None,
    keep_html: bool = True,
    timeout_seconds: float = 60.0,
    max_viewport_width: int = DEFAULT_MAX_VIEWPORT_WIDTH,
    max_viewport_height: int = DEFAULT_MAX_VIEWPORT_HEIGHT,
) -> RenderResult:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    if keep_html:
        html_path = output.with_suffix(".html")
        html_path.write_text(document.html, encoding="utf-8")
        temp_dir = None
    else:
        temp_dir = tempfile.TemporaryDirectory(prefix="table2img-")
        html_path = Path(temp_dir.name) / "table.html"
        html_path.write_text(document.html, encoding="utf-8")

    width = min(max(320, document.estimated_width), max_viewport_width)
    height = min(max(240, document.estimated_height), max_viewport_height)
    backend = backend.strip().lower()
    if backend not in {"auto", "browser", "pillow"}:
        raise ValueError(f"Unsupported table2img backend: {backend}")

    capture_errors = []
    browser = None
    if backend != "pillow":
        for candidate in find_browsers(browser_path):
            try:
                _capture_with_chrome_cli(
                    candidate,
                    html_path,
                    output,
                    width=width,
                    height=height,
                    scale=scale,
                    timeout_seconds=timeout_seconds,
                )
                browser = candidate
                break
            except RuntimeError as exc:
                capture_errors.append(str(exc))
    if browser is None and backend != "browser":
        _capture_with_pillow(html_path, output, scale=scale)
        browser = Path("pillow")
    if browser is None:
        raise RuntimeError("All browser screenshot attempts failed: " + " | ".join(capture_errors))
    image_width, image_height = trim_image(output)

    result = RenderResult(
        image_path=output,
        html_path=html_path if keep_html else None,
        width=image_width,
        height=image_height,
        browser_path=browser,
    )
    if temp_dir is not None:
        temp_dir.cleanup()
    return result


def find_browser(browser_path: str | Path | None = None) -> Path:
    return find_browsers(browser_path)[0]


def find_browsers(browser_path: str | Path | None = None) -> list[Path]:
    candidates: list[Path] = []
    if browser_path:
        candidates.append(Path(browser_path))

    env_browser = os.environ.get("TABLE2IMG_BROWSER")
    if env_browser:
        candidates.append(Path(env_browser))

    for name in ("msedge", "msedge.exe", "chrome", "chrome.exe", "google-chrome", "chromium", "chromium-browser"):
        found = shutil.which(name)
        if found:
            candidates.append(Path(found))

    program_files = [os.environ.get("ProgramFiles"), os.environ.get("ProgramFiles(x86)")]
    program_roots = [Path(p) for p in program_files if p]
    candidates.extend(root / "Microsoft" / "Edge" / "Application" / "msedge.exe" for root in program_roots)
    candidates.extend(root / "Google" / "Chrome" / "Application" / "chrome.exe" for root in program_roots)

    candidates.extend(
        [
            Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
            Path("/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge"),
            Path("/usr/bin/google-chrome"),
            Path("/usr/bin/chromium"),
            Path("/usr/bin/chromium-browser"),
        ]
    )

    puppeteer_root = Path.home() / ".cache" / "puppeteer" / "chrome"
    if puppeteer_root.exists():
        candidates.extend(puppeteer_root.glob("**/chrome.exe"))
        candidates.extend(puppeteer_root.glob("**/chrome"))

    resolved = []
    seen = set()
    for candidate in candidates:
        if candidate.is_file():
            path = candidate.resolve()
            normalized = str(path).lower()
            if normalized not in seen:
                resolved.append(path)
                seen.add(normalized)
    if resolved:
        return resolved

    raise FileNotFoundError(
        "No Chrome/Chromium/Edge browser found. Set TABLE2IMG_BROWSER or pass --browser."
    )


def estimate_rows_size(rows: list[list[Any]]) -> tuple[int, int]:
    rows = _rectangularize(rows)
    if not rows:
        return 640, 320

    col_count = max((len(row) for row in rows), default=1)
    col_widths = [64] * col_count
    row_heights = []

    for row in rows:
        max_lines = 1
        for col_index, cell in enumerate(row):
            text = _clean_text(cell)
            longest = max((len(part) for part in re.split(r"\s+", text) if part), default=0)
            char_count = min(max(len(text), longest), 48)
            col_widths[col_index] = max(col_widths[col_index], min(420, 28 + char_count * 8))
            max_lines = max(max_lines, max(1, (len(text) // 42) + 1))
        row_heights.append(28 + max_lines * 13)

    # CSS min-width applies to the content box; account for horizontal padding and
    # borders as well so a browser viewport never clips the final columns.
    width = int(sum(max(68, value + 20) for value in col_widths) + 48)
    height = int(sum(row_heights) + 96)
    return width, height


def estimate_html_size(fragment: str) -> tuple[int, int]:
    soup = BeautifulSoup(fragment, "html.parser")
    rows = []
    for tr in soup.find_all("tr"):
        rows.append([cell.get_text(" ", strip=True) for cell in tr.find_all(["td", "th"], recursive=False)])
    if rows:
        return estimate_rows_size(rows)
    text_len = len(_clean_text(soup.get_text(" ", strip=True)))
    return min(16000, max(640, 16 * min(text_len, 600))), 800


def trim_image(image_path: str | Path, *, padding: int = 8) -> tuple[int | None, int | None]:
    try:
        from PIL import Image, ImageChops
    except ImportError:
        return None, None

    path = Path(image_path)
    Image.MAX_IMAGE_PIXELS = None
    with Image.open(path) as img:
        if img.mode != "RGBA":
            img = img.convert("RGBA")
        background = Image.new("RGBA", img.size, (255, 255, 255, 255))
        diff = ImageChops.difference(img, background)
        bbox = diff.getbbox()
        if not bbox:
            return img.size

        left, top, right, bottom = bbox
        left = max(0, left - padding)
        top = max(0, top - padding)
        right = min(img.width, right + padding)
        bottom = min(img.height, bottom + padding)
        cropped = img.crop((left, top, right, bottom))
        cropped.save(path)
        return cropped.size


def _capture_with_chrome_cli(
    browser: Path,
    html_path: Path,
    output_path: Path,
    *,
    width: int,
    height: int,
    scale: float,
    timeout_seconds: float,
) -> None:
    output_path = output_path.resolve()
    output_path.unlink(missing_ok=True)
    html_uri = html_path.resolve().as_uri()
    with tempfile.TemporaryDirectory(prefix="table2img-chrome-") as profile_dir:
        base_args = [
            str(browser),
            f"--user-data-dir={profile_dir}",
            "--disable-gpu",
            "--disable-extensions",
            "--hide-scrollbars",
            "--no-first-run",
            "--no-default-browser-check",
            "--no-sandbox",
            f"--force-device-scale-factor={scale:g}",
            f"--window-size={width},{height}",
            f"--screenshot={output_path}",
            html_uri,
        ]

        last_error = None
        for headless_arg in ("--headless=new", "--headless"):
            args = [base_args[0], headless_arg, *base_args[1:]]
            completed = subprocess.run(
                args,
                text=True,
                capture_output=True,
                timeout=timeout_seconds,
            )
            if _wait_for_screenshot(output_path):
                return
            last_error = completed

    stderr = (last_error.stderr if last_error else "").strip()
    stdout = (last_error.stdout if last_error else "").strip()
    returncode = last_error.returncode if last_error else None
    raise RuntimeError(
        f"Chrome screenshot failed. browser={str(browser)!r} returncode={returncode!r} "
        f"stdout={stdout!r} stderr={stderr!r}"
    )


def _wait_for_screenshot(output_path: Path, grace_seconds: float = 2.0) -> bool:
    deadline = time.monotonic() + grace_seconds
    while time.monotonic() < deadline:
        if output_path.is_file() and output_path.stat().st_size > 0:
            return True
        time.sleep(0.05)
    return output_path.is_file() and output_path.stat().st_size > 0


def _capture_with_pillow(html_path: Path, output_path: Path, *, scale: float) -> None:
    from PIL import Image, ImageDraw, ImageFont

    soup = BeautifulSoup(html_path.read_text(encoding="utf-8"), "html.parser")
    table = soup.find("table")
    if table is None:
        raise RuntimeError("Pillow table rendering failed: HTML contains no table")

    cells = []
    occupied = set()
    max_column = 0
    rows = table.find_all("tr")
    for row_index, row in enumerate(rows):
        column_index = 0
        for cell in row.find_all(["th", "td"], recursive=False):
            while (row_index, column_index) in occupied:
                column_index += 1
            rowspan = max(1, int(cell.get("rowspan", 1)))
            colspan = max(1, int(cell.get("colspan", 1)))
            for covered_row in range(row_index, row_index + rowspan):
                for covered_column in range(column_index, column_index + colspan):
                    occupied.add((covered_row, covered_column))
            classes = set(cell.get("class", []))
            cells.append({
                "row": row_index,
                "column": column_index,
                "rowspan": rowspan,
                "colspan": colspan,
                "text": cell.get_text(" ", strip=True),
                "header": cell.name == "th" or row_index == 0 or "excel-coord" in classes,
                "style": _parse_css_declarations(cell.get("style", "")),
            })
            column_index += colspan
            max_column = max(max_column, column_index)

    if not cells or not rows or max_column == 0:
        raise RuntimeError("Pillow table rendering failed: table is empty")

    factor = max(1.0, float(scale))
    font_size = max(12, round(14 * factor))
    characters = {char for cell in cells for char in cell["text"]}
    regular_font = _load_table_fonts(ImageFont, font_size, bold=False, characters=characters)
    bold_font = _load_table_fonts(ImageFont, font_size, bold=True, characters=characters)
    padding = round(8 * factor)

    column_widths = [round(90 * factor)] * max_column
    col_elements = table.find_all("col")
    for column_index, column in enumerate(col_elements[:max_column]):
        width = _css_pixels(_parse_css_declarations(column.get("style", "")).get("width"))
        if width:
            column_widths[column_index] = round(width * factor)
    for cell in cells:
        if cell["colspan"] != 1:
            continue
        font_faces = bold_font if cell["header"] else regular_font
        text_width = _text_size(cell["text"], font_faces)[0] + padding * 2
        column_widths[cell["column"]] = min(round(360 * factor), max(column_widths[cell["column"]], text_width))

    wrapped = {}
    row_heights = [round(38 * factor)] * len(rows)
    for row_index, row in enumerate(rows):
        height = _css_pixels(_parse_css_declarations(row.get("style", "")).get("height"))
        if height:
            row_heights[row_index] = round(height * factor)
    for index, cell in enumerate(cells):
        available_width = sum(column_widths[cell["column"]:cell["column"] + cell["colspan"]]) - padding * 2
        chars = max(8, int(available_width / max(1, font_size * 0.58)))
        lines = textwrap.wrap(cell["text"], width=chars, break_long_words=True) or [""]
        wrapped[index] = "\n".join(lines)
        if cell["rowspan"] == 1:
            row_heights[cell["row"]] = max(row_heights[cell["row"]], len(lines) * round(font_size * 1.3) + padding * 2)

    x_positions = [0]
    for column_width in column_widths:
        x_positions.append(x_positions[-1] + column_width)
    y_positions = [0]
    for row_height in row_heights:
        y_positions.append(y_positions[-1] + row_height)

    image = Image.new("RGB", (x_positions[-1] + 1, y_positions[-1] + 1), "white")
    draw = ImageDraw.Draw(image)
    for index, cell in enumerate(cells):
        left = x_positions[cell["column"]]
        top = y_positions[cell["row"]]
        right = x_positions[cell["column"] + cell["colspan"]]
        bottom = y_positions[min(len(rows), cell["row"] + cell["rowspan"])]
        style = cell["style"]
        fill = style.get("background-color", "#e5e7eb" if cell["header"] else "white")
        text_fill = style.get("color", "#111827")
        is_bold = cell["header"] or style.get("font-weight") in {"bold", "600", "700", "800", "900"}
        draw.rectangle((left, top, right, bottom), fill=fill, outline="#4b5563", width=max(1, round(factor)))
        text_width, _ = _text_size(wrapped[index], bold_font if is_bold else regular_font)
        text_x = left + padding
        if style.get("text-align") == "center":
            text_x = left + max(padding, (right - left - text_width) / 2)
        elif style.get("text-align") == "right":
            text_x = right - padding - text_width
        _draw_text_with_fallback(
            draw,
            (text_x, top + padding),
            wrapped[index],
            bold_font if is_bold else regular_font,
            fill=text_fill,
            spacing=round(3 * factor),
        )
    image.save(output_path, format="PNG")


def _parse_css_declarations(style: str) -> dict[str, str]:
    declarations = {}
    for item in str(style or "").split(";"):
        name, separator, value = item.partition(":")
        if separator and name.strip() and value.strip():
            declarations[name.strip().lower()] = value.strip()
    return declarations


def _css_pixels(value: str | None) -> float | None:
    match = re.fullmatch(r"([0-9]+(?:\.[0-9]+)?)px", str(value or "").strip())
    return float(match.group(1)) if match else None


def _load_table_fonts(
    image_font,
    size: int,
    *,
    bold: bool,
    characters: set[str] | None = None,
) -> list[_FontFace]:
    candidates = [
        Path("C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/segoeuib.ttf" if bold else "C:/Windows/Fonts/segoeui.ttf"),
        Path("/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf"),
        Path("C:/Windows/Fonts/malgunbd.ttf" if bold else "C:/Windows/Fonts/malgun.ttf"),
        Path("C:/Windows/Fonts/NanumGothicBold.ttf" if bold else "C:/Windows/Fonts/NanumGothic.ttf"),
        Path("C:/Windows/Fonts/meiryob.ttc" if bold else "C:/Windows/Fonts/meiryo.ttc"),
        Path("C:/Windows/Fonts/msjhbd.ttc" if bold else "C:/Windows/Fonts/msjh.ttc"),
        Path("C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/NirmalaB.ttf" if bold else "C:/Windows/Fonts/Nirmala.ttf"),
        Path("C:/Windows/Fonts/leelawdb.ttf" if bold else "C:/Windows/Fonts/leelawad.ttf"),
        Path("C:/Windows/Fonts/ebrimabd.ttf" if bold else "C:/Windows/Fonts/ebrima.ttf"),
        Path("C:/Windows/Fonts/gadugib.ttf" if bold else "C:/Windows/Fonts/gadugi.ttf"),
        Path("C:/Windows/Fonts/seguisb.ttf" if bold else "C:/Windows/Fonts/seguisym.ttf"),
        Path("C:/Windows/Fonts/seguiemj.ttf"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc" if bold else "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf" if bold else "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
        Path("C:/Windows/Fonts/arialuni.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    faces = []
    loaded_paths = set()
    for candidate in candidates:
        if candidate.is_file():
            font = image_font.truetype(str(candidate), size=size)
            faces.append(_FontFace(
                font,
                _font_codepoints(candidate),
                _glyph_signature(font, "\U0010ffff"),
            ))
            loaded_paths.add(candidate.resolve())
    missing = {
        char for char in characters or set()
        if not char.isspace() and not any(_face_has_glyph(face, char) for face in faces)
    }
    if missing:
        for candidate in _system_font_candidates():
            if candidate.resolve() in loaded_paths:
                continue
            try:
                font = image_font.truetype(str(candidate), size=size)
            except OSError:
                continue
            face = _FontFace(
                font,
                _font_codepoints(candidate),
                _glyph_signature(font, "\U0010ffff"),
            )
            supported = {char for char in missing if _face_has_glyph(face, char)}
            if not supported:
                continue
            faces.append(face)
            loaded_paths.add(candidate.resolve())
            missing.difference_update(supported)
            if not missing:
                break
    if faces:
        return faces
    return [_FontFace(image_font.load_default(), None)]


def _font_codepoints(path: Path) -> frozenset[int] | None:
    try:
        from fontTools.ttLib import TTFont

        font = TTFont(str(path), lazy=True, fontNumber=0)
        try:
            codepoints = {
                codepoint
                for table in font["cmap"].tables
                for codepoint in table.cmap
            }
        finally:
            font.close()
        return frozenset(codepoints)
    except Exception:
        return None


def _text_size(text: str, font_faces: list[_FontFace]) -> tuple[int, int]:
    lines = str(text or " ").splitlines() or [" "]
    widths = []
    heights = []
    for line in lines:
        width = 0
        line_height = 0
        for char in line or " ":
            font = _font_for_char(font_faces, char)
            box = font.getbbox(char)
            width += box[2] - box[0]
            line_height = max(line_height, box[3] - box[1])
        widths.append(width)
        heights.append(line_height)
    return max(widths, default=0), sum(heights)


def _draw_text_with_fallback(draw, xy: tuple[int, int], text: str, font_faces: list[_FontFace], *, fill: str, spacing: int) -> None:
    x, y = xy
    line_height = _text_size("Ag", font_faces)[1] + spacing
    for line in str(text).splitlines() or [""]:
        current_x = x
        for char in line:
            font = _font_for_char(font_faces, char)
            draw.text((current_x, y), char, fill=fill, font=font)
            box = font.getbbox(char)
            current_x += box[2] - box[0]
        y += line_height


def _font_for_char(font_faces: list[_FontFace], char: str):
    for face in font_faces:
        if _face_has_glyph(face, char):
            return face.font
    return font_faces[0].font


def _face_has_glyph(face: _FontFace, char: str) -> bool:
    if char.isspace() or ord(char) < 32:
        return True
    if face.codepoints is not None:
        return ord(char) in face.codepoints
    signature = _glyph_signature(face.font, char)
    if signature is None:
        return False
    return face.missing_glyph is None or signature != face.missing_glyph


def _glyph_signature(font: Any, char: str) -> tuple[tuple[int, int], bytes] | None:
    """Fingerprint a glyph so fallback works even when fontTools is unavailable."""
    try:
        mask = font.getmask(char, mode="L")
        return mask.size, bytes(mask)
    except (OSError, ValueError):
        return None


def _system_font_candidates() -> list[Path]:
    windows_root = Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts"
    roots = [
        windows_root,
        Path("/usr/share/fonts"),
        Path.home() / ".fonts",
        Path("/Library/Fonts"),
        Path.home() / "Library/Fonts",
    ]
    candidates = []
    for root in roots:
        if not root.is_dir():
            continue
        try:
            candidates.extend(
                path for path in root.rglob("*")
                if path.suffix.lower() in {".ttf", ".ttc", ".otf"}
            )
        except OSError:
            continue
    return sorted(set(candidates), key=lambda path: str(path).lower())


def _load_table_font(image_font, size: int, *, bold: bool):
    return _load_table_fonts(image_font, size, bold=bold)[0].font


def _detect_format(path: Path, input_format: str) -> str:
    fmt = input_format.lower().strip()
    if fmt != "auto":
        aliases = {
            "markdown": "md",
            "htm": "html",
            "xls": "xlsx",
            "text": "txt",
        }
        return aliases.get(fmt, fmt)

    suffix = path.suffix.lower().lstrip(".")
    aliases = {
        "htm": "html",
        "markdown": "md",
        "xls": "xlsx",
        "txt": "txt",
    }
    return aliases.get(suffix, suffix or "txt")


def _read_jsonl(path: Path, *, json_index: int | None) -> Any:
    target_index = 0 if json_index is None else json_index
    with path.open("r", encoding="utf-8") as f:
        for index, line in enumerate(f):
            if not line.strip():
                continue
            if index == target_index:
                return json.loads(line)
    raise IndexError(f"JSONL index out of range: {target_index}")


def _select_json_record(data: Any, *, json_index: int | None) -> Any:
    if json_index is None:
        if isinstance(data, list) and data and not _looks_like_rows(data):
            for item in data:
                if isinstance(item, dict) and ("tables" in item or "texts" in item):
                    return item
        return data

    if not isinstance(data, list):
        if json_index == 0:
            return data
        raise TypeError("--json-index can only select from top-level JSON lists.")
    return data[_bounded_index(json_index, len(data), "json_index")]


def _looks_like_rows(data: list[Any]) -> bool:
    return all(isinstance(row, (list, tuple, dict)) for row in data) and not any(
        isinstance(row, dict) and ("tables" in row or "texts" in row)
        for row in data
    )


def _rows_from_any(rows: Iterable[Iterable[Any] | dict[str, Any]]) -> list[list[Any]]:
    materialized = list(rows)
    if not materialized:
        return []

    if all(isinstance(row, dict) for row in materialized):
        keys = []
        for row in materialized:
            for key in row.keys():
                if key not in keys:
                    keys.append(key)
        return [keys] + [[row.get(key, "") for key in keys] for row in materialized]

    normalized = []
    for row in materialized:
        if isinstance(row, dict):
            normalized.append([f"{key}: {value}" for key, value in row.items()])
        elif isinstance(row, (list, tuple)):
            normalized.append(list(row))
        else:
            normalized.append([row])
    return _rectangularize(normalized)


def _rectangularize(rows: Iterable[Iterable[Any]]) -> list[list[Any]]:
    normalized = [list(row) for row in rows]
    width = max((len(row) for row in normalized), default=0)
    return [row + [""] * (width - len(row)) for row in normalized]


def _split_markdown_row(line: str) -> list[str]:
    stripped = line.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    return [_clean_text(cell.replace("\\|", "|")) for cell in stripped.split("|")]


def _is_markdown_separator(cells: list[str]) -> bool:
    return all(re.fullmatch(r":?-{3,}:?", cell.strip()) for cell in cells if cell.strip())


def _clean_text(value: Any) -> str:
    return " ".join(str(value if value is not None else "").replace("\xa0", " ").split())


def _nbsp_pad(value: str) -> str:
    return f"\xa0{value}\xa0" if value else "\xa0"


def _safe_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _bounded_index(index: int, length: int, name: str) -> int:
    if index < 0 or index >= length:
        raise IndexError(f"{name} out of range: {index}; available range is 0..{length - 1}")
    return index


def positive_float(value: str) -> float:
    parsed = float(value)
    if parsed <= 0:
        raise argparse.ArgumentTypeError("value must be positive")
    return parsed
