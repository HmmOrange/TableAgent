import json
import unittest
from pathlib import Path
from subprocess import CompletedProcess
from tempfile import TemporaryDirectory
from unittest.mock import patch

from table2img.core import (
    _FontFace,
    _capture_with_chrome_cli,
    _font_for_char,
    _glyph_signature,
    document_from_json,
    document_from_xlsx,
    find_browsers,
    render_document,
    rows_from_markdown,
)


class Table2ImgCoreTests(unittest.TestCase):
    def test_font_fallback_without_fonttools_skips_missing_glyph(self):
        class Mask:
            size = (1, 1)

            def __init__(self, value):
                self.value = value

            def __bytes__(self):
                return self.value

        class Font:
            def __init__(self, supported):
                self.supported = supported

            def getmask(self, char, mode="L"):
                return Mask(char.encode("utf-8") if char in self.supported else b"missing")

        latin = Font({"A"})
        korean = Font({"한"})
        faces = [
            _FontFace(latin, None, _glyph_signature(latin, "\U0010ffff")),
            _FontFace(korean, None, _glyph_signature(korean, "\U0010ffff")),
        ]

        self.assertIs(_font_for_char(faces, "한"), korean)

    def test_pillow_backend_renders_without_launching_browser(self):
        with TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "table.png"
            document = document_from_json({"texts": [["Year", "Revenue"], ["2024", "100"]]})

            with patch("table2img.core._capture_with_chrome_cli") as browser_capture:
                result = render_document(document, output_path, backend="pillow")

            browser_capture.assert_not_called()
            self.assertTrue(output_path.is_file())
            self.assertGreater(output_path.stat().st_size, 0)
            self.assertEqual(result.browser_path, Path("pillow"))

    def test_browser_discovery_deduplicates_candidates(self):
        browser = Path("browser.exe").resolve()
        with patch("table2img.core.shutil.which", return_value=str(browser)), patch.object(Path, "is_file", return_value=True):
            browsers = find_browsers()

        self.assertEqual(browsers.count(browser), 1)

    def test_chrome_capture_uses_isolated_profile_and_accepts_written_screenshot(self):
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            html_path = root / "table.html"
            output_path = root / "table.png"
            html_path.write_text("<html><body>table</body></html>", encoding="utf-8")
            output_path.write_bytes(b"stale")

            def fake_run(args, **kwargs):
                self.assertTrue(any(arg.startswith("--user-data-dir=") for arg in args))
                self.assertFalse(output_path.exists())
                output_path.write_bytes(b"png")
                return CompletedProcess(args, returncode=0, stdout="", stderr="")

            with patch("table2img.core.subprocess.run", side_effect=fake_run):
                _capture_with_chrome_cli(
                    Path("chrome.exe"),
                    html_path,
                    output_path,
                    width=800,
                    height=600,
                    scale=2.0,
                    timeout_seconds=10,
                )

            self.assertEqual(output_path.read_bytes(), b"png")

    def test_markdown_rows(self):
        rows = rows_from_markdown(
            """
            | A | B |
            |---|---|
            | 1 | 2 |
            """
        )
        self.assertEqual(rows, [["A", "B"], ["1", "2"]])

    def test_hitab_merged_region_becomes_table_span(self):
        document = document_from_json(
            {
                "title": "sample",
                "texts": [["Region", "", "2018"], ["Allowance", "100", "90"]],
                "merged_regions": [
                    {
                        "first_row": 0,
                        "last_row": 0,
                        "first_column": 0,
                        "last_column": 1,
                    }
                ],
            }
        )

        self.assertIn('colspan="2"', document.html)
        self.assertIn("Region", document.html)
        self.assertIn("2018", document.html)
        self.assertIn("90", document.html)

    def test_mulhi_json_table_html(self):
        payload = {"tables": ["<table><tr><td>A</td><td>B</td></tr></table>"]}
        document = document_from_json(payload)
        self.assertIn("<table>", document.html)
        self.assertIn("<td>A</td>", document.html)

    def test_json_rows_from_dicts(self):
        document = document_from_json(json.loads('[{"name":"A","score":1},{"name":"B","score":2}]'))
        self.assertIn("name", document.html)
        self.assertIn("score", document.html)
        self.assertIn("B", document.html)

    def test_document_from_xlsx_coordinates(self):
        import openpyxl
        from tempfile import TemporaryDirectory
        from pathlib import Path
        from table2img.core import document_from_xlsx

        with TemporaryDirectory() as tmpdir:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "SheetTest"
            ws["B2"] = "ValB2"
            ws["C3"] = "ValC3"
            
            xlsx_path = Path(tmpdir) / "test.xlsx"
            wb.save(xlsx_path)
            
            doc = document_from_xlsx(xlsx_path, add_coordinates=True)
            
            self.assertIn("B", doc.html)
            self.assertIn("C", doc.html)
            self.assertIn("2", doc.html)
            self.assertIn("3", doc.html)
            self.assertIn('class="excel-coord"', doc.html)
            self.assertIn("ValB2", doc.html)
            self.assertIn("ValC3", doc.html)

    def test_document_from_xlsx_preserves_style_merge_dimensions_and_unicode(self):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        with TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "styled.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet["A1"] = "Tiếng Việt / 한국어"
            sheet["A1"].fill = PatternFill("solid", fgColor="FF00CC66")
            sheet["A1"].font = Font(bold=True, color="FFFF0000")
            sheet["A1"].alignment = Alignment(horizontal="center")
            sheet.merge_cells("A1:B1")
            sheet.column_dimensions["A"].width = 24
            sheet.row_dimensions[1].height = 30
            workbook.save(path)

            document = document_from_xlsx(path, add_coordinates=False)
            clipped = document_from_xlsx(path, add_coordinates=False, cell_range="B1:B1")

            self.assertIn('colspan="2"', document.html)
            self.assertIn("Tiếng Việt / 한국어", document.html)
            self.assertIn("background-color: #00CC66", document.html)
            self.assertIn("color: #FF0000", document.html)
            self.assertIn("font-weight: 700", document.html)
            self.assertIn("text-align: center", document.html)
            self.assertIn('style="width: 173px"', document.html)
            self.assertIn('style="height: 40px"', document.html)
            self.assertIn("Tiếng Việt / 한국어", clipped.html)

            output = Path(tmpdir) / "styled.png"
            render_document(document, output, backend="pillow", scale=1)
            from PIL import Image

            with Image.open(output) as image:
                colors = {color for _, color in image.getcolors(maxcolors=image.width * image.height)}
            self.assertTrue({(0, 204, 102), (0, 204, 102, 255)} & colors)


if __name__ == "__main__":
    unittest.main()
