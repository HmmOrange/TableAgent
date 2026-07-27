from __future__ import annotations

import re
import threading
import time
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import yaml

from TableAgent.configs import load_config
from TableAgent.pipeline import TableAgentPipeline
from TableAgent.pipeline.base import PipelineOutput
from TableAgent.llm import LLMResponse
from service.runtime import TableAgentService


class FakePipeline:
    instances = []

    def __init__(self, llm_client, layout_vlm_client, config):
        self.llm_client = llm_client
        self.layout_vlm_client = layout_vlm_client
        self.config = config
        self.prepared = []
        self.runs = []
        self.forces = []
        type(self).instances.append(self)

    def verify_samples(self, samples, force=False):
        self.forces.append(force)
        workbook_path = Path(samples[0].table_path.split(";")[0])
        workbook = openpyxl.load_workbook(workbook_path, read_only=True)
        try:
            sheet_names = list(workbook.sheetnames)
        finally:
            workbook.close()
        selected = samples[0].raw.get("selected_sheets") or sheet_names
        records = []
        for sheet_name in selected:
            structure_path = (
                Path(self.config["source_artifact_dir"])
                / "fake"
                / sheet_name
                / "structure.yaml"
            )
            structure_path.parent.mkdir(parents=True, exist_ok=True)
            structure_path.write_text(
                f"table1:\n  name: {sheet_name}\n  headers: []\n",
                encoding="utf-8",
            )
            records.append(
                SimpleNamespace(
                    workbook_path=workbook_path,
                    sheet_name=sheet_name,
                    structure_path=structure_path,
                    status="good",
                    cache_hit=False,
                )
            )
        return records

    def prepare_samples(self, samples):
        self.prepared.extend(samples)

    def run(self, sample):
        self.runs.append(sample)
        workbook_path = sample.table_path.split(";")[0]
        return PipelineOutput(
            sample_id=sample.sample_id,
            structured_table="table1:\n  headers: []\n",
            predicted_answer=f"answer: {sample.question}",
            latency=0.25,
            token_usage={"prompt": 4, "completion": 2},
            metadata={
                "workbook_path": workbook_path,
                "workbook_sheets": ["Sheet"],
                "verification": {"status": "good"},
                "qa": {"success": True, "artifacts": {"private": "path"}},
            },
        )


class FakeSummaryClient:
    def generate(self, prompt, system_prompt=None):
        description = "Workbook summary" if "workbook as a whole" in prompt else "Sheet summary"
        return LLMResponse(content=f'{{"description": "{description}"}}')


class FakeIndexedPipeline:
    instances = []
    calls = []

    def __init__(self, llm_client, layout_vlm_client, config):
        self.llm_client = llm_client
        self.layout_vlm_client = layout_vlm_client
        self.config = config
        self.qa_agent = SimpleNamespace(
            run=lambda prompt: LLMResponse(content="combined indexed answer", prompt_tokens=2, completion_tokens=1)
        )
        type(self).instances.append(self)
        self.progress_callback = None

    def set_progress_callback(self, callback):
        self.progress_callback = callback

    def _run_verified_qa(self, **kwargs):
        type(self).calls.append(kwargs)
        assert Path(kwargs["structure_path"]).is_file()
        if self.progress_callback is not None:
            self.progress_callback("[qa] planning start")
            self.progress_callback("[qa] planning done")
        return (
            LLMResponse(content="indexed answer", prompt_tokens=3, completion_tokens=2),
            {"success": True, "fallback_used": False, "replan_count": 0},
        )


def _workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    workbook.active["A1"] = "value"
    workbook.save(path)
    workbook.close()
    return path


def _multi_sheet_workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    workbook.active.title = "Summary"
    workbook.create_sheet("Detail")
    workbook.create_sheet("Archive")
    workbook.save(path)
    workbook.close()
    return path


def test_service_runs_structure_once_and_answers_all_queries(tmp_path: Path):
    FakePipeline.instances = []
    answer_client = FakeSummaryClient()
    layout_client = object()
    source = _workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=answer_client,
        layout_vlm_client=layout_client,
        pipeline_factory=FakePipeline,
    )

    result = service.run(
        stage="all",
        workbooks=[source],
        queries=["first question", "second question"],
        job_id="job-one",
    )

    assert len(FakePipeline.instances) == 2
    assert FakePipeline.instances[0].layout_vlm_client is layout_client
    assert FakePipeline.instances[1].llm_client is answer_client
    assert len(FakePipeline.instances[1].runs) == 2
    assert result["workbooks"] == ["book.xlsx"]
    assert result["structures"][0]["artifact"].endswith(".yaml")
    assert result["schema_artifacts"] == [
        {"workbook": "book.xlsx", "artifact": "workbooks/book.xlsx/schema.yaml"}
    ]
    assert result["metadata_artifacts"] == [
        {"workbook": "book.xlsx", "artifact": "workbooks/book.xlsx/metadata.json"}
    ]
    assert [item["answer"] for item in result["answers"]] == [
        "answer: first question",
        "answer: second question",
    ]
    assert "artifacts" not in result["answers"][0]["qa"]
    assert (service.root_dir / "job-one" / "run.json").is_file()
    assert not (service.root_dir / "jobs").exists()
    assert not (service.root_dir / "inputs").exists()
    assert not (service.root_dir / "structure").exists()


def test_service_accepts_persisted_not_good_structure_artifacts(tmp_path: Path):
    class NotGoodPipeline(FakePipeline):
        def verify_samples(self, samples, force=False):
            records = super().verify_samples(samples, force=force)
            for record in records:
                record.status = "not_good"
            return records

    source = _workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=NotGoodPipeline,
    )

    result = service.run(
        stage="structure",
        workbooks=[source],
        queries=[],
        job_id="not-good-structure",
    )

    structure = result["structures"][0]
    assert structure["status"] == "good"
    assert structure["verification_status"] == "not_good"
    assert structure["structure"]
    assert structure["artifact"]
    assert (service.root_dir / "not-good-structure" / structure["artifact"]).is_file()
    assert (service.root_dir / "not-good-structure" / "run.json").is_file()


def test_service_runs_structure_and_qa_units_concurrently(tmp_path: Path):
    state = {
        "structure_active": 0,
        "structure_max": 0,
        "qa_active": 0,
        "qa_max": 0,
    }
    state_lock = threading.Lock()
    instances = []

    def enter(phase: str) -> None:
        with state_lock:
            active_key = f"{phase}_active"
            max_key = f"{phase}_max"
            state[active_key] += 1
            state[max_key] = max(state[max_key], state[active_key])

    def leave(phase: str) -> None:
        with state_lock:
            state[f"{phase}_active"] -= 1

    class ConcurrentPipeline:
        def __init__(self, llm_client, layout_vlm_client, config):
            self.config = config
            instances.append(self)

        def verify_samples(self, samples, force=False):
            sample = samples[0]
            workbook_path = Path(sample.table_path)
            sheet_name = sample.raw["selected_sheets"][0]
            enter("structure")
            try:
                time.sleep(0.08)
                structure_dir = (
                    Path(self.config["source_artifact_dir"])
                    / workbook_path.stem
                    / sheet_name
                )
                structure_dir.mkdir(parents=True, exist_ok=True)
                structure_path = structure_dir / "structure.yaml"
                structure_path.write_text(
                    f"table1:\n  name: {workbook_path.stem}\n  headers: []\n",
                    encoding="utf-8",
                )
                return [
                    SimpleNamespace(
                        workbook_path=workbook_path,
                        sheet_name=sheet_name,
                        structure_path=structure_path,
                        status="good",
                        cache_hit=False,
                    )
                ]
            finally:
                leave("structure")

        def prepare_samples(self, samples):
            pass

        def run(self, sample):
            enter("qa")
            try:
                time.sleep(0.08)
                return PipelineOutput(
                    sample_id=sample.sample_id,
                    structured_table="table1: {}\n",
                    predicted_answer=f"answer: {sample.question}",
                    latency=0.08,
                    token_usage={"prompt": 1, "completion": 1},
                    metadata={
                        "workbook_path": sample.table_path.split(";")[0],
                        "workbook_sheets": ["Sheet"],
                        "verification": {"status": "good"},
                        "qa": {"success": True},
                    },
                )
            finally:
                leave("qa")

    source = _multi_sheet_workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=ConcurrentPipeline,
    )

    result = service.run(
        stage="all",
        workbooks=[source],
        queries=["first question", "second question"],
        max_workers=2,
        persist=False,
    )

    assert state["structure_max"] == 2
    assert state["qa_max"] == 2
    assert len(instances) == 5
    assert all(instance.config["max_workers"] == 2 for instance in instances)
    assert [(item["workbook"], item["sheet"]) for item in result["structures"]] == [
        ("book.xlsx", "Summary"),
        ("book.xlsx", "Detail"),
        ("book.xlsx", "Archive"),
    ]
    assert [item["answer"] for item in result["answers"]] == [
        "answer: first question",
        "answer: second question",
    ]


def test_structure_stage_always_generates_schema_and_metadata(tmp_path: Path):
    FakePipeline.instances = []
    source = _workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=FakePipeline,
    )

    result = service.run(
        stage="structure",
        workbooks=[source],
        job_id="structure-artifacts",
    )

    assert len(FakePipeline.instances) == 1
    assert result["schema_artifacts"][0]["artifact"] == "workbooks/book.xlsx/schema.yaml"
    assert result["metadata_artifacts"][0]["artifact"] == "workbooks/book.xlsx/metadata.json"
    assert result["retrieval_records"]
    assert result["retrieval_records"][0]["schema_yaml"]
    assert all("structure_yaml" not in record for record in result["retrieval_records"])
    assert all("structure_yamls" not in record for record in result["retrieval_records"])
    schema = yaml.safe_load(result["retrieval_records"][0]["schema_yaml"])
    assert schema["Sheet"]["structure"]["table1"]["name"] == "Sheet"


def test_indexed_qa_uses_persisted_structures_without_layout_extraction(tmp_path: Path):
    FakeIndexedPipeline.instances = []
    FakeIndexedPipeline.calls = []
    source = _workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=FakeIndexedPipeline,
    )

    result = service.run_indexed_qa(
        query="question",
        workbooks=[source],
        qa_max_replans=2,
        artifacts=[
            {
                "id": "book:Sheet:table-1",
                "upload_name": "book.xlsx",
                "document_name": "book.xlsx",
                "score": 0.9,
                "sheet": "Sheet",
                "retrieval_level": "table",
                "retrieval_card": "Workbook: book.xlsx\nSheet: Sheet",
                "structure_yaml": "table1:\n  name: Sheet\n  sheet: Sheet\n  headers: []\n",
            }
        ],
    )

    assert len(FakeIndexedPipeline.instances) == 1
    assert FakeIndexedPipeline.instances[0].layout_vlm_client is None
    assert FakeIndexedPipeline.instances[0].config["qa_max_replans"] == 2
    assert "indexed_schema_text" not in FakeIndexedPipeline.calls[0]
    assert result["answers"][0]["answer"] == "indexed answer"
    assert result["answers"][0]["retrieval"]["mode"] == "indexed_vector_multi_candidate"
    workspace_path = Path(result["workspace_path"])
    assert workspace_path.is_dir()
    assert result["answers"][0]["qa"]["workspace_retained"] is True
    assert Path(result["answers"][0]["qa"]["workspace_path"]) == workspace_path
    assert (workspace_path / "input" / "001" / "book.xlsx").is_file()


def test_production_indexed_qa_removes_temporary_workspace(tmp_path: Path, monkeypatch):
    FakeIndexedPipeline.instances = []
    FakeIndexedPipeline.calls = []
    source = _workbook(tmp_path / "book.xlsx")
    monkeypatch.setenv("TABLE_AGENT_QA_RETAIN_WORKSPACES", "false")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=FakeIndexedPipeline,
    )

    result = service.run_indexed_qa(
        query="question",
        workbooks=[source],
        artifacts=[
            {
                "id": "book:Sheet:table-1",
                "upload_name": "book.xlsx",
                "sheet": "Sheet",
                "retrieval_card": "Workbook: book.xlsx\nSheet: Sheet",
                "structure_yaml": "table1:\n  name: Sheet\n  sheet: Sheet\n  headers: []\n",
            }
        ],
    )

    assert result["workspace_path"] is None
    assert result["answers"][0]["qa"]["workspace_retained"] is False
    assert result["answers"][0]["qa"]["workspace_path"] is None


def test_indexed_qa_forwards_progress_without_changing_answer(tmp_path: Path):
    FakeIndexedPipeline.instances = []
    FakeIndexedPipeline.calls = []
    source = _workbook(tmp_path / "book.xlsx")
    progress_messages: list[str] = []
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=FakeIndexedPipeline,
    )

    result = service.run_indexed_qa(
        query="question",
        workbooks=[source],
        progress_callback=progress_messages.append,
        artifacts=[
            {
                "id": "book:Sheet:table-1",
                "upload_name": "book.xlsx",
                "sheet": "Sheet",
                "retrieval_card": "Workbook metadata",
                "structure_yaml": "table1:\n  name: Sheet\n  sheet: Sheet\n  headers: []\n",
            }
        ],
    )

    assert result["answers"][0]["answer"] == "indexed answer"
    assert progress_messages == ["[qa] planning start", "[qa] planning done"]


def test_instant_indexed_qa_disables_thinking_and_uses_instant_limits(tmp_path: Path):
    FakeIndexedPipeline.instances = []
    FakeIndexedPipeline.calls = []
    source = _workbook(tmp_path / "book.xlsx")
    answer_client = SimpleNamespace(
        extra_body={"chat_template_kwargs": {"enable_thinking": True}},
        max_tokens=8192,
    )
    service = TableAgentService(
        {
            "service": {"root_dir": str(tmp_path / "service")},
            "table_agent": {
                "qa_max_retries": 3,
                "generation_max_tokens": 8192,
                "qa_instant_max_retries": 1,
                "qa_instant_generation_max_tokens": 2048,
            },
        },
        llm_client=answer_client,
        layout_vlm_client=object(),
        pipeline_factory=FakeIndexedPipeline,
    )

    service.run_indexed_qa(
        query="question",
        workbooks=[source],
        mode="instant",
        artifacts=[
            {
                "id": "book:Sheet:table-1",
                "upload_name": "book.xlsx",
                "sheet": "Sheet",
                "retrieval_card": "Workbook: book.xlsx\nSheet: Sheet",
                "structure_yaml": "table1:\n  name: Sheet\n  sheet: Sheet\n  headers: []\n",
            }
        ],
    )

    pipeline = FakeIndexedPipeline.instances[0]
    assert pipeline.llm_client is not answer_client
    assert pipeline.llm_client.extra_body["chat_template_kwargs"]["enable_thinking"] is False
    assert answer_client.extra_body["chat_template_kwargs"]["enable_thinking"] is True
    assert pipeline.config["qa_max_retries"] == 1
    assert pipeline.config["generation_max_tokens"] == 2048


def test_thinking_indexed_qa_disables_reasoning_and_preserves_limits(tmp_path: Path):
    FakeIndexedPipeline.instances = []
    FakeIndexedPipeline.calls = []
    source = _workbook(tmp_path / "book.xlsx")
    answer_client = SimpleNamespace(
        extra_body={"chat_template_kwargs": {"enable_thinking": True}},
        max_tokens=8192,
    )
    service = TableAgentService(
        {
            "service": {"root_dir": str(tmp_path / "service")},
            "table_agent": {
                "qa_max_retries": 3,
                "generation_max_tokens": 8192,
            },
        },
        llm_client=answer_client,
        layout_vlm_client=object(),
        pipeline_factory=FakeIndexedPipeline,
    )

    service.run_indexed_qa(
        query="question",
        workbooks=[source],
        mode="thinking",
        artifacts=[
            {
                "id": "book:Sheet:table-1",
                "upload_name": "book.xlsx",
                "sheet": "Sheet",
                "retrieval_card": "Workbook: book.xlsx\nSheet: Sheet",
                "structure_yaml": "table1:\n  name: Sheet\n  sheet: Sheet\n  headers: []\n",
            }
        ],
    )

    pipeline = FakeIndexedPipeline.instances[0]
    assert pipeline.llm_client is not answer_client
    assert pipeline.llm_client.extra_body["chat_template_kwargs"]["enable_thinking"] is False
    assert answer_client.extra_body["chat_template_kwargs"]["enable_thinking"] is True
    assert pipeline.config["qa_max_retries"] == 3
    assert pipeline.config["generation_max_tokens"] == 8192


def test_real_indexed_qa_hybrid_routes_only_the_matching_workbook(tmp_path: Path, monkeypatch):
    sales = _workbook(tmp_path / "sales.xlsx")
    sales_book = openpyxl.load_workbook(sales)
    sales_book.create_sheet("Archive")
    sales_book.save(sales)
    sales_book.close()
    maintenance = _workbook(tmp_path / "maintenance.xlsx")
    maintenance_book = openpyxl.load_workbook(maintenance)
    maintenance_book.active["B1"] = "maintenance"
    maintenance_book.save(maintenance)
    maintenance_book.close()
    config = load_config("config.example.yaml")
    config["service"]["root_dir"] = str(tmp_path / "service")
    config["table_agent"]["retrieval_rerank_with_llm"] = False
    config["table_agent"]["retrieval_embedding_provider"] = "mock"
    qa_calls = []

    def run_verified_qa(_pipeline, **kwargs):
        qa_calls.append(
            {
                **kwargs,
                "structure_text": kwargs["structure_path"].read_text(
                    encoding="utf-8"
                ),
            }
        )
        return (
            LLMResponse(content="Sales answer", prompt_tokens=3, completion_tokens=2),
            {"success": True, "fallback_used": False, "replan_count": 0},
        )

    monkeypatch.setattr(TableAgentPipeline, "_run_verified_qa", run_verified_qa)
    service = TableAgentService(
        config,
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
    )

    selection = service.select_indexed_artifact(
        query="regional revenue score",
        mode="instant",
        artifacts=[
            {
                "id": "sales:summary",
                "document_id": "doc-sales",
                "upload_name": "sales.xlsx",
                "sheet": "Sheet",
                "retrieval_type": "data",
                "retrieval_level": "table",
                "retrieval_card": "Regional revenue score and quarterly sales results",
                "structure_yaml": "table1:\n  sheet: Sheet\n  headers: []\n",
            },
            {
                "id": "maintenance:plan",
                "document_id": "doc-maintenance",
                "upload_name": "maintenance.xlsx",
                "sheet": "Sheet",
                "retrieval_type": "data",
                "retrieval_level": "table",
                "retrieval_card": "Equipment maintenance schedule and spare parts",
                "structure_yaml": "table1:\n  sheet: Sheet\n  headers: []\n",
            },
            {
                "id": "sales:archive",
                "document_id": "doc-sales",
                "upload_name": "sales.xlsx",
                "sheet": "Archive",
                "retrieval_type": "data",
                "retrieval_level": "table",
                "retrieval_card": "Historical discontinued products",
                "structure_yaml": "table1:\n  sheet: Archive\n  headers: []\n",
            },
        ],
    )
    assert selection["document_id"] == "doc-sales"
    assert selection["retrieval"]["candidate_count"] == 3

    result = service.run_indexed_qa(
        query="regional revenue score",
        workbooks=[sales, maintenance],
        qa_enable_final_review=False,
        artifacts=[
            {
                "id": "sales:summary",
                "document_id": "doc-sales",
                "upload_name": "sales.xlsx",
                "sheet": "Sheet",
                "retrieval_type": "data",
                "retrieval_level": "table",
                "retrieval_card": "Regional revenue score and quarterly sales results",
                "structure_yaml": "table1:\n  sheet: Sheet\n  headers: []\n",
            },
            {
                "id": "maintenance:plan",
                "document_id": "doc-maintenance",
                "upload_name": "maintenance.xlsx",
                "sheet": "Sheet",
                "retrieval_type": "data",
                "retrieval_level": "table",
                "retrieval_card": "Equipment maintenance schedule and spare parts",
                "structure_yaml": "table1:\n  sheet: Sheet\n  headers: []\n",
            },
            {
                "id": "sales:archive",
                "document_id": "doc-sales",
                "upload_name": "sales.xlsx",
                "sheet": "Archive",
                "retrieval_type": "data",
                "retrieval_level": "table",
                "retrieval_card": "Historical discontinued products",
                "structure_yaml": "table1:\n  sheet: Archive\n  headers: []\n",
            },
        ],
    )

    answer = result["answers"][0]
    assert len(qa_calls) == 1
    assert qa_calls[0]["workbook_path"].name == "sales.xlsx"
    assert qa_calls[0]["structure_text"] == (
        "table1:\n  sheet: Sheet\n  headers: []\n"
    )
    assert qa_calls[0]["related_structure_paths"] == []
    assert qa_calls[0]["enable_final_answer_review"] is False
    assert answer["workbook"] == "sales.xlsx"
    assert answer["workbooks"] == ["sales.xlsx"]
    assert answer["retrieval"]["mode"] == "table_agent_hybrid"
    assert answer["retrieval"]["document_id"] == "doc-sales"
    assert answer["retrieval"]["embedding_used"] is True
    assert answer["retrieval"]["candidate_count"] == 3
    assert answer["retrieval"]["workbook_count"] == 2
    assert sum(bool(row["selected"]) for row in answer["retrieval"]["audit"]) == 1


def test_instant_indexed_retrieval_rejects_unrelated_candidates(tmp_path: Path):
    config = load_config("config.example.yaml")
    config["service"]["root_dir"] = str(tmp_path / "service")
    config["table_agent"]["retrieval_rerank_with_llm"] = True
    config["table_agent"]["retrieval_embedding_provider"] = "mock"
    service = TableAgentService(
        config,
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
    )

    selection = service.select_indexed_artifact(
        query="name in the CV",
        mode="instant",
        artifacts=[
            {
                "id": "maintenance:plan",
                "document_id": "doc-maintenance",
                "upload_name": "maintenance.xlsx",
                "sheet": "Plan",
                "retrieval_type": "data",
                "retrieval_level": "table",
                "retrieval_card": "Equipment maintenance schedule and spare parts",
                "structure_yaml": "table1:\n  sheet: Plan\n  headers: []\n",
            }
        ],
    )

    assert selection["status"] == "no_evidence"
    assert selection["document_id"] is None
    assert selection["retrieval"]["selected_artifact_id"] is None
    assert selection["retrieval"]["audit"][0]["selected"] is False


def test_indexed_retrieval_without_candidate_returns_no_evidence(tmp_path: Path):
    config = load_config("config.example.yaml")
    config["service"]["root_dir"] = str(tmp_path / "service")

    class EmptyRetriever:
        def select_indexed(self, **_kwargs):
            return None

    class EmptyPipeline:
        source_retriever = EmptyRetriever()

        @staticmethod
        def _fit_context(value):
            return value

    service = TableAgentService(
        config,
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=lambda **_kwargs: EmptyPipeline(),
    )

    selection = service.select_indexed_artifact(
        query="unmatched question",
        mode="instant",
        artifacts=[
            {
                "id": "maintenance:plan",
                "document_id": "doc-maintenance",
                "upload_name": "maintenance.xlsx",
                "sheet": "Plan",
                "retrieval_type": "data",
                "retrieval_level": "table",
                "retrieval_card": "Maintenance schedule",
                "structure_yaml": "table1:\n  sheet: Plan\n  headers: []\n",
            }
        ],
    )

    assert selection["status"] == "no_evidence"
    assert selection["selected_artifact_id"] is None
    assert selection["retrieval"]["candidate_count"] == 1
    assert selection["retrieval"]["rejection_reason"] == "no_usable_candidate"


def test_thinking_indexed_retrieval_uses_strong_candidate_fallback(tmp_path: Path):
    candidate = SimpleNamespace(
        artifact_id="aoi:standards",
        workbook_path=Path("aoi.xlsx"),
        sheet_name="AOI Standards",
        table_id="aoi_standards",
        table_name="AOI Standards",
        matched_terms=["aoi", "inspection", "tool"],
        missing_terms=["pin", "hole"],
        lexical_score=3.0,
        embedding_used=True,
        embedding_score=0.7,
        retrieval_audit=[{"artifact_id": "aoi:standards", "rank": 1}],
        retrieval_trace=[{"status": "need_more", "query_type": "data"}],
    )

    class NeedMorePipeline:
        def __init__(self, llm_client, layout_vlm_client, config):
            del llm_client, layout_vlm_client, config
            self.source_retriever = SimpleNamespace(
                select_indexed=lambda **kwargs: candidate
            )
            self._fit_context = lambda value: value

    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=NeedMorePipeline,
    )

    selection = service.select_indexed_artifact(
        query="AOI inspection tool for pin hole",
        mode="thinking",
        artifacts=[
            {
                "id": "aoi:standards",
                "document_id": "doc-aoi",
                "upload_name": "aoi.xlsx",
                "sheet": "AOI Standards",
                "structure_yaml": "table1:\n  headers: []\n",
            }
        ],
    )

    assert selection["status"] == "selected"
    assert selection["document_id"] == "doc-aoi"
    assert selection["retrieval"]["selection_fallback"] == (
        "deterministic_relevance"
    )
    assert selection["retrieval"]["audit"][0]["selected"] is True


def test_thinking_indexed_retrieval_respects_disabled_llm_reranker(tmp_path: Path):
    captured_config = {}
    candidate = SimpleNamespace(
        artifact_id="aoi:standards",
        workbook_path=Path("aoi.xlsx"),
        sheet_name="AOI Standards",
        table_id="aoi_standards",
        table_name="AOI Standards",
        matched_terms=["aoi", "inspection"],
        missing_terms=[],
        lexical_score=2.0,
        embedding_used=False,
        embedding_score=0.0,
        retrieval_audit=[{"artifact_id": "aoi:standards", "rank": 1}],
        retrieval_trace=[{"query_type": "data"}],
    )

    class ConfigCapturePipeline:
        def __init__(self, llm_client, layout_vlm_client, config):
            del llm_client, layout_vlm_client
            captured_config.update(config)
            self.source_retriever = SimpleNamespace(
                select_indexed=lambda **kwargs: candidate
            )
            self._fit_context = lambda value: value

    service = TableAgentService(
        {
            "service": {"root_dir": str(tmp_path / "service")},
            "table_agent": {"retrieval_rerank_with_llm": False},
        },
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=ConfigCapturePipeline,
    )

    selection = service.select_indexed_artifact(
        query="AOI inspection standards",
        mode="thinking",
        artifacts=[
            {
                "id": "aoi:standards",
                "document_id": "doc-aoi",
                "upload_name": "aoi.xlsx",
                "sheet": "AOI Standards",
                "structure_yaml": "table1:\n  headers: []\n",
            }
        ],
    )

    assert selection["status"] == "selected"
    assert captured_config["retrieval_rerank_with_llm"] is False


def test_thinking_indexed_retrieval_rejects_weak_candidate_when_reranker_needs_more():
    candidate = SimpleNamespace(
        matched_terms=[],
        missing_terms=["pin", "hole"],
        lexical_score=0.0,
        embedding_used=True,
        embedding_score=0.2,
    )

    rejection = TableAgentService._indexed_rejection_reason(
        candidate,
        {"reranker": {"status": "need_more"}},
        mode="thinking",
    )

    assert rejection == (
        "The TableAgent reranker found no sufficiently relevant indexed table."
    )


def test_qa_stage_generates_fresh_structure_before_answering(tmp_path: Path):
    FakePipeline.instances = []
    source = _workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=FakePipeline,
    )

    result = service.run(stage="qa", workbooks=[source], queries=["question"])

    assert len(FakePipeline.instances) == 2
    assert FakePipeline.instances[0].layout_vlm_client is not None
    assert FakePipeline.instances[0].forces == [True]
    assert FakePipeline.instances[1].layout_vlm_client is None
    assert result["answers"][0]["answer"] == "answer: question"


def test_qa_stage_accepts_per_run_max_replans_override(tmp_path: Path):
    FakePipeline.instances = []
    source = _workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {
            "service": {"root_dir": str(tmp_path / "service")},
            "table_agent": {"qa_max_replans": 5},
        },
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=FakePipeline,
    )

    service.run(
        stage="qa",
        workbooks=[source],
        queries=["question"],
        qa_max_replans=2,
    )

    assert FakePipeline.instances[1].config["qa_max_replans"] == 2


def test_service_rejects_negative_qa_max_replans(tmp_path: Path):
    service = TableAgentService({"service": {"root_dir": str(tmp_path / "service")}})
    source = _workbook(tmp_path / "book.xlsx")

    try:
        service.run(
            stage="qa",
            workbooks=[source],
            queries=["question"],
            qa_max_replans=-1,
        )
    except ValueError as exc:
        assert "qa_max_replans" in str(exc)
    else:
        raise AssertionError("Expected a negative qa_max_replans validation error")


def test_service_always_regenerates_structure_in_a_fresh_workspace(tmp_path: Path):
    FakePipeline.instances = []
    source = _workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=FakePipeline,
    )

    service.run(stage="structure", workbooks=[source])

    assert FakePipeline.instances[0].forces == [True]


def test_service_generates_readable_timestamp_job_id(tmp_path: Path):
    source = _workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=FakePipeline,
    )

    result = service.run(stage="structure", workbooks=[source])

    assert re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}-\d{2}-\d{2}\.\d{6}Z", result["job_id"])


def test_service_normalizes_repeated_comma_separated_sheet_filters(tmp_path: Path):
    FakePipeline.instances = []
    source = _multi_sheet_workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=FakePipeline,
    )

    result = service.run(
        stage="structure",
        workbooks=[source],
        sheets=["Summary, Detail", "Summary"],
        job_id="selected-sheets",
    )

    assert [item["sheet"] for item in result["structures"]] == ["Summary", "Detail"]
    schema_path = service.root_dir / "selected-sheets" / result["schema_artifacts"][0]["artifact"]
    assert list(yaml.safe_load(schema_path.read_text(encoding="utf-8"))) == ["Summary", "Detail"]


def test_ephemeral_service_run_returns_contents_without_persisting(tmp_path: Path):
    FakePipeline.instances = []
    source = _workbook(tmp_path / "book.xlsx")
    output_root = tmp_path / "output"
    service = TableAgentService(
        {"service": {"root_dir": str(output_root)}},
        llm_client=FakeSummaryClient(),
        layout_vlm_client=object(),
        pipeline_factory=FakePipeline,
    )

    result = service.run(stage="structure", workbooks=[source], persist=False)

    assert result["artifacts"] == []
    assert result["structures"][0]["artifact"] is None
    assert "table1" in result["schema_artifacts"][0]["schema"]
    assert result["metadata_artifacts"][0]["metadata"]["name"] == "book.xlsx"
    assert not output_root.exists()


def test_service_deletes_selected_or_all_saved_runs(tmp_path: Path):
    root = tmp_path / "output"
    service = TableAgentService({"service": {"root_dir": str(root)}})
    for run_id in ("run-one", "run-two"):
        run_dir = root / run_id
        run_dir.mkdir(parents=True)
        (run_dir / "run.json").write_text(f'{{"job_id": "{run_id}"}}', encoding="utf-8")
    unrelated = root / "unrelated"
    unrelated.mkdir()

    selected = service.delete_runs(["run-one", "missing"])
    all_runs = service.delete_runs(all_runs=True)

    assert selected == {"deleted": ["run-one"], "missing": ["missing"]}
    assert all_runs == {"deleted": ["run-two"], "missing": []}
    assert unrelated.is_dir()


def test_service_rejects_missing_sheet_before_pipeline_work(tmp_path: Path):
    FakePipeline.instances = []
    source = _workbook(tmp_path / "book.xlsx")
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        pipeline_factory=FakePipeline,
    )

    try:
        service.run(
            stage="structure",
            workbooks=[source],
            sheets=["Missing"],
        )
    except ValueError as exc:
        assert "book.xlsx: Missing" in str(exc)
    else:
        raise AssertionError("Expected a missing-sheet validation error")

    assert FakePipeline.instances == []


def test_service_rejects_queries_missing_from_qa_stage(tmp_path: Path):
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_client=object(),
        pipeline_factory=FakePipeline,
    )
    source = _workbook(tmp_path / "book.xlsx")

    try:
        service.run(stage="qa", workbooks=[source], queries=[])
    except ValueError as exc:
        assert "query" in str(exc)
    else:
        raise AssertionError("Expected an empty-query validation error")


def test_local_paths_are_disabled_by_default(tmp_path: Path):
    service = TableAgentService({"service": {"root_dir": str(tmp_path / "service")}})
    source = _workbook(tmp_path / "book.xlsx")

    try:
        service.validate_local_workbook(source)
    except PermissionError as exc:
        assert "disabled" in str(exc)
    else:
        raise AssertionError("Expected local paths to be disabled")


def test_service_uses_explicit_model_profiles(monkeypatch, tmp_path: Path):
    calls = []

    def fake_create_model_client(config, *, kind, profile):
        calls.append((kind, profile))
        return object()

    monkeypatch.setattr("service.runtime.create_model_client", fake_create_model_client)
    service = TableAgentService(
        {"service": {"root_dir": str(tmp_path / "service")}},
        llm_profile="alternate_answer",
        vlm_profile="alternate_layout",
    )

    service._answer_client()
    service._layout_client()

    assert calls == [
        ("llm", "alternate_answer"),
        ("vlm", "alternate_layout"),
    ]
