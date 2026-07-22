from __future__ import annotations

from datetime import datetime
import json
from pathlib import Path

import openpyxl
import pytest
import yaml

from TableAgent.artifacts import SummaryGenerator, build_workbook_metadata, build_workbook_schema
from TableAgent.llm import LLMResponse


class FakeSummaryLLM:
    def __init__(self, responses: list[str]):
        self.responses = list(responses)
        self.calls = []

    def generate(self, prompt, system_prompt=None):
        self.calls.append((prompt, system_prompt))
        return LLMResponse(content=self.responses.pop(0))


def test_schema_embeds_selected_sheet_structures_and_llm_descriptions(tmp_path: Path):
    summary = tmp_path / "summary.yaml"
    detail = tmp_path / "detail.yaml"
    summary.write_text("table1:\n  name: Revenue summary\n", encoding="utf-8")
    detail.write_text("table1:\n  name: Revenue detail\n", encoding="utf-8")
    llm = FakeSummaryLLM(
        [
            '{"description":"Summary of revenue."}',
            '{"description":"Detailed revenue records."}',
        ]
    )
    output = tmp_path / "schema.yaml"

    build_workbook_schema(
        [("Summary", summary), ("Detail View", detail)],
        output,
        SummaryGenerator(llm),
    )

    payload = yaml.safe_load(output.read_text(encoding="utf-8"))
    assert list(payload) == ["Summary", "Detail View"]
    assert payload["Summary"] == {
        "id": "summary",
        "description": "Summary of revenue.",
        "structure": {"table1": {"name": "Revenue summary"}},
    }
    assert payload["Detail View"]["id"] == "detail_view"
    assert payload["Detail View"]["description"] == "Detailed revenue records."
    assert len(llm.calls) == 2


def test_summary_generator_repairs_invalid_json_once():
    llm = FakeSummaryLLM(["not-json", '{"description":"Recovered"}'])

    description = SummaryGenerator(llm).sheet_description("Sheet1", "table1: {}")

    assert description == "Recovered"
    assert len(llm.calls) == 2


def test_summary_generator_fails_after_repair_retry():
    llm = FakeSummaryLLM(["not-json", "still-not-json"])

    with pytest.raises(ValueError, match="JSON object"):
        SummaryGenerator(llm).sheet_description("Sheet1", "table1: {}")


def test_metadata_uses_stable_fields_and_empty_description_without_schema(tmp_path: Path):
    source = tmp_path / "book.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "Summary"
    workbook.create_sheet("Detail")
    workbook.properties.creator = "Alice"
    workbook.properties.created = datetime(2024, 1, 2, 3, 4, 5)
    workbook.properties.modified = datetime(2024, 2, 3, 4, 5, 6)
    workbook.save(source)
    workbook.close()
    output = tmp_path / "metadata.json"

    build_workbook_metadata(source, "book.xlsx", output)

    payload = json.loads(output.read_text(encoding="utf-8"))
    date_modified = payload.pop("date_modified")
    assert payload == {
        "name": "book.xlsx",
        "description": "",
        "sheet_names": ["Summary", "Detail"],
        "author": "Alice",
        "date_created": "2024-01-02T03:04:05",
        "size_bytes": source.stat().st_size,
    }
    assert datetime.fromisoformat(date_modified)


def test_metadata_summarizes_existing_schema(tmp_path: Path):
    source = tmp_path / "book.xlsx"
    workbook = openpyxl.Workbook()
    workbook.save(source)
    workbook.close()
    schema = tmp_path / "schema.yaml"
    schema.write_text("Sheet:\n  description: Values\n", encoding="utf-8")
    output = tmp_path / "metadata.json"
    llm = FakeSummaryLLM(['{"description":"Workbook values."}'])

    build_workbook_metadata(
        source,
        "book.xlsx",
        output,
        schema_path=schema,
        summarizer=SummaryGenerator(llm),
    )

    assert json.loads(output.read_text(encoding="utf-8"))["description"] == "Workbook values."
    assert len(llm.calls) == 1
