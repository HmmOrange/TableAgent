from __future__ import annotations

from pathlib import Path

import openpyxl
import yaml

from TableAgent.environment.qa_env import QAEnvironment


def _write_structure(path: Path) -> None:
    headers = []
    for column, header_id, label in (
        ("A", "group", "Group"),
        ("B", "left", "Left"),
        ("C", "right", "Right"),
        ("D", "blank", "Blank"),
    ):
        headers.append({
            "id": header_id,
            "label": label,
            "description": label,
            "orientation": "column",
            "header_range": f"{column}1",
            "data_range": f"{column}2:{column}4",
            "sub_headers": [],
        })
    path.write_text(
        yaml.safe_dump(
            {
                "table1": {
                    "id": "table1",
                    "name": "Merged fixture",
                    "description": "Exercises merged table data.",
                    "sheet": "Sheet1",
                    "headers": headers,
                }
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )


def test_table_dataframe_expands_only_actual_merged_cells(tmp_path: Path):
    workbook_path = tmp_path / "merged.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Group", "Left", "Right", "Blank"])
    sheet["A2"] = "Group A"
    sheet.merge_cells("A2:A4")
    sheet["B2"] = "Shared"
    sheet.merge_cells("B2:C2")
    sheet["B3"], sheet["C3"] = "B3", "C3"
    sheet["B4"], sheet["C4"] = "B4", "C4"
    workbook.save(workbook_path)

    structure_path = tmp_path / "structure.yaml"
    _write_structure(structure_path)
    env = QAEnvironment(str(structure_path), str(workbook_path))
    try:
        # Raw range reads retain physical workbook storage semantics.
        assert env.operators.read_range("A2:A4", sheet="Sheet1") == [["Group A"], [None], [None]]

        dataframe = env.operators.read_table_as_dataframe("table1", has_headers=True)
        assert dataframe["group"].tolist() == ["Group A", "Group A", "Group A"]
        assert dataframe.loc[0, "left"] == "Shared"
        assert dataframe.loc[0, "right"] == "Shared"
        assert dataframe["blank"].isna().all()
    finally:
        env.workbook.close()


def test_structure_sheet_merged_equipment_name_expands_to_rows_1_through_10():
    root = Path(__file__).resolve().parents[1]
    workbook_path = (
        root
        / "data/SiFlex/golden_tests/data/설비/"
        / "LV01_설비_REPORT 2026년  설비유지보수 계획 VER 1.0_KR_202603.26.xlsx"
    )
    structure_path = root / "TableAgent/structure_sheet/structure.yaml"
    env = QAEnvironment(str(structure_path), str(workbook_path))
    try:
        dataframe = env.operators.read_table_as_dataframe("equipment_maintenance_plan", has_headers=True)
        numbers = dataframe["no"].astype(str).str.strip()
        first_ten = dataframe[numbers.isin({str(number) for number in range(1, 11)})]

        assert len(first_ten) == 10
        assert first_ten["equipment_name"].tolist() == ["화학동#1"] * 10
        assert first_ten["process"].tolist() == ["동도금"] * 10
    finally:
        env.workbook.close()
