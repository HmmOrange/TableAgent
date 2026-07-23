import json
from pathlib import Path

import yaml
from PIL import Image
import pytest

from TableAgent.schema import EvalSample
from TableAgent.pipeline import TableAgentPipeline
from TableAgent.llm import LLMResponse
from TableAgent.configs import TableAgentConfig


class FakeLLM:
    model_name = "fake"
    temperature = 0.0

    def __init__(self):
        self.calls = []

    def generate(self, prompt: str, system_prompt: str | None = None) -> LLMResponse:
        self.calls.append((prompt, system_prompt))
        if system_prompt and "layout agent" in system_prompt:
            return LLMResponse(
                content=yaml.safe_dump(
                    {
                        "structure": {
                            "table1": {
                                "name": "Revenue",
                                "description": "Company revenue values by year",
                                "headers": [
                                    {
                                        "label": "Revenue",
                                        "description": "Company revenue values by year",
                                        "orientation": "column",
                                        "header_range": "B1:B1",
                                        "data_range": "B2:B2",
                                        "sub_headers": [],
                                    }
                                ],
                            }
                        },
                        "changelog": "Added Revenue header.",
                        "remaining_directions": [],
                    },
                    sort_keys=False,
                ),
                prompt_tokens=10,
                completion_tokens=5,
            )
        if system_prompt and "verification agent" in system_prompt:
            return LLMResponse(
                content="status: good\nfeedback: Structure uses meaningful detected labels.\n",
                prompt_tokens=4,
                completion_tokens=2,
            )
        return LLMResponse(content="100", prompt_tokens=6, completion_tokens=1)


class SuccessfulQALLM(FakeLLM):
    def generate(self, prompt: str, system_prompt: str | None = None) -> LLMResponse:
        self.calls.append((prompt, system_prompt))
        if system_prompt and "verification agent" in system_prompt:
            return LLMResponse(
                content="status: good\nfeedback: Structure uses meaningful detected labels.\n",
                prompt_tokens=4,
                completion_tokens=2,
            )
        if system_prompt and "expert spreadsheet data planner" in system_prompt:
            return LLMResponse(
                content=json.dumps(
                    {
                        "subtasks": [
                            {
                                "id": "inspect_table",
                                "layer": "inspect",
                                "depends_on": [],
                                "description": "Inspect the available table.",
                            },
                            {
                                "id": "synthesize_answer",
                                "layer": "synthesis",
                                "depends_on": ["inspect_table"],
                                "description": "Compute the final answer.",
                            },
                        ]
                    },
                    sort_keys=False,
                ),
                prompt_tokens=7,
                completion_tokens=3,
            )
        if system_prompt and "spreadsheet analysis ReAct agent" in system_prompt:
            return LLMResponse(
                content=json.dumps(
                    {
                        "reasoning": "Inspect a compact preloaded table summary.",
                        "code": "inspected_shape = table_df.shape\nprint(inspected_shape)",
                        "description": "Stores and prints the table shape.",
                    }
                ),
                prompt_tokens=11,
                completion_tokens=4,
            )
        if system_prompt and "spreadsheet synthesis agent" in system_prompt:
            return LLMResponse(
                content=json.dumps(
                    {
                        "reasoning": "The sample answer is directly computed for this fixture.",
                        "code": "final_answer = '100'\nprint(final_answer)",
                        "description": "Sets the final answer.",
                    }
                ),
                prompt_tokens=13,
                completion_tokens=4,
            )
        if system_prompt and "strict table-QA reviewer" in system_prompt:
            return LLMResponse(
                content=json.dumps(
                    {
                        "accepted": True,
                        "score": 1.0,
                        "feedback": "Accepted.",
                    }
                ),
                prompt_tokens=5,
                completion_tokens=2,
            )
        return super().generate(prompt, system_prompt=system_prompt)


class FakeLayoutVLM:
    model_name = "fake-vlm"
    temperature = 0.0

    def __init__(self):
        self.calls = []

    def generate_with_image(self, prompt: str, image_path: Path, system_prompt: str | None = None) -> LLMResponse:
        self.calls.append((prompt, Path(image_path), system_prompt))
        structure = yaml.safe_dump(
            {
                "structure": {
                    "table1": {
                        "name": "Revenue",
                        "description": "Company revenue values by year",
                        "headers": [
                            {
                                "label": "Revenue",
                                "description": "Company revenue values by year",
                                "orientation": "column",
                                "header_range": "B1:B1",
                                "data_range": "B2:B2",
                                "sub_headers": [],
                            }
                        ],
                    }
                },
                "changelog": "Added Revenue header.",
                "remaining_directions": [],
            },
            sort_keys=False,
        )
        return LLMResponse(
            content=f"```yaml\n{structure}```",
            prompt_tokens=10,
            completion_tokens=5,
        )


@pytest.fixture(autouse=True)
def fake_libreoffice_workbook_render(monkeypatch, tmp_path):
    def fake_render(workbook_path, sheet_name, cell_range, image_path, **kwargs):
        image_path.parent.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", (100, 80), "white").save(image_path)

    monkeypatch.setattr("TableAgent.rendering.workbook._render_xlsx_range_with_libreoffice", fake_render)
    from TableAgent.configs import load_config
    real_from_config = TableAgentConfig.from_config

    def resolved_config(config=None):
        merged = dict(load_config("config.example.yaml")["table_agent"])
        explicit = config or {}
        if "table_agent" in explicit:
            explicit = explicit["table_agent"]
        merged.update(explicit)
        merged.setdefault("structure_cache_dir", str(tmp_path / "structure-cache"))
        if merged.get("structure_cache_dir") == "cache/table_agent/structure":
            merged["structure_cache_dir"] = str(tmp_path / "structure-cache")
        return real_from_config(merged)

    monkeypatch.setattr(TableAgentConfig, "from_config", staticmethod(resolved_config))


def test_table_agent_writes_verified_structure(tmp_path: Path):
    sample = EvalSample(
        index=0,
        sample_id="sample/1",
        table_id="table-1",
        table_content="Year | Revenue\n2024 | 100",
        question="What is the 2024 revenue?",
        answer=["100"],
    )
    llm = FakeLLM()
    layout_vlm = FakeLayoutVLM()
    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=layout_vlm,
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 1},
    )

    output = pipeline.run(sample)

    assert output.predicted_answer == "100"
    assert output.metadata["verification"]["status"] == "good"
    structure_path = Path(output.metadata["structure_path"])
    assert structure_path.is_file()
    structure = yaml.safe_load(structure_path.read_text(encoding="utf-8"))
    assert structure["table1"]["headers"][0]["label"] == "Revenue"
    assert Path(output.metadata["workbook_path"]).is_file()
    assert Path(output.metadata["image_path"]).is_file()
    assert output.metadata["html_path"] is None
    assert Path(output.metadata["changelog_path"]).is_file()
    assert Path(output.metadata["events_path"]).is_file()
    assert output.metadata["workbook_sheets"] == ["table-1"]
    assert len(layout_vlm.calls) == 1
    assert layout_vlm.calls[0][1].name == "viewport.png"
    assert output.metadata["qa"]["token_usage"] == {"prompt": 72, "completion": 12}
    assert output.token_usage == {"prompt": 78, "completion": 13}


def test_table_agent_counts_successful_qa_runner_tokens(tmp_path: Path):
    sample = EvalSample(
        index=0,
        sample_id="sample/1",
        table_id="table-1",
        table_content="Year | Revenue\n2024 | 100",
        question="What is the 2024 revenue?",
        answer=["100"],
    )
    pipeline = TableAgentPipeline(
        llm_client=SuccessfulQALLM(),
        layout_vlm_client=FakeLayoutVLM(),
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 1},
    )

    output = pipeline.run(sample)

    assert output.predicted_answer == "100"
    assert output.metadata["qa"]["success"] is True
    assert output.metadata["qa"]["fallback_used"] is False
    assert output.metadata["qa"]["token_usage"] == {"prompt": 41, "completion": 15}
    assert output.token_usage == {"prompt": 41, "completion": 15}


def test_table_agent_can_disable_source_retrieval(tmp_path: Path, monkeypatch):
    sample = EvalSample(
        index=0,
        sample_id="siflex/disabled",
        table_id="table-1",
        table_content="Year | Revenue\n2024 | 100",
        question="What is the 2024 revenue?",
        answer=["100"],
        sample_path="data/SiFlex/golden_tests/compiled/golden_cases.json:cases[0]",
        table_path="/unused/source.xlsx",
    )
    pipeline = TableAgentPipeline(
        llm_client=SuccessfulQALLM(),
        layout_vlm_client=FakeLayoutVLM(),
        config={
            "artifact_dir": str(tmp_path),
            "structure_cache_dir": str(tmp_path / "cache"),
            "max_refinement_rounds": 1,
            "run_retrieval": False,
        },
    )

    def fail_prepare(*args, **kwargs):
        raise AssertionError("source preparation should be disabled")

    def fail_select(*args, **kwargs):
        raise AssertionError("source retrieval should be disabled")

    monkeypatch.setattr(pipeline.source_preparer, "prepare", fail_prepare)
    monkeypatch.setattr(pipeline.source_retriever, "select", fail_select)

    pipeline.prepare_samples([sample])
    output = pipeline.run(sample)

    assert output.predicted_answer == "100"
    assert output.metadata["workbook_source_format"] == "verification-cache"
    assert "retrieval_info" not in output.metadata


def test_table_agent_all_phase_aborts_on_invalid_prepared_cache(tmp_path: Path, monkeypatch):
    from TableAgent.pipeline.structure_cache import StructureCacheRecord

    sample = EvalSample(
        index=0,
        sample_id="sample/invalid-cache",
        table_id="table-1",
        table_content="Year | Revenue\n2024 | 100",
        question="What is the 2024 revenue?",
        answer=["100"],
    )
    pipeline = TableAgentPipeline(
        llm_client=SuccessfulQALLM(),
        layout_vlm_client=FakeLayoutVLM(),
        config={
            "artifact_dir": str(tmp_path),
            "structure_cache_dir": str(tmp_path / "cache"),
            "max_refinement_rounds": 1,
            "phase": "all",
        },
    )
    invalid = StructureCacheRecord(
        key="bad",
        directory=tmp_path,
        workbook_path=tmp_path / "workbook.xlsx",
        sheet_name="Sheet1",
        structure_path=tmp_path / "structure.yaml",
        manifest_path=tmp_path / "manifest.json",
        status="not_good",
        cache_hit=False,
    )
    monkeypatch.setattr(pipeline, "verify_samples", lambda samples, force=True: [invalid])

    with pytest.raises(RuntimeError, match="verification failed for 1 cache entries"):
        pipeline.prepare_samples([sample])


def test_table_agent_structure_phase_aborts_on_invalid_prepared_cache(tmp_path: Path, monkeypatch):
    from TableAgent.pipeline.structure_cache import StructureCacheRecord

    sample = EvalSample(
        index=0,
        sample_id="sample/invalid-cache",
        table_id="table-1",
        table_content="Year | Revenue\n2024 | 100",
        question="What is the 2024 revenue?",
        answer=["100"],
    )
    pipeline = TableAgentPipeline(
        llm_client=None,
        layout_vlm_client=FakeLayoutVLM(),
        config={
            "artifact_dir": str(tmp_path),
            "structure_cache_dir": str(tmp_path / "cache"),
            "max_refinement_rounds": 1,
            "phase": "structure",
        },
    )
    invalid = StructureCacheRecord(
        key="bad",
        directory=tmp_path,
        workbook_path=tmp_path / "workbook.xlsx",
        sheet_name="Sheet1",
        structure_path=tmp_path / "structure.yaml",
        manifest_path=tmp_path / "manifest.json",
        status="not_good",
        cache_hit=False,
    )
    monkeypatch.setattr(pipeline, "verify_samples", lambda samples, force=True: [invalid])

    with pytest.raises(RuntimeError, match="verification failed for 1 cache entries"):
        pipeline.prepare_samples([sample])


def test_prepared_source_qa_uses_retrieved_table_structure(tmp_path: Path, monkeypatch):
    from TableAgent.pipeline.common import SourceCandidate

    source_dir = tmp_path / "source"
    source_dir.mkdir()
    full_structure = "table1:\n  id: table1\n  name: Sales\ntable2:\n  id: table2\n  name: Costs\n"
    selected_structure = "table2:\n  id: table2\n  name: Costs\n"
    (source_dir / "structure.yaml").write_text(full_structure, encoding="utf-8")
    candidate = SourceCandidate(
        directory=source_dir,
        workbook_path=tmp_path / "book.xlsx",
        sheet_name="Sheet1",
        image_path=source_dir / "table.png",
        html_path=None,
        structure_text=selected_structure,
        sheet_text="cost data",
        score=1.0,
        table_id="table2",
        table_name="Costs",
    )
    sample = EvalSample(
        index=0,
        sample_id="siflex/table-level",
        table_id="source",
        table_content="",
        question="What are the costs?",
        answer=["100"],
        sample_path="siflex",
    )
    pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=FakeLayoutVLM(),
        config={"artifact_dir": str(tmp_path / "artifacts")},
    )
    captured = {}

    def fake_run_verified_qa(**kwargs):
        captured["structure_path"] = kwargs["structure_path"]
        return LLMResponse(content="100"), {"success": True}

    monkeypatch.setattr(pipeline, "_run_verified_qa", fake_run_verified_qa)
    monkeypatch.setattr(pipeline, "_format_siflex_answer", lambda sample, answer, responses: answer)

    output = pipeline._run_prepared_source(sample, candidate, [], pipeline.start_timer())
    selected_path = captured["structure_path"]

    assert selected_path.name == "retrieved_structure.yaml"
    assert selected_path.read_text(encoding="utf-8") == selected_structure
    assert output.metadata["structure_path"] == str(selected_path).replace("\\", "/")
    assert output.metadata["retrieval_info"]["table_id"] == "table2"
    assert "table1:" not in output.structured_table


def test_table_agent_qa_phase_reuses_structure_cache(tmp_path: Path):
    sample = EvalSample(
        index=0,
        sample_id="cache/1",
        table_id="table-1",
        table_content="Year | Revenue\n2024 | 100",
        question="What is the revenue?",
        answer=["100"],
    )
    cache_dir = tmp_path / "structure-cache"
    all_pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=FakeLayoutVLM(),
        config={"artifact_dir": str(tmp_path / "all"), "structure_cache_dir": str(cache_dir), "phase": "all"},
    )
    first = all_pipeline.run(sample)

    qa_pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=None,
        config={"artifact_dir": str(tmp_path / "qa"), "structure_cache_dir": str(cache_dir), "phase": "qa"},
    )
    second = qa_pipeline.run(sample)

    assert second.metadata["cache_hit"] is True
    assert second.metadata["cache_key"] == first.metadata["cache_key"]
    assert second.predicted_answer == "100"


def test_table_agent_qa_phase_fails_on_missing_cache(tmp_path: Path):
    sample = EvalSample(0, "missing/1", "table-1", "A | B\n1 | 2", "What is B?", ["2"])
    pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=None,
        config={"artifact_dir": str(tmp_path), "structure_cache_dir": str(tmp_path / "missing"), "phase": "qa"},
    )
    with pytest.raises(RuntimeError, match="Missing or stale structure cache"):
        pipeline.run(sample)


def test_table_agent_structure_phase_does_not_require_answer_llm(tmp_path: Path):
    sample = EvalSample(0, "verify/1", "table-1", "Year | Revenue\n2024 | 100", "Unused", ["100"])
    pipeline = TableAgentPipeline(
        llm_client=None,
        layout_vlm_client=FakeLayoutVLM(),
        config={"artifact_dir": str(tmp_path), "structure_cache_dir": str(tmp_path / "cache"), "phase": "structure"},
    )
    records = pipeline.verify_samples([sample])
    assert records[0].valid
    with pytest.raises(RuntimeError, match="structure phase does not run question answering"):
        pipeline.run(sample)


def test_table_agent_structure_progress_counts_files_and_sheets(tmp_path: Path):
    import openpyxl

    workbook_path = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    workbook.create_sheet("Second")
    workbook.save(workbook_path)
    workbook.close()
    samples = [
        EvalSample(0, "hitab/1", "table-1", "A | B", "Question", ["Answer"]),
        EvalSample(
            1,
            "siflex/1",
            "table-2",
            "",
            "Question",
            ["Answer"],
            table_path=str(workbook_path),
        ),
    ]

    totals = TableAgentPipeline.structure_progress_totals(samples)

    assert totals["files"] == 2
    assert totals["sheets"] == 3
    assert totals["sheets_per_file"] == {
        "sample:hitab/1": 1,
        "book:source.xlsx": 2,
    }


def test_source_preparer_force_regenerates_valid_structure(tmp_path: Path):
    import openpyxl

    workbook_path = tmp_path / "source.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet["B1"] = "Revenue"
    worksheet["B2"] = 100
    workbook.save(workbook_path)
    workbook.close()
    sample = EvalSample(
        0,
        "siflex/force",
        "table-1",
        "",
        "Question",
        ["100"],
        table_path=str(workbook_path),
    )
    layout_vlm = FakeLayoutVLM()
    pipeline = TableAgentPipeline(
        llm_client=None,
        layout_vlm_client=layout_vlm,
        config={"artifact_dir": str(tmp_path / "artifacts"), "phase": "structure"},
    )

    pipeline.source_preparer.prepare([sample])
    initial_call_count = len(layout_vlm.calls)
    assert initial_call_count > 0

    pipeline.source_preparer.prepare([sample])
    assert len(layout_vlm.calls) == initial_call_count

    pipeline.source_preparer.prepare([sample], force=True)
    assert len(layout_vlm.calls) > initial_call_count


def test_table_agent_pipeline_does_not_preselect_table_for_qa(tmp_path: Path, monkeypatch):
    import openpyxl
    import TableAgent.pipeline.table_agent_pipeline as pipeline_module

    captured_configs = []

    class _WorkbookHandle:
        def close(self):
            pass

    class _Env:
        workbook = _WorkbookHandle()

    class CapturingRunner:
        def __init__(
            self,
            *,
            structure_path,
            workbook_path,
            llm_client,
            config,
            table_retriever=None,
            related_structure_paths=None,
        ):
            captured_configs.append(config)
            self.env = _Env()

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, traceback):
            self.env.workbook.close()

        def run(self, question):
            class Result:
                success = True
                final_answer = "ok"
                error = None
                execution_time = 0.0
                token_usage = {"prompt": 1, "completion": 1}
                artifacts = {}

            return Result()

        def token_usage(self):
            return {"prompt": 1, "completion": 1}

    monkeypatch.setattr(pipeline_module, "TableQARunner", CapturingRunner)

    workbook_path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    wb.save(workbook_path)

    structure_path = tmp_path / "structure.yaml"
    structure_path.write_text(
        yaml.safe_dump(
            {
                "table1": {
                    "id": "table1",
                    "name": "Wrong lexical match",
                    "description": "This table should not be preselected.",
                    "sheet": "Sheet1",
                    "headers": [],
                },
                "table2": {
                    "id": "table2",
                    "name": "Right table",
                    "description": "The table inspect layer should choose this later.",
                    "sheet": "Sheet1",
                    "headers": [],
                },
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=FakeLayoutVLM(),
        config={"artifact_dir": str(tmp_path)},
    )

    response, qa_info = pipeline._run_verified_qa(
        question="Which table is right?",
        structure_path=structure_path,
        workbook_path=workbook_path,
        qa_artifact_dir=tmp_path / "qa",
        fallback_prompt="fallback",
    )

    assert response.content == "ok"
    assert qa_info["success"] is True
    assert captured_configs
    assert "table_id" not in captured_configs[0]


def legacy_table_agent_siflex_retrieval(tmp_path: Path):
    import openpyxl
    # Create two dummy workbooks
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "SheetA"
    ws1["A1"] = "Apple"
    ws1["B1"] = "Banana"
    path_a = tmp_path / "doc_a.xlsx"
    wb1.save(path_a)

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "SheetB"
    ws2["A1"] = "Orange"
    ws2["B1"] = "Grape"
    path_b = tmp_path / "doc_b.xlsx"
    wb2.save(path_b)

    sample = EvalSample(
        index=0,
        sample_id="siflex/1",
        table_id="siflex-table",
        table_content="Apple | Banana\nOrange | Grape",
        question="Which sheet has the Apple and Banana?",
        answer=["SheetA"],
        sample_path="data/SiFlex/golden_tests/compiled/golden_cases.json:cases[0]",
        table_path=f"{path_a};{path_b}",
    )

    llm = FakeLLM()
    # Mock generate_with_image on LLM for XLSX answering flow
    llm.generate_with_image = lambda prompt, image_path, system_prompt=None: LLMResponse(
        content="SheetA", prompt_tokens=15, completion_tokens=3
    )

    layout_vlm = FakeLayoutVLM()

    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=layout_vlm,
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 1},
    )

    # 1. Pre-encode sheets
    pipeline.prepare_samples([sample])

    # Verify that structures/sources are generated
    sources_dir = tmp_path / "sources"
    assert sources_dir.is_dir()
    assert (sources_dir / "doc_a.xlsx_SheetA" / "structure.yaml").is_file()
    assert (sources_dir / "doc_a.xlsx_SheetA" / "sheet_text.txt").is_file()
    assert (sources_dir / "doc_b.xlsx_SheetB" / "structure.yaml").is_file()
    assert (sources_dir / "doc_b.xlsx_SheetB" / "sheet_text.txt").is_file()

    # 2. Run the question answering flow
    output = pipeline.run(sample)

    assert output.predicted_answer == "SheetA"
    assert output.metadata["workbook_path"] == str(path_a.resolve())
    assert output.metadata["workbook_sheets"] == ["SheetA"]
    assert "doc_a.xlsx_SheetA/table.png" in output.metadata["image_path"]


def legacy_table_agent_run_prepares_siflex_source_lazily(tmp_path: Path):
    import openpyxl

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "SheetA"
    worksheet["A1"] = "Apple"
    workbook_path = tmp_path / "doc_a.xlsx"
    workbook.save(workbook_path)

    sample = EvalSample(
        index=0,
        sample_id="siflex/lazy",
        table_id="siflex-table",
        table_content="Apple",
        question="Which sheet has Apple?",
        answer=["SheetA"],
        sample_path="data/SiFlex/golden_tests/compiled/golden_cases.json:cases[0]",
        table_path=str(workbook_path),
    )

    llm = FakeLLM()
    llm.generate_with_image = lambda prompt, image_path, system_prompt=None: LLMResponse(
        content="SheetA", prompt_tokens=15, completion_tokens=3
    )
    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=FakeLayoutVLM(),
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 1},
    )
    progress_messages = []
    pipeline.set_progress_callback(progress_messages.append)

    assert pipeline.prepare_samples_before_run is False
    assert not (tmp_path / "sources").exists()

    output = pipeline.run(sample)

    assert output.predicted_answer == "SheetA"
    assert (tmp_path / "sources" / "doc_a.xlsx_SheetA" / "structure.yaml").is_file()
    assert output.metadata["workbook_path"] == str(workbook_path.resolve())
    assert any(message.startswith("prepare:extract") for message in progress_messages)
    assert any(message.startswith("prepare:layout") for message in progress_messages)
    assert any(message.startswith("render | range=") for message in progress_messages)


def legacy_table_agent_prepare_samples_regenerates_invalid_structure(tmp_path: Path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SheetX"
    ws["A1"] = "Data"
    path_x = tmp_path / "doc_x.xlsx"
    wb.save(path_x)

    sample = EvalSample(
        index=0,
        sample_id="siflex/invalid_test",
        table_id="siflex-table-invalid",
        table_content="Data",
        question="What is Data?",
        answer=["Data"],
        sample_path="data/SiFlex/golden_tests/compiled/golden_cases.json:cases[0]",
        table_path=str(path_x),
    )

    llm = FakeLLM()
    layout_vlm = FakeLayoutVLM()

    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=layout_vlm,
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 1},
    )

    # Pre-create the directory and put an invalid structure.yaml in there
    sheet_dir = tmp_path / "sources" / "doc_x.xlsx_SheetX"
    sheet_dir.mkdir(parents=True, exist_ok=True)
    invalid_structure_path = sheet_dir / "structure.yaml"
    invalid_structure_path.write_text("ERROR: Connection error.", encoding="utf-8")

    # Run prepare_samples
    pipeline.prepare_samples([sample])

    # Assert that structure.yaml was regenerated and contains valid headers YAML
    assert invalid_structure_path.is_file()
    content = invalid_structure_path.read_text(encoding="utf-8")
    assert "ERROR" not in content
    assert "headers" in content


def legacy_table_agent_run_ignores_invalid_structures(tmp_path: Path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SheetY"
    ws["A1"] = "Apple"
    path_y = tmp_path / "doc_y.xlsx"
    wb.save(path_y)

    sample = EvalSample(
        index=0,
        sample_id="siflex/ignore_test",
        table_id="siflex-table-ignore",
        table_content="Apple",
        question="Which sheet?",
        answer=["SheetY"],
        sample_path="data/SiFlex/golden_tests/compiled/golden_cases.json:cases[0]",
        table_path=str(path_y),
    )

    llm = FakeLLM()
    layout_vlm = FakeLayoutVLM()

    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=layout_vlm,
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 1},
    )

    # 1. Pre-encode sheets
    pipeline.prepare_samples([sample])

    # 2. Overwrite structure.yaml with connection error
    structure_path = tmp_path / "sources" / "doc_y.xlsx_SheetY" / "structure.yaml"
    assert structure_path.is_file()
    structure_path.write_text("ERROR: Connection error.", encoding="utf-8")

    # 3. Running should ignore the candidate directory because of the invalid structure.yaml,
    # and fall back. Since the fallback path executes, we verify it falls through.
    output = pipeline.run(sample)

    assert output.metadata["verification"]["feedback"] != "Retrieved from encoded source"


def legacy_table_agent_fallback_invalid_structure_marked_not_good(tmp_path: Path):
    sample = EvalSample(
        index=0,
        sample_id="siflex/fallback_error",
        table_id="siflex-table-err",
        table_content="Apple",
        question="Which sheet?",
        answer=["SheetY"],
    )

    class ErrorLLM:
        model_name = "error-llm"
        temperature = 0.0
        def generate(self, prompt: str, system_prompt: str | None = None) -> LLMResponse:
            return LLMResponse(content="ERROR: Connection error.", prompt_tokens=0, completion_tokens=0)

    class ErrorLayoutVLM:
        model_name = "error-vlm"
        temperature = 0.0
        def generate_with_image(self, prompt: str, image_path: Path, system_prompt: str | None = None) -> LLMResponse:
            return LLMResponse(content="ERROR: Connection error.", prompt_tokens=0, completion_tokens=0)

    pipeline = TableAgentPipeline(
        llm_client=ErrorLLM(),
        layout_vlm_client=ErrorLayoutVLM(),
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 1},
    )

    output = pipeline.run(sample)
    assert output.metadata["verification"]["status"] == "not_good"
    assert "invalid" in output.metadata["verification"]["feedback"].lower() or "empty" in output.metadata["verification"]["feedback"].lower()


def test_table_agent_siflex_answer_prompt_formatting(tmp_path: Path):
    pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=FakeLayoutVLM(),
        config={"artifact_dir": str(tmp_path)},
    )

    # 1. Non-SiFlex sample
    non_siflex_sample = EvalSample(
        index=0,
        sample_id="hitab/1",
        table_id="t1",
        table_content="Content",
        question="What is the answer?",
        answer=["Ans"],
        sample_path="data/HiTab/split.json:cases[0]",
    )
    prompt = pipeline._answer_prompt(non_siflex_sample, "Content", "headers: []")
    assert "FORMAT INSTRUCTIONS" not in prompt
    assert "Verified structure.yaml" in prompt

    # 2. SiFlex table sample
    siflex_table_sample = EvalSample(
        index=0,
        sample_id="siflex/1",
        table_id="t1",
        table_content="Content",
        question="What is the answer?",
        answer=["Ans"],
        sample_path="data/SiFlex/compiled/golden_cases.json:cases[0]",
        raw={"answer_type": "table"},
    )
    prompt = pipeline._answer_prompt(siflex_table_sample, "Content", "headers: []")
    assert "FORMAT INSTRUCTIONS" in prompt
    assert "CRITICAL EXPECTED FORMAT: TABLE" in prompt
    assert "Format your final answer as a markdown table" in prompt

    # 3. SiFlex list sample
    siflex_list_sample = EvalSample(
        index=0,
        sample_id="siflex/2",
        table_id="t1",
        table_content="Content",
        question="What is the answer?",
        answer=["Ans"],
        sample_path="data/SiFlex/compiled/golden_cases.json:cases[1]",
        raw={"answer_type": "list"},
    )
    prompt = pipeline._answer_prompt(siflex_list_sample, "Content", "headers: []")
    assert "FORMAT INSTRUCTIONS" in prompt
    assert "CRITICAL EXPECTED FORMAT: LIST" in prompt
    assert "Format your final answer as a bulleted list" in prompt

    # 4. SiFlex form sample
    siflex_form_sample = EvalSample(
        index=0,
        sample_id="siflex/3",
        table_id="t1",
        table_content="Content",
        question="What is the answer?",
        answer=["Ans"],
        sample_path="data/SiFlex/compiled/golden_cases.json:cases[2]",
        raw={"answer_type": "form"},
    )
    prompt = pipeline._answer_prompt(siflex_form_sample, "Content", "headers: []")
    assert "FORMAT INSTRUCTIONS" in prompt
    assert "CRITICAL EXPECTED FORMAT: FORM/DOCUMENT" in prompt
    assert "Organize your final answer in a clear document structure" in prompt


def test_table_agent_default_applies_generation_cap_and_early_breaks(tmp_path: Path):
    llm = FakeLLM()
    llm.max_tokens = None
    layout_vlm = FakeLayoutVLM()
    layout_vlm.max_tokens = 4096
    
    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=layout_vlm,
        config={"artifact_dir": str(tmp_path)},
    )
    assert llm.max_tokens == 8192
    assert layout_vlm.max_tokens == 8192

    # Test fit_context truncation
    long_content = "X" * 70000
    fitted = pipeline._fit_context(long_content)
    assert len(fitted) <= 60000 + 50
    assert "...TRUNCATED..." in fitted


def test_table_agent_max_tokens_config_driven(tmp_path: Path):
    llm = FakeLLM()
    llm.max_tokens = 100
    layout_vlm = FakeLayoutVLM()
    layout_vlm.max_tokens = 200
    
    # Config overrides default cap
    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=layout_vlm,
        config={"artifact_dir": str(tmp_path), "generation_max_tokens": 1024},
    )
    assert llm.max_tokens == 1024
    assert layout_vlm.max_tokens == 1024
    assert pipeline.get_config()["agent"]["generation_max_tokens"] == 1024

    # Config None disables capping/mutating
    llm.max_tokens = 500
    layout_vlm.max_tokens = 600
    pipeline_none = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=layout_vlm,
        config={"artifact_dir": str(tmp_path), "generation_max_tokens": None},
    )
    assert llm.max_tokens == 500
    assert layout_vlm.max_tokens == 600
    assert pipeline_none.get_config()["agent"]["generation_max_tokens"] is None


def test_is_valid_structure_rules():
    from TableAgent.structure.layout.parsing import _is_valid_structure
    
    # 1. Reject structures with 'error' key
    assert not _is_valid_structure("headers: []\nerror: Failed to generate structure")
    assert not _is_valid_structure("error: Failed to generate structure")
    
    # 2. Reject empty headers
    assert not _is_valid_structure("headers: []")
    
    # 3. Reject structures with only placeholder/vague headers
    assert not _is_valid_structure("headers:\n  - Column 1\n  - Column 2")
    assert not _is_valid_structure("headers:\n  - col1\n  - col 2\n  - Placeholder")
    assert not _is_valid_structure("headers:\n  - -")
    assert not _is_valid_structure("headers:\n  - Empty\n  - None")
    
    # 4. Accept if at least one header is not placeholder
    assert _is_valid_structure("headers:\n  - Column 1\n  - Revenue\n  - Column 2")
    assert _is_valid_structure("headers:\n  - label: Column 1\n  - label: Total Revenue")
    assert _is_valid_structure("headers:\n  - name: Column 1\n  - name: Net Profit")


def test_table_agent_layout_prompt_uses_deterministic_feedback():
    from TableAgent.prompts.structure import (
        LAYOUT_MAS_SYSTEM_PROMPT,
        LAYOUT_MAS_USER_PROMPT_TEMPLATE,
    )

    assert "never output null, UNKNOWN, or placeholder range values" in LAYOUT_MAS_SYSTEM_PROMPT
    assert "header_range` is only the cell or merged/spanned cells" in LAYOUT_MAS_USER_PROMPT_TEMPLATE
    assert "data starts below all header and sub-header rows" in LAYOUT_MAS_USER_PROMPT_TEMPLATE
    assert "must not include child `header_range` cells" in LAYOUT_MAS_USER_PROMPT_TEMPLATE
    assert "Never write `null`, `UNKNOWN`, `N/A`, or placeholder range values" in LAYOUT_MAS_USER_PROMPT_TEMPLATE
    assert "use the union of the old range and newly visible" in LAYOUT_MAS_USER_PROMPT_TEMPLATE
    assert "Do not create separate headers for blank cells inside a merged" in LAYOUT_MAS_USER_PROMPT_TEMPLATE
    assert "deterministic verifier" in LAYOUT_MAS_USER_PROMPT_TEMPLATE.lower()


def test_strict_structure_normalizes_uncertain_ranges_to_null():
    from TableAgent.structure.layout.parsing import extract_strict_structure

    structure_text, _ = extract_strict_structure(
        "headers:\n"
        "- label: Revenue\n"
        "  range: UNKNOWN\n"
        "  sub_headers:\n"
        "  - label: 2024\n"
        "    range: uncertain\n"
    )
    structure = yaml.safe_load(structure_text)

    assert structure["headers"][0]["range"] is None
    assert structure["headers"][0]["sub_headers"][0]["range"] is None


def test_strict_structure_extraction_discards_reasoning_and_extra_keys():
    from TableAgent.structure.layout.parsing import extract_strict_structure

    content = """I analyzed the table before producing the result.
```yaml
reasoning: this must never be persisted
headers:
  - label: Revenue
    description: Annual revenue
    orientation: column
    range: B1:B3
    confidence: 0.9
    sub_headers:
      - label: "2024"
        description: Fiscal year 2024
        orientation: column
        range: B2
```
This trailing explanation is also logging-only."""

    structure_text, discarded = extract_strict_structure(content)
    structure = yaml.safe_load(structure_text)

    assert set(structure) == {"headers"}
    assert set(structure["headers"][0]) == {"label", "description", "orientation", "range", "sub_headers"}
    assert set(structure["headers"][0]["sub_headers"][0]) == {"label", "description", "orientation", "range"}
    assert "reasoning" not in structure_text
    assert "confidence" not in structure_text
    assert "analyzed the table" in discarded
    assert "trailing explanation" in discarded
    assert "reasoning" in discarded
    assert "confidence" in discarded


def test_table_agent_persists_only_strict_structure(tmp_path: Path):
    class ReasoningLayoutVLM(FakeLayoutVLM):
        def generate_with_image(self, prompt: str, image_path: Path, system_prompt: str | None = None) -> LLMResponse:
            return LLMResponse(
                content="""Analysis that belongs in logs.
```yaml
structure:
  table1:
    name: Revenue
    description: Annual revenue
    headers:
      - label: Revenue
        description: Annual revenue
        orientation: column
        header_range: B1:B1
        data_range: B2:B2
        sub_headers: []
changelog: Added revenue.
remaining_directions: []
```""",
                prompt_tokens=10,
                completion_tokens=5,
            )

    sample = EvalSample(
        index=0,
        sample_id="strict/1",
        table_id="table-1",
        table_content="Year | Revenue\n2024 | 100",
        question="What is the revenue?",
        answer=["100"],
    )
    pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=ReasoningLayoutVLM(),
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 0},
    )

    output = pipeline.run(sample)
    persisted = Path(output.metadata["structure_path"]).read_text(encoding="utf-8")

    assert persisted.startswith("table1:")
    assert "Analysis that belongs in logs" not in persisted


def test_layout_parser_preserves_unquoted_no_header():
    from TableAgent.structure.layout.parsing import extract_layout_structure

    content = """structure:
  table1:
    id: people
    name: People
    headers:
      - id: no
        label: No
        orientation: column
        header_range: A1:A2
        data_range: A3:A22
        sub_headers: []
changelog: Added the row number header.
remaining_directions: []
"""

    structure_text, discarded, directions, changelog = extract_layout_structure(content)
    structure = yaml.safe_load(structure_text)

    assert structure["table1"]["headers"][0]["id"] == "no"
    assert structure["table1"]["headers"][0]["label"] == "No"
    assert discarded == ""
    assert directions == []
    assert changelog == "Added the row number header."


def test_layout_parser_salvages_structure_when_changelog_breaks_yaml():
    from TableAgent.structure.layout.parsing import extract_layout_structure

    content = """```yaml
structure:
  table1:
    id: hours_of_work
    name: Hours of work
    headers:
      - id: hours
        label: Hours
        orientation: column
        header_range: A3
        data_range: A5:A17
        sub_headers: []
changelog: Added headers (sub-headers: All industries and Agriculture).
remaining_directions: [right]
```"""

    structure_text, discarded, directions, changelog = extract_layout_structure(content)
    structure = yaml.safe_load(structure_text)

    assert structure["table1"]["headers"][0]["data_range"] == "A5:A17"
    assert directions == ["right"]
    assert changelog == "Added headers (sub-headers: All industries and Agriculture)."
    assert "changelog:" in discarded


def test_layout_parser_does_not_salvage_malformed_structure_block():
    from TableAgent.structure.layout.parsing import extract_layout_structure

    content = """structure:
  table1:
    headers:
      - label: Revenue
        header_range: [invalid
changelog: Added headers (sub-headers: Revenue).
remaining_directions: []
"""

    structure_text, _, directions, changelog = extract_layout_structure(content)

    assert structure_text == ""
    assert directions == []
    assert changelog == ""


def test_table_agent_separates_artifacts_by_benchmark_repeat(tmp_path: Path):
    from TableAgent.configs import run_scoped_table_agent_config

    scoped_config = run_scoped_table_agent_config(
        {
            "table_agent": {
                "artifact_root": str(tmp_path),
                "run_dir_template": "{run_name}",
                "repeat_dir_template": "repeat_{run_id}",
                "shared_dir_name": "shared",
            }
        },
        "hitab-table_agent-20260621_163851",
    )
    pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=FakeLayoutVLM(),
        config=scoped_config,
    )
    sample = EvalSample(
        index=0,
        sample_id="sample/1",
        table_id="table-1",
        table_content="Year | Revenue\n2024 | 100",
        question="What is the revenue?",
        answer=["100"],
    )

    pipeline.set_run_id(1)
    first_output = pipeline.run(sample)
    pipeline.set_run_id(3)
    third_output = pipeline.run(sample)

    assert Path(first_output.metadata["structure_path"]) == Path(third_output.metadata["structure_path"])
    assert first_output.metadata["cache_key"] == third_output.metadata["cache_key"]
    assert pipeline.settings.source_artifact_dir == Path("cache/table_agent/structure/v5/prepared")


def test_table_agent_scoped_config_preserves_explicit_source_artifacts(tmp_path: Path):
    from TableAgent.configs import run_scoped_table_agent_config

    source_artifacts = tmp_path / "prior" / "prepared"
    scoped = run_scoped_table_agent_config(
        {
            "table_agent": {
                "artifact_root": str(tmp_path / "new"),
                "source_artifact_dir": str(source_artifacts),
                "phase": "qa",
                "perfect_retrieval": True,
            }
        },
        "new-run",
    )

    assert scoped["source_artifact_dir"] == str(source_artifacts)
    assert scoped["phase"] == "qa"
    assert scoped["perfect_retrieval"] is True


def test_perfect_retrieval_filters_excluded_sheet_samples(tmp_path: Path, monkeypatch):
    pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=None,
        config={
            "artifact_dir": str(tmp_path / "run"),
            "source_artifact_dir": str(tmp_path / "prior"),
            "phase": "qa",
            "perfect_retrieval": True,
        },
    )
    keep = EvalSample(0, "keep", "book", "", "keep", [""], sample_path="siflex")
    skip = EvalSample(1, "skip", "book", "", "skip", [""], sample_path="siflex")

    def select_perfect(sample):
        if sample.sample_id == "skip":
            raise RuntimeError("Perfect retrieval excludes sheet 'Sheet3' from 'book.xlsx'.")
        return object()

    monkeypatch.setattr(pipeline.source_retriever, "select_perfect", select_perfect)

    assert pipeline.filter_samples([keep, skip]) == [keep]


def test_perfect_retrieval_skips_benchmark_exclusions(tmp_path: Path, monkeypatch):
    pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=None,
        config={
            "artifact_dir": str(tmp_path / "run"),
            "phase": "qa",
            "perfect_retrieval": True,
        },
    )
    skipped = [
        EvalSample(
            0,
            "plasma",
            "book",
            "",
            "❓ Câu hỏi: Sheet PLASMA quản lý spare parts cho công đoạn nào? "
            "Các loại phụ tùng chính được quản lý trong sheet này là gì?",
            [""],
            sample_path="siflex",
        ),
        EvalSample(
            1,
            "maintenance-summary",
            "book",
            "",
            "전체 점검 건수와 보수 건수, 그리고 완료 처리된 항목은?",
            [""],
            sample_path="siflex",
        ),
    ]
    keep = EvalSample(2, "keep", "book", "", "Sheet OIL quản lý loại phụ tùng nào?", [""], sample_path="siflex")
    monkeypatch.setattr(pipeline.source_retriever, "select_perfect", lambda _sample: object())

    assert pipeline.filter_samples([*skipped, keep]) == [keep]


def test_perfect_retrieval_exclusions_are_disabled_by_default(tmp_path: Path):
    pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=None,
        config={
            "artifact_dir": str(tmp_path / "run"),
            "phase": "qa",
            "perfect_retrieval": False,
        },
    )
    sample = EvalSample(
        0,
        "plasma",
        "book",
        "",
        "Sheet PLASMA quản lý spare parts cho công đoạn nào?",
        [""],
        sample_path="siflex",
    )

    assert pipeline.filter_samples([sample]) == [sample]


def test_table_agent_separates_structure_caches_by_dataset(tmp_path: Path):
    cache_dir = tmp_path / "structure-cache"
    common = {
        "artifact_dir": str(tmp_path / "artifacts"),
        "structure_cache_dir": str(cache_dir),
        "phase": "structure",
    }
    siflex = TableAgentPipeline(
        llm_client=None,
        layout_vlm_client=FakeLayoutVLM(),
        config={**common, "cache_namespace": "siflex"},
    )
    realhitbench = TableAgentPipeline(
        llm_client=None,
        layout_vlm_client=FakeLayoutVLM(),
        config={**common, "cache_namespace": "realhitbench"},
    )

    assert siflex.structure_cache.root == cache_dir / "v5" / "datasets" / "siflex"
    assert realhitbench.structure_cache.root == cache_dir / "v5" / "datasets" / "realhitbench"
    assert siflex.structure_cache.root != realhitbench.structure_cache.root


def test_table_agent_default_outputs_stay_under_table_agent_outputs():
    from TableAgent.configs import resolve_table_agent_run_roots

    config = {
        "table_agent": {
            "evaluation_output_dir": "TableAgent/outputs/evaluations",
            "log_dir": "TableAgent/outputs/logs",
        }
    }

    output_dir, log_dir = resolve_table_agent_run_roots("table_agent", "outputs", config)

    assert output_dir == Path("TableAgent") / "outputs"
    assert log_dir == Path("TableAgent") / "outputs"
    assert resolve_table_agent_run_roots("table_agent", "custom", config)[0] == Path("custom")
    assert resolve_table_agent_run_roots("graphotter", "outputs", config) == (Path("outputs"), Path("logs"))


def legacy_table_agent_retrieval_rejects_placeholder_structures(tmp_path: Path):
    import openpyxl
    # Create two dummy workbooks
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "SheetA"
    ws1["A1"] = "Apple"
    path_a = tmp_path / "doc_a.xlsx"
    wb1.save(path_a)

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "SheetB"
    ws2["A1"] = "Orange"
    path_b = tmp_path / "doc_b.xlsx"
    wb2.save(path_b)

    sample = EvalSample(
        index=0,
        sample_id="siflex/1",
        table_id="siflex-table",
        table_content="Apple | Banana\nOrange | Grape",
        question="Which sheet has the Apple and Banana?",
        answer=["SheetA"],
        sample_path="data/SiFlex/golden_tests/compiled/golden_cases.json:cases[0]",
        table_path=f"{path_a};{path_b}",
    )

    llm = FakeLLM()
    llm.generate_with_image = lambda prompt, image_path, system_prompt=None: LLMResponse(
        content="SheetA", prompt_tokens=15, completion_tokens=3
    )

    layout_vlm = FakeLayoutVLM()

    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=layout_vlm,
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 1},
    )

    # Pre-encode sheets
    pipeline.prepare_samples([sample])

    # Now manually overwrite structure.yaml of SheetA and SheetB with placeholder/error/invalid structures
    sources_dir = tmp_path / "sources"
    struct_a_path = sources_dir / "doc_a.xlsx_SheetA" / "structure.yaml"
    struct_b_path = sources_dir / "doc_b.xlsx_SheetB" / "structure.yaml"
    
    # Overwrite both structure.yaml with placeholder structures
    struct_a_path.write_text("headers: []\nerror: Failed to generate structure", encoding="utf-8")
    struct_b_path.write_text("headers: []\nerror: Failed to generate structure", encoding="utf-8")

    # Run the pipeline - it should skip both since their structures are invalid (rejected)
    output = pipeline.run(sample)
    
    # Verify that it didn't retrieve either (status won't be good or verification metadata won't say retrieved)
    assert output.metadata["verification"]["feedback"] != "Retrieved from encoded source"


def legacy_table_agent_prepare_samples_uses_error_sidecar(tmp_path: Path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SheetZ"
    ws["A1"] = "Data"
    path_z = tmp_path / "doc_z.xlsx"
    wb.save(path_z)

    sample = EvalSample(
        index=0,
        sample_id="siflex/sidecar_test",
        table_id="siflex-table-sidecar",
        table_content="Data",
        question="What is Data?",
        answer=["Data"],
        sample_path="data/SiFlex/golden_tests/compiled/golden_cases.json:cases[0]",
        table_path=str(path_z),
    )

    # Mock VLM layout client that returns invalid/empty layout to simulate failure
    class ErrorLayoutVLM:
        model_name = "error-vlm"
        temperature = 0.0
        def generate_with_image(self, prompt: str, image_path: Path, system_prompt: str | None = None) -> LLMResponse:
            return LLMResponse(content="", prompt_tokens=0, completion_tokens=0)

    llm = FakeLLM()

    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=ErrorLayoutVLM(),
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 1},
    )

    # First prepare_samples call: should attempt generation, fail (VLM returns empty),
    # write structure.error, and not write structure.yaml
    pipeline.prepare_samples([sample])

    sheet_dir = tmp_path / "sources" / "doc_z.xlsx_SheetZ"
    error_path = sheet_dir / "structure.error"
    struct_path = sheet_dir / "structure.yaml"

    assert error_path.is_file()
    assert not struct_path.is_file()

    # Now let's mock a layout VLM that succeeds
    class SuccessLayoutVLM:
        model_name = "success-vlm"
        temperature = 0.0
        def generate_with_image(self, prompt: str, image_path: Path, system_prompt: str | None = None) -> LLMResponse:
            # Return a valid structure
            return LLMResponse(
                content="headers:\n  - label: SuccessHeader",
                prompt_tokens=5,
                completion_tokens=5,
            )

    pipeline_success = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=SuccessLayoutVLM(),
        config={"artifact_dir": str(tmp_path), "max_refinement_rounds": 1},
    )

    # Run prepare_samples again. Since structure.error exists, it should NOT regenerate,
    # and structure.yaml should still NOT exist.
    pipeline_success.prepare_samples([sample])
    assert not struct_path.is_file()

    # If we delete structure.error, then running prepare_samples should regenerate and succeed
    error_path.unlink()
    pipeline_success.prepare_samples([sample])
    assert struct_path.is_file()
    assert "SuccessHeader" in struct_path.read_text(encoding="utf-8")


def test_table_agent_image_dimension_config_and_render_kwargs(tmp_path: Path):
    sample = EvalSample(
        index=0,
        sample_id="sample/1",
        table_id="table-1",
        table_content="Year | Revenue\n2024 | 100",
        question="What is the 2024 revenue?",
        answer=["100"],
    )
    llm = FakeLLM()
    layout_vlm = FakeLayoutVLM()
    
    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=layout_vlm,
        config={
            "artifact_dir": str(tmp_path),
            "max_refinement_rounds": 1,
            "max_image_dimension": 200,
            "libreoffice_image_resolution": 240,
        },
    )
    
    # Expose configs check
    cfg = pipeline.get_config()
    assert cfg["agent"]["max_image_dimension"] == 200
    assert cfg["agent"]["libreoffice_image_resolution"] == 240
    
    # Run the pipeline
    output = pipeline.run(sample)
    
    assert Path(output.metadata["image_path"]).is_file()


def test_table_agent_image_resizing_and_tiling(tmp_path: Path):
    from PIL import Image
    from TableAgent.rendering.image_utils import (
        _generate_image_tiles,
        _resize_image_file_to_fit,
    )

    large_img_path = tmp_path / "large_table.png"
    # Create a dummy image of size 1200 x 800
    img = Image.new("RGB", (1200, 800), color="white")
    img.save(large_img_path)

    # Slice into tiles of size 500 with 100 overlap
    tiles = _generate_image_tiles(large_img_path, tile_size=500, overlap=100)
    assert len(tiles) == 6
    for tile in tiles:
        tile_file = tmp_path / tile["filename"]
        assert tile_file.is_file()
        with Image.open(tile_file) as t_img:
            assert t_img.width <= 500
            assert t_img.height <= 500

    # Resize the large image to max_dim 600 and max_pixels 200000
    _resize_image_file_to_fit(large_img_path, max_dim=600, max_pixels=200000)
    with Image.open(large_img_path) as resized_img:
        assert resized_img.width <= 600
        assert resized_img.height <= 400
        assert resized_img.width * resized_img.height <= 200000

    # Test decompression bomb safety.
    bomb_path = tmp_path / "bomb.png"
    try:
        # Create a 10000x9000 (90MP) image in mode "1" (binary, keeps RAM tiny ~11MB)
        bomb_img = Image.new("1", (10000, 9000), color=0)
        bomb_img.save(bomb_path)
        
        # Verify opening it does not crash/throw with our bypass, and it successfully resizes to max_dim 1000
        _resize_image_file_to_fit(bomb_path, max_dim=1000)
        assert bomb_path.is_file()
        Image.MAX_IMAGE_PIXELS = None
        with Image.open(bomb_path) as opened_bomb:
            assert opened_bomb.width == 1000
    except MemoryError:
        pass


def legacy_table_agent_llm_reranker(tmp_path: Path):
    import openpyxl
    # Create two dummy workbooks: doc_a has low lexical score, doc_b has high lexical score
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "LowLexical"
    ws1["A1"] = "Cherry"
    path_a = tmp_path / "doc_a.xlsx"
    wb1.save(path_a)

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "HighLexical"
    ws2["A1"] = "Apple"
    path_b = tmp_path / "doc_b.xlsx"
    wb2.save(path_b)

    sample = EvalSample(
        index=0,
        sample_id="siflex/rerank_test",
        table_id="siflex-table",
        table_content="Apple | Cherry",
        question="Which sheet has the Apple?",
        answer=["SomeSecretGoldAnswer"],  # gold answer is distinct secret string
        sample_path="data/SiFlex/golden_tests/compiled/golden_cases.json:cases[0]",
        table_path=f"{path_a};{path_b}",
    )

    class CustomFakeLLM:
        model_name = "custom-fake"
        temperature = 0.0
        def __init__(self):
            self.calls = []
            self.generate_content = "selected_index: 1\nrationale: LLM chose LowLexical"
        
        def generate(self, prompt: str, system_prompt: str | None = None) -> LLMResponse:
            self.calls.append((prompt, system_prompt))
            if system_prompt and "selection agent" in system_prompt:
                return LLMResponse(content=self.generate_content, prompt_tokens=100, completion_tokens=10)
            return LLMResponse(content="LowLexical", prompt_tokens=15, completion_tokens=3)

    llm = CustomFakeLLM()
    # Mock generate_with_image for answering
    llm.generate_with_image = lambda prompt, image_path, system_prompt=None: LLMResponse(
        content="LowLexical", prompt_tokens=15, completion_tokens=3
    )

    layout_vlm = FakeLayoutVLM()

    pipeline = TableAgentPipeline(
        llm_client=llm,
        layout_vlm_client=layout_vlm,
        config={
            "artifact_dir": str(tmp_path),
            "max_refinement_rounds": 1,
            "retrieval_rerank_with_llm": True,
            "retrieval_top_k": 2,
        },
    )

    # 1. Pre-encode sheets
    pipeline.prepare_samples([sample])

    # 2. Run the QA flow with LLM reranking choosing the lower lexical candidate
    output = pipeline.run(sample)
    
    # Verify we selected LowLexical (from path_a)
    assert output.metadata["workbook_sheets"] == ["LowLexical"]
    assert output.metadata["workbook_path"] == str(path_a.resolve())
    assert output.metadata["retrieval_info"]["reranker_selected_index"] == 1
    assert output.metadata["retrieval_info"]["fallback_used"] is False
    assert output.metadata["retrieval_info"]["reranker_rationale"] == "LLM chose LowLexical"
    
    # Verify the prompt does not contain the gold answer "SomeSecretGoldAnswer"
    rerank_call = [call for call in llm.calls if call[1] and "selection agent" in call[1]]
    assert len(rerank_call) == 1
    rerank_prompt = rerank_call[0][0]
    assert "SomeSecretGoldAnswer" not in rerank_prompt
    assert "gold" not in rerank_prompt

    # Test 2: Invalid/empty reranker output falls back to lexical best (HighLexical/doc_b) and does not crash.
    llm.calls.clear()
    llm.generate_content = "invalid yaml here: index: 9999"
    llm.generate_with_image = lambda prompt, image_path, system_prompt=None: LLMResponse(
        content="HighLexical", prompt_tokens=15, completion_tokens=3
    )
    
    output2 = pipeline.run(sample)
    assert output2.metadata["workbook_sheets"] == ["HighLexical"]
    assert output2.metadata["workbook_path"] == str(path_b.resolve())
    assert output2.metadata["retrieval_info"]["fallback_used"] is True


def test_pure_common_info_answer_bypasses_siflex_formatter():
    draft = "## Sheet: OIL\n\n| Header | Description |\n| --- | --- |\n| Code | Part code |"
    responses = []

    answer = TableAgentPipeline._finalize_siflex_answer(
        None,
        None,
        draft,
        responses,
        {"answer_route": "common_info"},
    )

    assert answer == draft
    assert responses == []


def test_table_agent_default_max_replans_is_five(tmp_path: Path):
    pipeline = TableAgentPipeline(
        llm_client=FakeLLM(),
        layout_vlm_client=FakeLayoutVLM(),
        config={"artifact_dir": str(tmp_path)},
    )

    assert pipeline.settings.qa_max_replans == 5


def test_verified_fallback_uses_only_successful_normal_inspections():
    from types import SimpleNamespace

    from TableAgent.schema.qa import AgentOutput, QAResult
    from TableAgent.schema.subtask import SubTask

    result = QAResult(
        question="Describe the matching record.",
        plan=[
            SubTask(id="route", description="Select a table.", layer="table_inspect"),
            SubTask(id="inspect", description="Read the exact matching row.", layer="inspect"),
            SubTask(
                id="common",
                description="Describe the sheet.",
                layer="inspect",
                category="common_info",
                metadata={"common_info_scope": "sheet"},
            ),
        ],
        subtask_outputs=[
            AgentOutput("route", "", "", True, "selected table", namespace_updates={}),
            AgentOutput("inspect", "", "", True, "error_code=E-1; fault_content=blocked", namespace_updates={}),
            AgentOutput("common", "", "", True, "sheet headers", namespace_updates={}),
        ],
    )
    pipeline = SimpleNamespace(settings=SimpleNamespace(max_context_chars=10000))

    prompt = TableAgentPipeline._verified_observation_fallback_prompt(
        pipeline,
        "Describe error E-1.",
        result,
    )

    assert prompt is not None
    assert "error_code=E-1; fault_content=blocked" in prompt
    assert "selected table" not in prompt
    assert "sheet headers" not in prompt


def test_verified_fallback_rejects_unsafe_evidence():
    from types import SimpleNamespace

    from TableAgent.schema.qa import AgentOutput, QAResult
    from TableAgent.schema.subtask import SubTask

    result = QAResult(
        question="Return the matching record.",
        plan=[SubTask(id="inspect", description="Inspect the requested field.", layer="inspect")],
        subtask_outputs=[
            AgentOutput(
                "inspect",
                "Inspect the requested field.",
                "print(wrong_header_value)",
                True,
                "value=from a neighboring header",
                layer="inspect",
                category="normal",
            )
        ],
        error="Final answer review rejected the plan: incorrect header mapping.",
    )
    pipeline = SimpleNamespace(settings=SimpleNamespace(max_context_chars=10000))

    assert TableAgentPipeline._verified_observation_fallback_prompt(
        pipeline,
        result.question,
        result,
    ) is None
