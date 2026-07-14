from __future__ import annotations

import json
import sys
import threading
import time
import types
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import openpyxl
import yaml
from PIL import Image

from TableAgent.agents import (
    LayoutAgent,
    VerificationAgent,
    _VERIFIER_CODE,
    _execute_verifier,
    _union_existing_data_ranges,
)
from TableAgent.config import TableAgentConfig
from TableAgent.perception.metadata import ExStructMetadataExtractor, SheetMetadata
from TableAgent.pipeline.layout_workflow import TableLayoutWorkflow, _has_enough_data, _range_fully_covered
from TableAgent.pipeline.traversal import Direction, DirectionQueue, TraversalTask, Viewport, corner_viewports
from TableAgent.rendering.workbook import WorkbookRenderer
from TableAgent.rendering.workbook import _render_xlsx_range_with_libreoffice
from table2img.core import RenderResult
from utils.llm.base import LLMResponse


class StaticLayoutVLM:
    model_name = "layout"
    temperature = 0.0

    def __init__(self, header_range: str = "A1:A1"):
        self.header_range = header_range

    def generate_with_image(self, prompt, image_path, system_prompt=None):
        structure = {
            "table1": {
                "name": "Sales",
                "description": "Sales table",
                "headers": [{
                    "label": "Region",
                    "description": "Sales region",
                    "orientation": "column",
                    "header_range": self.header_range,
                    "data_range": "A2:A10",
                    "sub_headers": [],
                }],
            }
        }
        return LLMResponse(content=yaml.safe_dump({
            "structure": structure,
            "changelog": "Added the Region header.",
            "remaining_directions": [],
        }, sort_keys=False))


class GoodVerificationLLM:
    model_name = "verifier"
    temperature = 0.0

    def generate(self, prompt, system_prompt=None):
        return LLMResponse(content="status: good\nfeedback: Verified.\nnull_fields: []\n")


class SemanticRepairVerificationLLM:
    model_name = "semantic-verifier"
    temperature = 0.0

    def __init__(self, structure: dict, status: str = "good"):
        self.structure = structure
        self.status = status

    def generate(self, prompt, system_prompt=None):
        return LLMResponse(content=yaml.safe_dump({
            "thought": "The deterministic report identifies a range issue.",
            "action": "repair_structure",
            "observation": "Semantic review can correct the persisted structure.",
            "status": self.status,
            "feedback": "Applied semantic structure repair.",
            "null_fields": [],
            "updated_structure": self.structure,
        }, sort_keys=False))


class RecordingRenderer:
    def __init__(self):
        self.ranges = []

    def __call__(self, document, image_path, **kwargs):
        image_path = Path(image_path)
        self.ranges.append(document)
        image_path.write_bytes(b"viewport")
        html_path = image_path.with_suffix(".html")
        html_path.write_text(document.html, encoding="utf-8")
        return RenderResult(image_path, html_path, 100, 80, Path("fake"))


def _patch_libreoffice_workbook_render(monkeypatch):
    def fake_render(workbook_path, sheet_name, cell_range, image_path, **kwargs):
        image_path.parent.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", (100, 80), "white").save(image_path)

    monkeypatch.setattr("TableAgent.rendering.workbook._render_xlsx_range_with_libreoffice", fake_render)


def _settings(tmp_path: Path, **override) -> TableAgentConfig:
    return TableAgentConfig.from_config({
        "artifact_dir": str(tmp_path),
        "viewport_rows": 20,
        "viewport_columns": 20,
        "shift_cells": 15,
        "max_retry": 5,
        **override,
    })


def _workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "Region"
    worksheet["A2"] = "North"
    for row in range(3, 11):
        worksheet[f"A{row}"] = f"Region {row}"
    worksheet["B2"] = 12.5
    worksheet["B2"].number_format = "0.00"
    workbook.save(path)


def test_libreoffice_range_renderer_sets_print_area_and_coordinates(monkeypatch, tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["B2"] = "Revenue"
    worksheet["D5"] = 100
    workbook.save(workbook_path)
    workbook.close()

    image_path = tmp_path / "render.png"
    calls = {}

    monkeypatch.setattr(
        "TableAgent.rendering.workbook._resolve_libreoffice_path",
        lambda path: Path("C:/LibreOffice/program/soffice.exe"),
    )

    def fake_run(command, **kwargs):
        calls["command"] = command
        calls["kwargs"] = kwargs
        outdir = Path(command[command.index("--outdir") + 1])
        prepared = Path(command[-1])
        pdf_path = outdir / prepared.with_suffix(".pdf").name
        pdf_path.write_bytes(b"%PDF-1.4")

        saved = openpyxl.load_workbook(prepared)
        try:
            sheet = saved["Sheet1"]
            calls["print_area"] = str(sheet.print_area)
            calls["headings"] = sheet.print_options.headings
            calls["grid_lines"] = sheet.print_options.gridLines
            calls["fit_to_width"] = sheet.page_setup.fitToWidth
            calls["fit_to_height"] = sheet.page_setup.fitToHeight
        finally:
            saved.close()
        return types.SimpleNamespace(returncode=0, stdout="", stderr="")

    class FakeBitmap:
        def to_pil(self):
            return Image.new("RGB", (100, 80), "white")

        def close(self):
            calls["bitmap_closed"] = True

    class FakePage:
        def render(self, *, scale):
            calls["scale"] = scale
            return FakeBitmap()

        def close(self):
            calls["page_closed"] = True

    class FakePdf:
        def __init__(self, path):
            calls["pdf_path"] = path

        def __getitem__(self, index):
            calls["page_index"] = index
            return FakePage()

        def close(self):
            calls["closed"] = True

    monkeypatch.setattr("subprocess.run", fake_run)
    monkeypatch.setitem(sys.modules, "pypdfium2", types.SimpleNamespace(PdfDocument=FakePdf))

    _render_xlsx_range_with_libreoffice(
        workbook_path,
        "Sheet1",
        "B2:D5",
        image_path,
        libreoffice_path=Path("soffice.exe"),
        resolution=240,
        timeout_seconds=30,
        show_coordinates=True,
    )

    assert image_path.is_file()
    assert calls["command"][0] == "C:\\LibreOffice\\program\\soffice.exe"
    assert calls["print_area"] == "'Sheet1'!$B$2:$D$5"
    assert calls["headings"] is True
    assert calls["grid_lines"] is True
    assert calls["fit_to_width"] == 1
    assert calls["fit_to_height"] == 1
    assert calls["scale"] == 384 / 72
    assert calls["page_index"] == 0
    assert calls["bitmap_closed"] is True
    assert calls["page_closed"] is True
    assert calls["closed"] is True


def test_libreoffice_range_renderer_serializes_pdfium(monkeypatch, tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    _workbook(workbook_path)
    state = {"active": 0, "max_active": 0}
    state_lock = threading.Lock()

    monkeypatch.setattr(
        "TableAgent.rendering.workbook._resolve_libreoffice_path",
        lambda path: Path("C:/LibreOffice/program/soffice.exe"),
    )

    def fake_run(command, **kwargs):
        outdir = Path(command[command.index("--outdir") + 1])
        prepared = Path(command[-1])
        (outdir / prepared.with_suffix(".pdf").name).write_bytes(b"%PDF-1.4")
        return types.SimpleNamespace(returncode=0, stdout="", stderr="")

    class FakeBitmap:
        def to_pil(self):
            return Image.new("RGB", (100, 80), "white")

        def close(self):
            pass

    class FakePage:
        def render(self, *, scale):
            return FakeBitmap()

        def close(self):
            pass

    class FakePdf:
        def __init__(self, path):
            with state_lock:
                state["active"] += 1
                state["max_active"] = max(state["max_active"], state["active"])
            time.sleep(0.05)

        def __getitem__(self, index):
            return FakePage()

        def close(self):
            with state_lock:
                state["active"] -= 1

    monkeypatch.setattr("subprocess.run", fake_run)
    monkeypatch.setitem(sys.modules, "pypdfium2", types.SimpleNamespace(PdfDocument=FakePdf))

    def render(index: int):
        _render_xlsx_range_with_libreoffice(
            workbook_path,
            "Sheet1",
            "A1:B10",
            tmp_path / f"render_{index}.png",
            libreoffice_path=Path("soffice.exe"),
        )

    with ThreadPoolExecutor(max_workers=2) as executor:
        list(executor.map(render, range(2)))

    assert state["max_active"] == 1


def _hierarchical_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "Month"
    worksheet["A2"] = "In"
    worksheet["B2"] = "Out"
    worksheet["A3"] = 1
    worksheet["B3"] = 2
    workbook.save(path)


def _wide_hierarchical_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = "Quarter"
    for cell, value in zip(("A2", "B2", "C2", "D2"), ("Jan", "Feb", "Mar", "Apr")):
        worksheet[cell] = value
    for col in range(1, 5):
        worksheet.cell(row=3, column=col).value = col
    workbook.save(path)


def _merged_header_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = "Month\nPlan"
    worksheet["A2"] = "North"
    worksheet["B2"] = "South"
    workbook.save(path)


def _adjacent_merged_header_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.merge_cells("A1:B1")
    worksheet.merge_cells("C1:D1")
    worksheet["A1"] = "First"
    worksheet["C1"] = "Second"
    worksheet["C2"] = 1
    worksheet["D2"] = 2
    workbook.save(path)


def _run_verifier(tmp_path: Path, workbook_path: Path, structure: dict) -> dict:
    verifier_path = tmp_path / "verification.py"
    structure_path = tmp_path / "structure_after.yaml"
    verifier_path.write_text(_VERIFIER_CODE, encoding="utf-8")
    structure_path.write_text(yaml.safe_dump(structure, sort_keys=False), encoding="utf-8")
    return _execute_verifier(verifier_path, workbook_path, "Sheet1", structure_path)


def test_direction_queue_uses_required_priority():
    queue = DirectionQueue()
    viewport = Viewport(1, 1, 20, 20)
    assert viewport.clipped_a1_range("A1:B3") == "A1:B3"
    for direction in [Direction.UP, Direction.LEFT, Direction.DOWN, Direction.RIGHT, Direction.STAY]:
        queue.push(TraversalTask(direction, viewport.shifted(direction, 15)))

    assert [queue.pop().direction for _ in range(5)] == [
        Direction.STAY,
        Direction.RIGHT,
        Direction.DOWN,
        Direction.LEFT,
        Direction.UP,
    ]


def test_corner_viewports_cover_used_range_corners_without_duplicates():
    viewports = corner_viewports("A1:AX50", rows=20, columns=20)

    assert [viewport.clipped_a1_range("A1:AX50") for _, viewport in viewports] == [
        "A1:T20",
        "AE1:AX20",
        "A31:T50",
        "AE31:AX50",
    ]
    assert [direction for direction, _ in viewports] == [
        Direction.STAY,
        Direction.RIGHT,
        Direction.DOWN,
        Direction.RIGHT,
    ]


def test_corner_viewports_collapse_when_used_range_fits_one_viewport():
    viewports = corner_viewports("A1:B3", rows=20, columns=20)

    assert [(direction, viewport.clipped_a1_range("A1:B3")) for direction, viewport in viewports] == [
        (Direction.STAY, "A1:B3"),
    ]


def test_range_fully_covered_accepts_union_of_corner_ranges():
    assert _range_fully_covered("A46:AX95", {"A1:AX50", "A51:AX100"})
    assert not _range_fully_covered("A46:AX95", {"A1:AX50"})


def test_exstruct_payload_becomes_metadata_yaml(tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    _workbook(workbook_path)
    payload = {
        "sheets": {
            "Sheet1": {
                "rows": [
                    {"r": 1, "c": {"A": "Region"}},
                    {"r": 2, "c": {"A": "North", "B": 12.5}},
                ],
                "merged_ranges": ["A1:B1"],
            }
        }
    }

    metadata = ExStructMetadataExtractor("light").sheet_metadata(workbook_path, payload, "Sheet1")
    metadata_path = tmp_path / "metadata.yaml"
    metadata_path.write_text(metadata.to_yaml(), encoding="utf-8")

    assert metadata.used_range == "A1:B2"
    assert metadata.merged_ranges == ["A1:B1"]
    assert yaml.safe_load(metadata_path.read_text(encoding="utf-8")) == {
        "sheet_name": "Sheet1",
        "used_range": "A1:B2",
    }


def test_exstruct_metadata_falls_back_to_workbook_merged_ranges(tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    _merged_header_workbook(workbook_path)
    payload = {
        "sheets": {
            "Sheet1": {
                "rows": [
                    {"r": 1, "c": {"A": "Month Plan"}},
                    {"r": 2, "c": {"A": "North", "B": "South"}},
                ],
                "merged_ranges": [],
            }
        }
    }

    metadata = ExStructMetadataExtractor("light").sheet_metadata(workbook_path, payload, "Sheet1")

    assert metadata.used_range == "A1:B2"
    assert metadata.merged_ranges == ["A1:B1"]


def test_verifier_accepts_consistent_header_and_sub_header_ranges(tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    _hierarchical_workbook(workbook_path)
    structure = {
        "table1": {
            "name": "Movement",
            "description": "Movement by month",
            "headers": [{
                "label": "Month",
                "description": "Month group",
                "orientation": "column",
                "header_range": "A1:B1",
                "data_range": "A3:B3",
                "sub_headers": [
                    {
                        "label": "In",
                        "description": "Inbound",
                        "orientation": "column",
                        "header_range": "A2:A2",
                        "data_range": "A3:A3",
                    },
                    {
                        "label": "Out",
                        "description": "Outbound",
                        "orientation": "column",
                        "header_range": "B2:B2",
                        "data_range": "B3:B3",
                    },
                ],
            }],
        }
    }

    report = _run_verifier(tmp_path, workbook_path, structure)

    assert report["status"] == "good"
    assert report["errors"] == []


def test_verifier_rejects_missing_visible_layered_subheaders(tmp_path: Path):
    workbook_path = tmp_path / "wide.xlsx"
    _wide_hierarchical_workbook(workbook_path)
    structure = {
        "table1": {
            "name": "Quarter",
            "description": "Quarterly data",
            "headers": [{
                "label": "Quarter",
                "description": "Quarter group",
                "orientation": "column",
                "header_range": "A1:D1",
                "data_range": "A3:D3",
                "sub_headers": [{
                    "label": "Jan",
                    "description": "January",
                    "orientation": "column",
                    "header_range": "A2:A2",
                    "data_range": "A3:A3",
                }],
            }],
        }
    }

    report = _run_verifier(tmp_path, workbook_path, structure)

    assert report["status"] == "not_good"
    assert any("do not cover visible layered header cells" in error for error in report["errors"])


def test_verifier_rejects_subheader_without_data_range(tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    _hierarchical_workbook(workbook_path)
    structure = {
        "table1": {
            "name": "Movement",
            "description": "Movement by month",
            "headers": [{
                "label": "Month",
                "description": "Month group",
                "orientation": "column",
                "header_range": "A1:B1",
                "data_range": "A3:B3",
                "sub_headers": [{
                    "label": "In",
                    "description": "Inbound",
                    "orientation": "column",
                    "header_range": "A2:A2",
                    "data_range": None,
                }],
            }],
        }
    }

    report = _run_verifier(tmp_path, workbook_path, structure)

    assert report["status"] == "not_good"
    assert "table1.headers[0].sub_headers[0].data_range" in report["null_fields"]


def test_verifier_outputs_unicode_without_windows_codepage_crash(tmp_path: Path):
    workbook_path = tmp_path / "unicode.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "Hạng mục 검사항목"
    worksheet["A2"] = "Open"
    workbook.save(workbook_path)
    structure = {
        "table1": {
            "name": "Unicode",
            "description": "Unicode table",
            "headers": [{
                "label": "Hạng mục 검사항목",
                "description": "Inspection item",
                "orientation": "column",
                "header_range": "A1:A1",
                "data_range": "A2:A2",
                "sub_headers": [],
            }],
        }
    }

    report = _run_verifier(tmp_path, workbook_path, structure)

    assert report["status"] == "good"
    assert not report.get("tool_error")


def test_verifier_rejects_header_text_and_data_range_mismatches(tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    _hierarchical_workbook(workbook_path)
    structure = {
        "table1": {
            "name": "Movement",
            "description": "Movement by month",
            "headers": [{
                "label": "Month",
                "description": "Month group",
                "orientation": "column",
                "header_range": "A1:A2",
                "data_range": "A2:B3",
                "sub_headers": [{
                    "label": "In",
                    "description": "Inbound",
                    "orientation": "column",
                    "header_range": "A2:A2",
                    "data_range": "A3:A3",
                }],
            }],
        }
    }

    report = _run_verifier(tmp_path, workbook_path, structure)

    assert report["status"] == "not_good"
    assert any("header_range contains multiple unrelated texts" in error for error in report["errors"])
    assert "table1.headers[0].data_range" in report["null_fields"]


def test_verifier_repairs_header_label_and_ranges_from_workbook(tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    _merged_header_workbook(workbook_path)
    structure = {
        "table1": {
            "name": "Plan",
            "description": "Plan table",
            "headers": [{
                "label": "Month \\n Plan",
                "description": "Month plan",
                "orientation": "column",
                "header_range": "B1",
                "data_range": "B2:B2",
                "sub_headers": [],
            }],
        }
    }

    report = _run_verifier(tmp_path, workbook_path, structure)
    repaired = yaml.safe_load(report["repaired_structure_yaml"])
    header = repaired["table1"]["headers"][0]

    assert report["status"] == "good"
    assert header["label"] == "Month Plan"
    assert header["header_range"] == "A1:B1"
    assert header["data_range"] == "A2:B2"


def test_verifier_moves_blank_merged_follower_to_next_matching_header(tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    _adjacent_merged_header_workbook(workbook_path)
    structure = {
        "table1": {
            "name": "Adjacent",
            "description": "Adjacent merged headers",
            "headers": [{
                "label": "Second",
                "description": "Second group",
                "orientation": "column",
                "header_range": "B1",
                "data_range": "B2:B2",
                "sub_headers": [],
            }],
        }
    }

    report = _run_verifier(tmp_path, workbook_path, structure)
    repaired = yaml.safe_load(report["repaired_structure_yaml"])
    header = repaired["table1"]["headers"][0]

    assert report["status"] == "good"
    assert header["header_range"] == "C1:D1"
    assert header["data_range"] == "C2:D2"


def test_verification_agent_applies_semantic_updated_structure(tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    _workbook(workbook_path)
    bad_structure = {
        "table1": {
            "name": "Sales",
            "description": "Sales table",
            "headers": [{
                "label": "Region",
                "description": "Sales region",
                "orientation": "column",
                "header_range": "NOT_A_RANGE",
                "data_range": "A2:A10",
                "sub_headers": [],
            }],
        }
    }
    repaired_structure = {
        "table1": {
            "name": "Sales",
            "description": "Sales table",
            "headers": [{
                "label": "Region",
                "description": "Sales region",
                "orientation": "column",
                "header_range": "A1:A1",
                "data_range": "A2:A10",
                "sub_headers": [],
            }],
        }
    }
    iteration_dir = tmp_path / "iteration"
    iteration_dir.mkdir()
    structure_text = yaml.safe_dump(bad_structure, sort_keys=False)
    (iteration_dir / "structure_after.yaml").write_text(structure_text, encoding="utf-8")

    result = VerificationAgent(SemanticRepairVerificationLLM(repaired_structure)).run(
        workbook_path=workbook_path,
        sheet_name="Sheet1",
        metadata_text=SheetMetadata("Sheet1", "A1:A10", []).to_yaml(),
        structure_text=structure_text,
        changelog="Added bad range.",
        viewport_range="A1:T20",
        iteration=1,
        iteration_dir=iteration_dir,
    )

    assert result.status == "good"
    accepted = yaml.safe_load(result.structure_text)
    persisted = yaml.safe_load((iteration_dir / "structure_after.yaml").read_text(encoding="utf-8"))
    assert accepted["table1"]["headers"][0]["header_range"] == "A1:A1"
    assert accepted["table1"]["headers"][0]["data_range"] == "A2:A10"
    assert persisted == accepted
    assert (iteration_dir / "structure_semantic.yaml").is_file()
    assert (iteration_dir / "verification_output_after_semantic.json").is_file()


def test_verification_agent_semantic_update_can_override_strict_deterministic_mismatch(tmp_path: Path):
    workbook_path = tmp_path / "book.xlsx"
    _workbook(workbook_path)
    original_structure = {
        "table1": {
            "name": "Sales",
            "description": "Sales table",
            "headers": [{
                "label": "Region",
                "description": "Sales region",
                "orientation": "column",
                "header_range": "A1:A1",
                "data_range": "A2:A10",
                "sub_headers": [],
            }],
        }
    }
    semantic_structure = {
        "table1": {
            "name": "Sales semantic",
            "description": "Sales table",
            "headers": [{
                "label": "Region",
                "description": "Sales region",
                "orientation": "column",
                "header_range": "A1:A2",
                "data_range": "A3:A10",
                "sub_headers": [],
            }],
        }
    }
    iteration_dir = tmp_path / "semantic-iteration"
    iteration_dir.mkdir()
    structure_text = yaml.safe_dump(original_structure, sort_keys=False)
    (iteration_dir / "structure_after.yaml").write_text(structure_text, encoding="utf-8")

    result = VerificationAgent(SemanticRepairVerificationLLM(semantic_structure)).run(
        workbook_path=workbook_path,
        sheet_name="Sheet1",
        metadata_text=SheetMetadata("Sheet1", "A1:A10", []).to_yaml(),
        structure_text=structure_text,
        changelog="Semantic review adjusted header span.",
        viewport_range="A1:T20",
        iteration=1,
        iteration_dir=iteration_dir,
    )
    after_semantic = yaml.safe_load((iteration_dir / "verification_output_after_semantic.json").read_text(encoding="utf-8"))

    assert after_semantic["status"] == "not_good"
    assert result.status == "good"
    accepted = yaml.safe_load(result.structure_text)
    assert accepted["table1"]["name"] == "Sales semantic"
    assert accepted["table1"]["headers"][0]["header_range"] == "A1:A2"
    assert accepted["table1"]["headers"][0]["data_range"] == "A3:A10"


def test_verification_agent_does_not_accept_semantic_good_on_tool_error(tmp_path: Path, monkeypatch):
    workbook_path = tmp_path / "book.xlsx"
    _workbook(workbook_path)
    structure = {
        "table1": {
            "name": "Sales",
            "description": "Sales table",
            "headers": [{
                "label": "Region",
                "description": "Sales region",
                "orientation": "column",
                "header_range": "A1:A1",
                "data_range": "A2:A10",
                "sub_headers": [],
            }],
        }
    }
    iteration_dir = tmp_path / "tool-error"
    iteration_dir.mkdir()
    structure_text = yaml.safe_dump(structure, sort_keys=False)
    (iteration_dir / "structure_after.yaml").write_text(structure_text, encoding="utf-8")

    def broken_verifier(*args, **kwargs):
        return {
            "status": "not_good",
            "errors": ["Traceback: verifier crashed"],
            "tool_error": True,
            "feedback": "Deterministic verifier tool failed before validating the structure.",
        }

    monkeypatch.setattr("TableAgent.agents._execute_verifier", broken_verifier)

    result = VerificationAgent(SemanticRepairVerificationLLM(structure)).run(
        workbook_path=workbook_path,
        sheet_name="Sheet1",
        metadata_text=SheetMetadata("Sheet1", "A1:A10", []).to_yaml(),
        structure_text=structure_text,
        changelog="No change.",
        viewport_range="A1:T20",
        iteration=1,
        iteration_dir=iteration_dir,
    )

    assert result.status == "not_good"
    assert "tool failed" in result.feedback.lower()


def test_data_range_updates_preserve_union_with_existing_range():
    previous = yaml.safe_dump({
        "table1": {
            "name": "Sales",
            "description": "Sales table",
            "headers": [{
                "label": "Region",
                "description": "Sales region",
                "orientation": "column",
                "header_range": "A1:A1",
                "data_range": "A2:A20",
                "sub_headers": [],
            }],
        }
    }, sort_keys=False)
    updated = yaml.safe_dump({
        "table1": {
            "name": "Sales",
            "description": "Sales table",
            "headers": [{
                "label": "Region",
                "description": "Sales region",
                "orientation": "column",
                "header_range": "A1:A1",
                "data_range": "A16:A35",
                "sub_headers": [],
            }],
        }
    }, sort_keys=False)

    merged = yaml.safe_load(_union_existing_data_ranges(previous, updated))

    assert merged["table1"]["headers"][0]["data_range"] == "A2:A35"


def test_workflow_stops_same_direction_after_good_no_change(tmp_path: Path, monkeypatch):
    _patch_libreoffice_workbook_render(monkeypatch)
    workbook_path = tmp_path / "book.xlsx"
    _workbook(workbook_path)
    settings = _settings(tmp_path)
    recording_renderer = RecordingRenderer()
    renderer = WorkbookRenderer(settings, recording_renderer, logger=None)
    workflow = TableLayoutWorkflow(
        settings,
        renderer,
        LayoutAgent(StaticLayoutVLM()),
        VerificationAgent(GoodVerificationLLM()),
    )
    metadata = SheetMetadata("Sheet1", "A1:AN10", [])

    result = workflow.run(
        workbook_path=workbook_path,
        sheet_name="Sheet1",
        metadata=metadata,
        output_dir=tmp_path / "artifacts",
    )

    events = [json.loads(line) for line in (tmp_path / "artifacts" / "events.jsonl").read_text().splitlines()]
    assert [(event["direction"], event["viewport"]) for event in events] == [
        ("stay", "A1:T10"),
        ("right", "U1:AN10"),
    ]
    assert result.iterations == 2
    assert (tmp_path / "artifacts" / "metadata.yaml").is_file()
    assert (tmp_path / "artifacts" / "changelog.md").is_file()
    for iteration_dir in (tmp_path / "artifacts" / "iterations").iterdir():
        assert (iteration_dir / "viewport.png").is_file()
        assert (iteration_dir / "layout_prompt.txt").is_file()
        assert (iteration_dir / "verification.py").is_file()
        assert (iteration_dir / "verification_output.json").is_file()


def test_workflow_nulls_ranges_after_max_retry(tmp_path: Path, monkeypatch):
    _patch_libreoffice_workbook_render(monkeypatch)
    workbook_path = tmp_path / "book.xlsx"
    _workbook(workbook_path)
    settings = _settings(tmp_path, max_retry=2)
    renderer = WorkbookRenderer(settings, RecordingRenderer(), logger=None)
    workflow = TableLayoutWorkflow(
        settings,
        renderer,
        LayoutAgent(StaticLayoutVLM(header_range="NOT_A_RANGE")),
        VerificationAgent(GoodVerificationLLM()),
    )

    result = workflow.run(
        workbook_path=workbook_path,
        sheet_name="Sheet1",
        metadata=SheetMetadata("Sheet1", "A1:A2", []),
        output_dir=tmp_path / "retry-artifacts",
    )

    structure = yaml.safe_load(result.structure_text)
    header = structure["table1"]["headers"][0]
    assert result.iterations == 2
    assert header["header_range"] is None
    assert header["data_range"] is None
    assert "retries exhausted" in (tmp_path / "retry-artifacts" / "changelog.md").read_text().lower()


def test_workflow_ignores_suggested_direction_outside_used_range(tmp_path: Path, monkeypatch):
    _patch_libreoffice_workbook_render(monkeypatch)
    class RightSuggestingLayoutVLM(StaticLayoutVLM):
        def generate_with_image(self, prompt, image_path, system_prompt=None):
            response = yaml.safe_load(super().generate_with_image(prompt, image_path, system_prompt).content)
            response["remaining_directions"] = ["right"]
            return LLMResponse(content=yaml.safe_dump(response, sort_keys=False))

    workbook_path = tmp_path / "book.xlsx"
    _workbook(workbook_path)
    settings = _settings(tmp_path)
    renderer = WorkbookRenderer(settings, RecordingRenderer(), logger=None)
    workflow = TableLayoutWorkflow(
        settings,
        renderer,
        LayoutAgent(RightSuggestingLayoutVLM()),
        VerificationAgent(GoodVerificationLLM()),
    )

    workflow.run(
        workbook_path=workbook_path,
        sheet_name="Sheet1",
        metadata=SheetMetadata("Sheet1", "A1:Q51", []),
        output_dir=tmp_path / "bounded-artifacts",
    )

    events = [json.loads(line) for line in (tmp_path / "bounded-artifacts" / "events.jsonl").read_text().splitlines()]
    assert [event["viewport"] for event in events] == ["A1:Q20", "A32:Q51"]
    assert all(event["direction"] != "right" for event in events)


def test_has_enough_data_uses_value_coverage_not_styles(tmp_path: Path):
    workbook_path = tmp_path / "coverage.xlsx"
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    for row in range(1, 11):
        for col in range(1, 11):
            worksheet.cell(row=row, column=col).border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin")
            )
    for row in range(1, 6):
        worksheet.cell(row=row, column=1).value = f"Row {row}"
    workbook.save(workbook_path)
    workbook.close()

    assert _has_enough_data(workbook_path, "Sheet1", "A1:J10")
    assert not _has_enough_data(workbook_path, "Sheet1", "B1:J10")


def test_workflow_discards_vlm_suggested_empty_range(tmp_path: Path, monkeypatch):
    _patch_libreoffice_workbook_render(monkeypatch)

    class RightSuggestingLayoutVLM(StaticLayoutVLM):
        def generate_with_image(self, prompt, image_path, system_prompt=None):
            response = yaml.safe_load(super().generate_with_image(prompt, image_path, system_prompt).content)
            response["remaining_directions"] = ["right"]
            return LLMResponse(content=yaml.safe_dump(response, sort_keys=False))

    workbook_path = tmp_path / "book.xlsx"
    _workbook(workbook_path)
    settings = _settings(tmp_path)
    renderer = WorkbookRenderer(settings, RecordingRenderer(), logger=None)
    workflow = TableLayoutWorkflow(
        settings,
        renderer,
        LayoutAgent(RightSuggestingLayoutVLM()),
        VerificationAgent(GoodVerificationLLM()),
    )

    result = workflow.run(
        workbook_path=workbook_path,
        sheet_name="Sheet1",
        metadata=SheetMetadata("Sheet1", "A1:AN10", []),
        output_dir=tmp_path / "empty-right-artifacts",
    )

    events = [json.loads(line) for line in (tmp_path / "empty-right-artifacts" / "events.jsonl").read_text().splitlines()]
    assert [event["viewport"] for event in events] == ["A1:T10", "U1:AN10"]
    assert result.iterations == 2
