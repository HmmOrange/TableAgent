from pathlib import Path

from openpyxl import load_workbook

from datasets.base import EvalSample
from utils.workbook_converter import sample_to_xlsx


def test_html_table_converts_to_xlsx_with_merged_cells(tmp_path: Path):
    sample = EvalSample(
        index=0,
        sample_id="html",
        table_id="html-table",
        table_content=(
            "<table><tr><td rowspan='2'>Metric</td><td colspan='2'>2018</td></tr>"
            "<tr><td>A</td><td>B</td></tr></table>"
        ),
        question="",
        answer=[],
    )

    result = sample_to_xlsx(sample, tmp_path / "table.xlsx")
    workbook = load_workbook(result.path)
    sheet = workbook["html-table"]

    assert sheet["A1"].value == "Metric"
    assert sheet["B1"].value == "2018"
    assert "A1:A2" in {str(item) for item in sheet.merged_cells.ranges}
    assert "B1:C1" in {str(item) for item in sheet.merged_cells.ranges}


def test_realhitbench_latex_swap_converts_to_structured_xlsx(tmp_path: Path):
    sample = EvalSample(
        index=0,
        sample_id="latex",
        table_id="latex-table",
        table_content=r"""
\begin{table}[]
\begin{tabular}{lll}
\multicolumn{3}{c}{Annual results} \\
Metric & \begin{tabular}[c]{@{}l@{}}2022\\ total\end{tabular} & \textbf{2023} \\
Revenue & 10 & 12 \\
\end{tabular}
\end{table}
""",
        question="",
        answer=[],
    )

    result = sample_to_xlsx(sample, tmp_path / "latex.xlsx")
    workbook = load_workbook(result.path)
    sheet = workbook["latex-table"]

    assert result.source_format == "latex"
    assert [sheet.cell(1, column).value for column in range(1, 4)] == ["Annual results", None, None]
    assert [sheet.cell(2, column).value for column in range(1, 4)] == ["Metric", "2022 total", "2023"]
    assert [sheet.cell(3, column).value for column in range(1, 4)] == ["Revenue", "10", "12"]
