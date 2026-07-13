from __future__ import annotations
import os
import datetime
import json
import subprocess
import sys
import pytest
from pathlib import Path

from TableAgent.schema import AxisSelection, Cell, CellRange, Header, ExperienceRecord
from TableAgent.utils import (
    col_name_to_num,
    col_num_to_name,
    parse_a1_cell,
    parse_a1_range,
    cell_to_a1,
    range_to_a1,
    load_table_structures,
)
from TableAgent.environment.qa_env import QAEnvironment
from TableAgent.QA import TableQARunner
from tests.mock_policy import MockActionPolicy
from TableAgent.QA.agents import TableQAPlanner, TableQAAgent
from TableAgent.QA.actions.write_plan import parse_planner_output

# Setup paths
STRUCTURE_PATH = "sample/structure.yaml"
WORKBOOK_PATH = "sample/QA_sample.xlsx"


def _llm_json(data: dict) -> str:
    return "```json\n" + json.dumps(data) + "\n```"


def _two_step_plan_json() -> str:
    return _llm_json({
        "subtasks": [
            {
                "id": "inspect_fields",
                "layer": "inspect",
                "depends_on": [],
                "description": "Inspect the fields needed to answer the question.",
            },
            {
                "id": "synthesize_answer",
                "layer": "synthesis",
                "depends_on": ["inspect_fields"],
                "description": "Use inspected variables to compute final_answer.",
            },
        ],
    })

def test_a1_conversions():
    # column name to number
    assert col_name_to_num("A") == 1
    assert col_name_to_num("Z") == 26
    assert col_name_to_num("AA") == 27
    assert col_name_to_num("AZ") == 52

    # column number to name
    assert col_num_to_name(1) == "A"
    assert col_num_to_name(26) == "Z"
    assert col_num_to_name(27) == "AA"

    # parse cell
    assert parse_a1_cell("B2") == (2, 2)
    assert parse_a1_cell("AA15") == (15, 27)

    # cell to A1
    assert cell_to_a1(2, 2) == "B2"
    assert cell_to_a1(15, 27) == "AA15"

    # parse range
    rng = parse_a1_range("B2:D22", "Sheet1")
    assert rng.start_row == 2
    assert rng.start_col == 2
    assert rng.end_row == 22
    assert rng.end_col == 4
    assert rng.sheet == "Sheet1"

    # single cell range
    rng_single = parse_a1_range("A1")
    assert rng_single.start_row == 1
    assert rng_single.start_col == 1
    assert rng_single.end_row == 1
    assert rng_single.end_col == 1
    
    assert range_to_a1(rng) == "B2:D22"
    assert range_to_a1(rng_single) == "A1"

def test_range_operations():
    # Intersection
    r1 = CellRange(1, 1, 10, 10)
    r2 = CellRange(5, 5, 15, 15)
    intersect = r1.intersection(r2)
    assert intersect == CellRange(5, 5, 10, 10)

    # Crossing (different orientation)
    # Column G (col 7, rows 3 to 22)
    col_g = CellRange(3, 7, 22, 7)
    # Row 5 (row 5, cols 1 to 10)
    row_5 = CellRange(5, 1, 5, 10)
    crossing = col_g.intersection(row_5)
    # The crossing cell should be G5 (row 5, col 7)
    assert crossing == CellRange(5, 7, 5, 7)

    # Union (adjacent columns sharing same row bounds)
    c1 = CellRange(3, 2, 22, 2) # Col B
    c2 = CellRange(3, 3, 22, 3) # Col C
    union_res = c1.union(c2)
    assert len(union_res) == 1
    assert union_res[0] == CellRange(3, 2, 22, 3) # B3:C22

    # Difference (columns same row bounds)
    r_all = CellRange(3, 2, 22, 4) # B3:D22
    r_sub = CellRange(3, 3, 22, 3) # C3:C22
    diff_res = r_all.difference(r_sub)
    assert len(diff_res) == 2
    assert CellRange(3, 2, 22, 2) in diff_res # Col B
    assert CellRange(3, 4, 22, 4) in diff_res # Col D

def test_load_structures():
    structures = load_table_structures(STRUCTURE_PATH)
    assert "table1" in structures
    table = structures["table1"]
    assert table["name"] == "People Nested Headers"
    
    headers = table["headers"]
    # Top-level headers: no, name, date_of, score
    assert len(headers) == 4
    assert headers[0].id == "no"
    assert headers[0].label == "No"
    
    name_hdr = next(h for h in headers if h.id == "name")
    assert name_hdr.orientation == "column_group"
    assert len(name_hdr.sub_headers) == 3
    assert name_hdr.sub_headers[0].id == "first_name"
    assert name_hdr.sub_headers[0].orientation == "column"


def test_load_structures_allows_null_ranges(tmp_path: Path):
    import openpyxl
    import yaml

    workbook_path = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A2"] = "Revenue"
    ws["B2"] = 100
    wb.save(workbook_path)

    structure_path = tmp_path / "structure.yaml"
    structure_path.write_text(
        yaml.safe_dump(
            {
                "table1": {
                    "id": "table1",
                    "name": "Table 1",
                    "description": "Contains one valid header and one null-range header.",
                    "sheet": "Sheet1",
                    "headers": [
                        {
                            "id": "metric",
                            "label": "Metric",
                            "description": "Metric name",
                            "orientation": "column",
                            "header_range": "A1",
                            "data_range": "A2:A2",
                            "sub_headers": [],
                        },
                        {
                            "id": "missing",
                            "label": "Missing",
                            "description": "Unverified sparse field",
                            "orientation": "column",
                            "header_range": None,
                            "data_range": None,
                            "sub_headers": [],
                        },
                    ],
                }
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    env = QAEnvironment(str(structure_path), str(workbook_path))
    headers = env.operators.list_headers("table1")

    assert headers[1].header_range is None
    assert headers[1].data_range is None
    assert env.operators.read_table_as_dataframe("table1", has_headers=False).shape == (2, 1)

def test_environment_and_operators():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)
    
    # Check structures
    assert env.get_table_structure("table1") is not None
    
    # Check sheets
    active_sheet_name = env.get_active_sheet_name()
    assert env.get_sheet(active_sheet_name) is not None

    # Test operators list and find headers
    headers = env.operators.list_headers("table1")
    # no (1), name (1), first (1), middle (1), last (1), date_of (1), birth (1), admission (1), score (1) -> total 9 headers!
    # Let's count leaf headers: no, first, middle, last, birth, admission, score -> total 7 leaf headers.
    # Total headers = 9. Let's verify:
    assert len(headers) == 9

    # Find headers
    score_hdrs = env.operators.find_headers("table1", "score")
    assert len(score_hdrs) > 0
    assert score_hdrs[0].id == "score"

    # Read range
    val = env.operators.read_range("B3:C4")
    assert val == [["Ha", "Minh"], ["An", "Gia"]]

    # Read range flat
    flat_val = env.operators.read_range_flat("B3:C4")
    assert flat_val == ["Ha", "Minh", "An", "Gia"]

    # Basic stats
    scores = env.operators.read_range_flat("G3:G22")
    assert len(scores) == 20
    numeric_scores = [score for score in scores if isinstance(score, (int, float))]
    assert sum(numeric_scores) > 0
    assert 60 <= sum(numeric_scores) / len(numeric_scores) <= 95

def test_filter_operator_returns_axis_selection_for_row_linked_fields():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)
    table_id = env.default_table_id()
    last_hdr = env.operators.get_header(table_id, "last_name")
    birth_hdr = env.operators.get_header(table_id, "birth_date")
    score_hdr = env.operators.get_header(table_id, "score")

    tran_rows = env.operators.filter_values(last_hdr.data_range, contains="Trần", ignore_accents=True)
    birth_month_9_rows = env.operators.filter_values(
        birth_hdr.data_range,
        predicate=lambda value: getattr(value, "month", None) == 9,
    )
    selected_rows = env.operators.selection_intersection(tran_rows, birth_month_9_rows)
    score_ranges = env.operators.project_selection(selected_rows, score_hdr.data_range)
    scores = env.operators.read_selection(selected_rows, score_hdr.data_range)

    assert tran_rows.axis == "row"
    assert tran_rows.positions == (11, 12)
    assert birth_month_9_rows.axis == "row"
    assert selected_rows.axis == "row"
    assert selected_rows.positions == (11,)
    assert [range_to_a1(rng) for rng in score_ranges] == ["G11"]
    assert scores == [86]
    assert sum(scores) == 86

def test_filter_operator_adapts_to_column_axis_for_horizontal_ranges():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)

    selected_cols = env.operators.filter_values("B3:D3", contains="Đặng", ignore_accents=True)
    projected = env.operators.project_selection(selected_cols, "B22:D22")

    assert selected_cols == AxisSelection("col", (4,))
    assert [range_to_a1(rng) for rng in projected] == ["D22"]
    assert env.operators.read_selection(selected_cols, "B22:D22") == ["Nguyen"]

def test_planner_json_depends_on_and_runner_topological_sort():
    plan = parse_planner_output(
        """
        ```json
        {
          "subtasks": [
            {"id": "synthesis", "layer": "synthesis", "depends_on": ["join"], "description": "finish"},
            {"id": "filter", "layer": "inspect", "depends_on": [], "description": "filter rows"},
            {"id": "score", "layer": "inspect", "depends_on": [], "description": "find scores"},
            {"id": "join", "layer": "inspect", "depends_on": ["filter", "score"], "description": "read target scores"}
          ]
        }
        ```
        """
    )
    runner = TableQARunner(STRUCTURE_PATH, WORKBOOK_PATH, policy=MockActionPolicy())
    ordered = runner._topological_sort(plan)

    assert [task.id for task in ordered] == ["filter", "score", "join", "synthesis"]
    assert plan[0].depends_on == ["join"]

def test_topological_sort_invalid_planning():
    runner = TableQARunner(STRUCTURE_PATH, WORKBOOK_PATH, policy=MockActionPolicy())

    # 1. Cycle detection
    cycle_plan = parse_planner_output(
        """
        ```json
        {
          "subtasks": [
            {"id": "task_a", "layer": "inspect", "depends_on": ["task_b"], "description": "A"},
            {"id": "task_b", "layer": "inspect", "depends_on": ["task_a"], "description": "B"}
          ]
        }
        ```
        """
    )
    with pytest.raises(ValueError, match="Cycle detected"):
        runner._topological_sort(cycle_plan)

    # 2. Unknown dependency
    unknown_dep_plan = parse_planner_output(
        """
        ```json
        {
          "subtasks": [
            {"id": "task_a", "layer": "inspect", "depends_on": ["missing_task"], "description": "A"}
          ]
        }
        ```
        """
    )
    with pytest.raises(ValueError, match="depends on unknown subtasks"):
        runner._topological_sort(unknown_dep_plan)

    # 3. Duplicate subtask IDs
    from TableAgent.schema.subtask import SubTask
    dup_plan = [
        SubTask(id="task_a", description="A first", layer="inspect", depends_on=[]),
        SubTask(id="task_a", description="A second", layer="inspect", depends_on=[])
    ]
    with pytest.raises(ValueError, match="Duplicate subtask id"):
        runner._topological_sort(dup_plan)

def test_runner_fails_on_invalid_planning():
    # Setup FakeLLM to return a cyclic plan
    cyclic_plan_json = _llm_json({
        "subtasks": [
            {"id": "task_a", "layer": "inspect", "depends_on": ["task_b"], "description": "A"},
            {"id": "task_b", "layer": "inspect", "depends_on": ["task_a"], "description": "B"}
        ]
    })
    runner = TableQARunner(
        STRUCTURE_PATH,
        WORKBOOK_PATH,
        llm_client=FakeLLM({"Table Structure": cyclic_plan_json}),
        policy=MockActionPolicy(),
    )
    result = runner.run("Dummy question?")
    assert not result.success
    assert "Cycle detected" in result.error


def test_operator_catalog_is_prompt_ready_without_arithmetic_helpers():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)

    catalog = env.operators.operator_catalog()

    assert "operators.list_tables()" in catalog
    assert "operators.read_range" in catalog
    assert "operators.resolve_ranges" in catalog
    assert "operators.filter_values" in catalog
    assert "ignore_accents=True" in catalog
    assert "operators.read_selection" in catalog
    assert "Write normal Python/pandas/numpy code for arithmetic" in catalog
    assert "operators.mean(" not in catalog
    assert "operators.sum(" not in catalog

def test_shared_execution_namespace():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)
    
    # Execute code that assigns variables
    code = (
        "scores = operators.read_range_flat('G3:G7')\n"
        "numeric_scores = [score for score in scores if isinstance(score, (int, float))]\n"
        "mean_score = sum(numeric_scores) / len(numeric_scores) if numeric_scores else 0\n"
    )
    output, error, success, updates = env.execute_code(code)
    
    assert success
    assert not error
    assert "scores" in updates
    assert "mean_score" in updates
    expected_scores = env.operators.read_range_flat('G3:G7')
    expected_numeric_scores = [score for score in expected_scores if isinstance(score, (int, float))]
    assert updates["mean_score"] == sum(expected_numeric_scores) / len(expected_numeric_scores)
    
    # Execute subsequent code referencing the same variables
    code2 = "final_answer = f'Mean is {mean_score:.1f}'"
    output2, error2, success2, updates2 = env.execute_code(code2)
    
    assert success2
    assert not error2
    assert env.execution_namespace.get("final_answer") == f"Mean is {updates['mean_score']:.1f}"

def test_react_loop_and_retry_self_repair():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)
    
    # We will simulate an error on round 1 by passing simulate_error=True
    policy = MockActionPolicy(simulate_error=True)
    agent = TableQAAgent(env, policy=policy, max_retries=3)
    
    from TableAgent.schema.subtask import SubTask
    subtask = SubTask(id="test_subtask", description="Get scores", layer="inspect")
    
    output = agent.run_subtask(question="What is the average score?", subtask=subtask)
    
    # Subtask should succeed because of round 2 self-repair
    assert output.success
    assert subtask.status == "success"
    
    # Experience pool should contain 2 attempts (one failed with score 0.0, one succeeded with score 1.0)
    records = env.experience_pool.records
    assert len(records) >= 2
    assert any(r.score == 0.0 for r in records)
    assert any(r.score == 1.0 for r in records)

    # Check formatting of experience
    formatted_exp = env.experience_pool.format()
    assert "<attempt round=\"1\" subtask=\"test_subtask\">" in formatted_exp
    assert "<attempt round=\"2\" subtask=\"test_subtask\">" in formatted_exp
    assert "Error during execution:" in formatted_exp

def test_full_runner_pipeline():
    # Test 1: Average score question
    runner = TableQARunner(
        STRUCTURE_PATH,
        WORKBOOK_PATH,
        llm_client=FakeLLM({"Table Structure": _two_step_plan_json()}),
        policy=MockActionPolicy(),
    )
    result = runner.run("What is the average score of all people?")
    
    assert result.success
    assert not result.error
    assert len(result.plan) == 2
    assert result.plan[0].layer == "inspect"
    assert result.plan[1].layer == "synthesis"
    
    # Verify average score computation:
    # G3:G22 scores are: 92, 60, 92, 64, 91, 94, 74, 83, ...
    # The average of the sample rows (20 rows) should be evaluated to a float string.
    assert result.final_answer is not None
    # Check that it parses to a valid float
    avg_val = float(result.final_answer)
    assert 75 <= avg_val <= 85

    # Test 2: Birth date question
    runner_birth = TableQARunner(
        STRUCTURE_PATH,
        WORKBOOK_PATH,
        llm_client=FakeLLM({"Table Structure": _two_step_plan_json()}),
        policy=MockActionPolicy(),
    )
    result_birth = runner_birth.run("What is the birth date of An Gia Pham?")
    
    assert result_birth.success
    assert not result_birth.error
    assert result_birth.final_answer == "1995-02-14"


def test_runner_humanizes_header_id_in_final_answer():
    from TableAgent.QA.actions.base_action import CodeGenerationRequest, CodeGenerationResult

    class HeaderIdAnswerPolicy:
        def run(self, request: CodeGenerationRequest) -> CodeGenerationResult:
            if request.layer == "inspect":
                code = "selected_column = table_df.columns[-1]"
            else:
                code = "final_answer = selected_column"
            return CodeGenerationResult(
                code=code,
                description="Simulates a dataframe operation that returns an internal header ID.",
                reasoning="Exercise the user-facing header-label boundary.",
            )

    runner = TableQARunner(
        STRUCTURE_PATH,
        WORKBOOK_PATH,
        llm_client=FakeLLM({"Table Structure": _two_step_plan_json()}),
        policy=HeaderIdAnswerPolicy(),
    )

    result = runner.run("Which field is the last column?")

    assert result.success
    assert result.final_answer == "Score"
    assert runner.env.execution_namespace["final_answer"] == "Score"
    assert runner.env.operators.get_header("table1", "score").label == "Score"

def test_base_abstractions_usable():
    from TableAgent.QA import BaseCodeGenerationAction, BaseReActAgent
    from TableAgent.QA.actions.base_action import CodeGenerationRequest, CodeGenerationResult
    from TableAgent.schema.subtask import SubTask
    from TableAgent.schema.qa import AgentOutput
    
    # 1. Test subclassing BaseCodeGenerationAction
    class CustomPolicy(BaseCodeGenerationAction):
        def run(self, request: CodeGenerationRequest) -> CodeGenerationResult:
            return CodeGenerationResult(
                code="print('hello')",
                description="custom action description",
            )

    policy = CustomPolicy()
    code, desc = policy.generate("test?", "subtask_1", "inspect", 1)
    assert code == "print('hello')"
    assert desc == "custom action description"

    # 2. Test subclassing BaseReActAgent
    class CustomAgent(BaseReActAgent):
        def run_subtask(self, question: str, subtask: SubTask) -> AgentOutput:
            code, desc = self.code_action.generate(question, subtask.id, subtask.layer, 1)
            return AgentOutput(
                subtask_id=subtask.id,
                description=desc,
                code=code,
                success=True,
                observation="done",
                namespace_updates={}
            )

    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)
    agent = CustomAgent(env, policy=policy)
    assert agent.env is env
    assert agent.policy is policy
    assert agent.max_retries == 3

    subtask = SubTask(id="sub_1", description="dummy", layer="inspect")
    output = agent.run_subtask("hello", subtask)
    assert output.success
    assert output.code == "print('hello')"
    assert output.observation == "done"

class FakeLLM:
    def __init__(self, responses: dict[str, str]):
        self.responses = responses
        self.calls = []

    def generate(self, prompt: str, system_prompt: str = None) -> Any:
        from utils.llm.base import LLMResponse
        self.calls.append((prompt, system_prompt))
        if "Review whether this attempt" in prompt:
            return LLMResponse(content=_llm_json({
                "accepted": True,
                "score": 1.0,
                "feedback": "Accepted.",
            }))
        for key, val in self.responses.items():
            if key in prompt:
                return LLMResponse(content=val)
        return LLMResponse(content=_llm_json({
            "reasoning": "Default code response.",
            "code": "pass",
            "description": "Default no-op code.",
        }))

def test_notebook_restrictions_and_persistence():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)
    
    # 1. Verify restricted imports fail
    code_bad = "import os\nprint(os.getcwd())"
    out, error, success, updates = env.execute_code(code_bad)
    assert not success
    assert "restricted" in error

    # 2. Verify allowed imports succeed
    code_good = "import math\nval = math.sqrt(16)"
    out, error, success, updates = env.execute_code(code_good)
    assert success
    assert not error
    assert updates.get("val") == 4.0

    # 3. Verify namespace persistence
    code_persisted = "val_check = val * 2"
    out2, error2, success2, updates2 = env.execute_code(code_persisted)
    assert success2
    assert updates2.get("val_check") == 8.0

def test_operators_generic():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)
    # Check listing headers for table1
    headers = env.operators.list_headers("table1")
    assert len(headers) > 0
    # Check listing headers for non-existent table id returns empty list (generic behavior, doesn't crash)
    headers_empty = env.operators.list_headers("non_existent_table")
    assert len(headers_empty) == 0

def test_runner_with_fake_llm_and_logging():
    # Setup FakeLLM responses
    responses = {
        # Planner prompt matches
        "Table Structure": _two_step_plan_json(),
        # Inspect prompt matches
        "Assigned Subtask:": _llm_json({
            "reasoning": "Inspect the first-name field as a small targeted check.",
            "code": "table_id = env.default_table_id()\nfirst_hdr = operators.get_header(table_id, 'first_name')\nfirsts = operators.read_range_flat(first_hdr.data_range)",
            "description": "Reads first names into the notebook namespace.",
        }),
        # Synthesis prompt matches
        "Variables in namespace:": _llm_json({
            "reasoning": "The test expects a fixed final answer.",
            "code": "final_answer = '82.5'",
            "description": "Sets the final answer for the test.",
        })


    }
    fake_llm = FakeLLM(responses)
    
    runner = TableQARunner(STRUCTURE_PATH, WORKBOOK_PATH, llm_client=fake_llm)
    result = runner.run("What is the average score?")
    
    assert result.success
    assert result.final_answer == "82.5"
    assert len(result.plan) == 2
    assert any("operators.list_headers" in (system_prompt or "") for _, system_prompt in fake_llm.calls)
    assert all("operators.mean(" not in (system_prompt or "") for _, system_prompt in fake_llm.calls)
    
    # Verify logs/events
    logs = result.logs
    assert len(logs) > 0
    
    # Check that logs contain key info: question, subtask input, generated code, observations, final answer
    log_types = [event.get("event_type") for event in logs]
    assert "run_start" in log_types
    assert "planning_start" in log_types
    assert "subtask_start" in log_types
    assert "execute_code" in log_types
    assert "subtask_complete" in log_types
    assert "run_complete" in log_types
    
    # Verify events content
    start_event = next(e for e in logs if e.get("event_type") == "run_start")
    assert start_event["question"] == "What is the average score?"
    
    complete_event = next(e for e in logs if e.get("event_type") == "run_complete")
    assert complete_event["final_answer"] == "82.5"


def test_runner_persists_per_run_artifacts(tmp_path):
    runner = TableQARunner(
        STRUCTURE_PATH,
        WORKBOOK_PATH,
        llm_client=FakeLLM({"Table Structure": _two_step_plan_json()}),
        policy=MockActionPolicy(),
        config={"table_agent": {"qa_log_path": str(tmp_path / "qa_runner.log")}},
    )

    result = runner.run("What is the average score?")

    assert result.success
    artifacts = result.artifacts
    assert artifacts["run_dir"].startswith(str(tmp_path / "qa_runs"))
    assert os.path.exists(artifacts["events_jsonl"])
    assert os.path.exists(artifacts["plan_json"])
    assert os.path.exists(artifacts["result_json"])
    assert os.path.exists(artifacts["answer_py"])
    assert os.path.exists(artifacts["notebook_ipynb"])
    assert os.path.isdir(artifacts["cells_dir"])
    assert os.listdir(artifacts["cells_dir"])

    with open(artifacts["answer_py"], "r", encoding="utf-8") as f:
        assert "final_answer" in f.read()


def test_llm_code_generation_repairs_invalid_json_response():
    from TableAgent.QA.actions.base_action import CodeGenerationRequest
    from TableAgent.QA.actions.llm_code_generation import LLMCodeGenerationAction
    from TableAgent.schema.subtask import SubTask
    from utils.llm.base import LLMResponse

    class RepairingLLM:
        def __init__(self):
            self.calls = []

        def generate(self, prompt: str, system_prompt: str = None) -> Any:
            self.calls.append((prompt, system_prompt))
            if "previous response for a TableAgent code-generation action was invalid" in prompt:
                return LLMResponse(content=_llm_json({
                    "reasoning": "Repair the invalid prose by producing a compact executable inspection cell.",
                    "code": "table_id = env.default_table_id()\nprint(table_id)",
                    "description": "Prints the default table id.",
                }))
            return LLMResponse(content="We need to inspect the table first. Let's list tables.")

    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)
    subtask = SubTask(
        id="inspect_table",
        description="Inspect the default table id.",
        layer="inspect",
        metadata={"table_id": env.default_table_id()},
    )
    action = LLMCodeGenerationAction(RepairingLLM(), env=env, output_format_retries=1)

    result = action.run(CodeGenerationRequest(
        question="What table is loaded?",
        subtask_id=subtask.id,
        layer=subtask.layer,
        round_num=1,
        subtask=subtask,
    ))

    assert "env.default_table_id()" in result.code
    event_types = [event["event_type"] for event in env.logger.events]
    assert "generate_parse_error" in event_types
    assert "generate_repair_call" in event_types
    assert "generate_repair_parsed" in event_types


def test_security_import_bypass_fails():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)

    # Proving __builtins__['__import__']('os') fails
    code_bypass = "func = __builtins__['__import__']\nfunc('os')"
    out, error, success, updates = env.execute_code(code_bypass)
    assert not success
    assert "restricted" in error or "Access to" in error

    # Proving import os fails
    code_import_os = "import os"
    out2, error2, success2, updates2 = env.execute_code(code_import_os)
    assert not success2
    assert "restricted" in error2

    # Proving getattr(__builtins__, '__import__')('os') fails
    code_getattr_bypass = "getattr(__builtins__, '__import__')('os')"
    out3, error3, success3, updates3 = env.execute_code(code_getattr_bypass)
    assert not success3
    assert "restricted" in error3 or "Access to" in error3

    # Proving no leaked variables in namespace
    assert "os" not in env.execution_namespace

def test_allowed_imports_work():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)

    # math
    out, error, success, updates = env.execute_code("import math\nres = math.sqrt(25)")
    assert success
    assert not error
    assert updates.get("res") == 5.0

    # pandas
    out, error, success, updates = env.execute_code("import pandas as pd\ndf = pd.DataFrame({'a': [1, 2]})")
    assert success
    assert not error

    # datetime
    out, error, success, updates = env.execute_code("import datetime\nd = datetime.date(2026, 6, 26)")
    assert success
    assert not error

    # time
    out, error, success, updates = env.execute_code("import time\ntime.sleep(0.001)")
    assert success
    assert not error

def test_runner_with_non_default_table_id(tmp_path):
    # Read sample structure and replace table1 with people_table
    with open("sample/structure.yaml", "r") as f:
        struct_content = f.read()
    
    struct_content = struct_content.replace("table1:", "people_table:")
    
    struct_file = tmp_path / "structure_temp.yaml"
    with open(struct_file, "w") as f:
        f.write(struct_content)

    # Initialize runner with temp structure and check generic table_id selection
    runner = TableQARunner(
        str(struct_file),
        WORKBOOK_PATH,
        llm_client=FakeLLM({"Table Structure": _two_step_plan_json()}),
        policy=MockActionPolicy(),
    )
    assert runner.table_id is None
    
    # Run a question and make sure it plans and resolves table to "people_table"
    result = runner.run("What is the average score?")
    assert result.success
    # The default plan should resolve table_id to "people_table"
    assert result.plan[0].metadata["table_id"] == "people_table"
    assert runner.env.default_table_id() == "people_table"

def test_no_table1_fallback_in_production_code():
    import pathlib
    # Check all production code under TableAgent/QA
    qa_dir = pathlib.Path("TableAgent/QA")
    
    # Check planner.py, runner.py, code-generation action
    files_to_check = [
        qa_dir / "agents/planner.py",
        qa_dir / "runner.py",
        qa_dir / "actions/llm_code_generation.py",
        qa_dir / "actions/write_plan.py",
    ]
    
    for file_path in files_to_check:
        with open(file_path, "r") as f:
            content = f.read()
            # Assert "table1" is not hard-coded in the source code as a literal
            assert "table1" not in content, f"Production file {file_path} contains hardcoded 'table1'"
            # Assert sample header ids are not hard-coded in fallback
            if "planner.py" in str(file_path):
                assert "first_name" not in content, "planner.py contains hardcoded first_name"
                assert "middle_name" not in content, "planner.py contains hardcoded middle_name"
                assert "last_name" not in content, "planner.py contains hardcoded last_name"
                assert "birth_date" not in content, "planner.py contains hardcoded birth_date"


def test_large_output_is_compacted_but_full_output_is_recoverable():
    env = QAEnvironment(
        STRUCTURE_PATH,
        WORKBOOK_PATH,
        max_observation_chars=120,
        max_error_chars=120,
        max_value_repr_chars=120,
    )

    code = "big_text = '0123456789' * 100\nprint(big_text)"
    output, error, success, updates = env.execute_code(code, cell_id="large_output")

    assert success
    assert not error
    assert len(output) < 200
    assert "[truncated]" in output
    assert "big_text" in updates

    full_output = env.get_cell_output("large_output")
    assert len(full_output) >= 1000
    assert full_output.startswith("0123456789")

    history = env.get_history(last_n=1, max_output_len=80)
    assert "large_output" in history
    assert "[truncated]" in history


def test_notebook_records_nbformat_cells(tmp_path):
    pytest.importorskip("nbformat")
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH)

    output, error, success, updates = env.execute_code("answer_preview = 42\nprint(answer_preview)", cell_id="nb_cell")

    assert success
    assert not error
    assert output.strip() == "42"
    assert env.notebook.nb is not None
    assert env.notebook.nb.cells[-1].source == "answer_preview = 42\nprint(answer_preview)"
    assert env.notebook.nb.cells[-1].metadata["cell_id"] == "nb_cell"

    notebook_path = env.export_notebook(tmp_path / "qa_run.ipynb")
    assert notebook_path.exists()


def test_variable_preview_summarizes_large_values():
    env = QAEnvironment(STRUCTURE_PATH, WORKBOOK_PATH, max_value_repr_chars=200)
    code = "import pandas as pd\ndf_large = pd.DataFrame({'a': range(50), 'b': range(50, 100)})"
    output, error, success, updates = env.execute_code(code)

    assert success
    assert not error
    assert "DataFrame(shape=(50, 2)" in env.notebook.summarize_value(env.execution_namespace["df_large"])

    preview = env.preview_variable("df_large", rows=3, max_chars=300)
    assert "DataFrame(shape=(50, 2)" in preview
    assert "0" in preview
    assert len(preview) <= 330


def test_experience_format_truncates_large_observations():
    from TableAgent.schema.experience import ExperiencePool, ExperienceRecord

    pool = ExperiencePool(max_records=2, max_code_chars=80, max_observation_chars=80)
    pool.add(ExperienceRecord(
        subtask_id="inspect",
        description="large observation",
        code="x = 1\n" + ("#" * 200),
        observation="obs-" + ("y" * 200),
        reasoning="I should inspect only the relevant field before reading values.",
        score=1.0,
        round=1,
    ))

    formatted = pool.format()
    assert "large observation" in formatted
    assert "<reasoning>" in formatted
    assert "inspect only the relevant field" in formatted
    assert "[truncated]" in formatted
    assert len(formatted) < 400


def test_operator_modules_have_runnable_smoke_entrypoints():
    modules = [
        "TableAgent.QA.operators.base_operator",
        "TableAgent.QA.operators.range_operator",
        "TableAgent.QA.operators.filter_operator",
        "TableAgent.QA.operators.structure_operator",
        "TableAgent.QA.operators.workbook_operator",
        "TableAgent.QA.operators.table_operator",
    ]

    for module in modules:
        result = subprocess.run(
            [sys.executable, "-m", module],
            cwd=os.getcwd(),
            text=True,
            capture_output=True,
            timeout=20,
            check=False,
        )
        assert result.returncode == 0, f"{module} failed:\nSTDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        assert result.stdout.strip()
