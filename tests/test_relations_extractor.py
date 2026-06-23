import pytest
from pathlib import Path
import openpyxl
import yaml
from TableAgent.perception.relations import extract_relations

def test_relations_extraction(tmp_path):
    # 1. Create temporary Excel workbook
    xlsx_path = tmp_path / "test_workbook.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Fill in some dummy headers
    ws["A1"] = "Item"
    ws["B1"] = "Measure"
    ws["C1"] = "Calculated"
    ws["D1"] = "Input 1"
    ws["E1"] = "Input 2"
    ws["F1"] = "Row Sum"
    ws["G1"] = "Status"

    # Fill in some data
    # Row 2
    ws["A2"] = "A"
    ws["B2"] = 12
    ws["C2"] = "=B2*$H$1"
    ws["D2"] = 5
    ws["E2"] = 10
    ws["F2"] = "=SUM(D2:E2)"
    ws["G2"] = '=IF(B2>10,"NG","OK")'

    # Row 3
    ws["A3"] = "B"
    ws["B3"] = 8
    ws["C3"] = "=B3*$H$1"
    ws["D3"] = 4
    ws["E3"] = 6
    ws["F3"] = "=SUM(D3:E3)"
    ws["G3"] = "=#REF!+B3"

    # Row 4
    ws["A4"] = "C"
    ws["B4"] = 15
    ws["C4"] = "=B4*$H$1"
    ws["D4"] = 3
    ws["E4"] = 7
    ws["F4"] = "=SUM(D4:E4)"

    # Constant cell
    ws["H1"] = 1.5

    # Row 8 - Column repetition
    ws["D8"] = "=D2-D3"
    ws["E8"] = "=E2-E3"
    ws["F8"] = "=F2-F3"

    # Row 9 - Standalone aggregate SUM
    ws["B9"] = "=SUM(B2:B4)"

    wb.save(xlsx_path)
    wb.close()

    # 2. Create tables mapping structure file
    structure_path_multi = tmp_path / "structure_multi.yaml"
    structure_multi = {
        "table1": {
            "name": "Test Table",
            "sheet": "Sheet1",
            "headers": [
                {"id": "item", "data_range": "A2:A4"},
                {"id": "measure_value", "data_range": "B2:B4"},
                {"id": "calculated_value", "data_range": "C2:C4"},
                {
                    "id": "input_group",
                    "data_range": "D2:E4",
                    "sub_headers": [
                        {"id": "input_1", "data_range": "D2:D4"},
                        {"id": "input_2", "data_range": "E2:E4"}
                    ]
                },
                {"id": "row_sum", "data_range": "F2:F4"},
                {"id": "summary_row", "data_range": "D8:F8"},
                {"id": "status", "data_range": "G2:G3"}
            ]
        }
    }
    with open(structure_path_multi, "w", encoding="utf-8") as f:
        yaml.safe_dump(structure_multi, f)

    output_path_multi = tmp_path / "relations_multi.yaml"

    # Run extractor on tables mapping structure
    extract_relations(xlsx_path, structure_path_multi, output_path_multi)

    # Assert output exists
    assert output_path_multi.exists()
    with open(output_path_multi, "r", encoding="utf-8") as f:
        relations_multi = yaml.safe_load(f)

    # Check relations tables mapping
    assert "table1" in relations_multi
    table1 = relations_multi["table1"]

    assert "normal_formulas" in table1
    assert "aggregate_formulas" in table1
    assert "cell_formulas" in table1
    assert "invalid_formulas" in table1

    normal = table1["normal_formulas"]
    assert len(normal) == 3

    # Check that C2:C4 is captured
    c_rep = next(n for n in normal if n["range"] == "C2:C4")
    assert c_rep["expression"] == "calculated_value = measure_value * $H$1"
    assert c_rep["pattern"] == "C{row} = B{row} * $H$1"
    assert c_rep["agg_function"] is None

    # Check that F2:F4 is captured (row SUM as normal)
    f_rep = next(n for n in normal if n["range"] == "F2:F4")
    assert f_rep["expression"] == "row_sum = SUM(input_group)"
    assert f_rep["pattern"] == "F{row} = SUM(D{row}:E{row})"
    assert f_rep["agg_function"] == "SUM"

    # Check that D8:F8 is captured (col repetition)
    col_rep = next(n for n in normal if n["range"] == "D8:F8")
    assert col_rep["pattern"] == "{col}8 = {col}2 - {col}3"
    assert col_rep["agg_function"] is None

    # Verify standalone aggregate SUM (B9)
    agg = table1["aggregate_formulas"]
    assert len(agg) == 1
    assert agg[0]["formula"]["cell"] == "B9"
    assert agg[0]["expression"] == "total_measure_value = SUM(measure_value)"

    # Verify standalone IF (G2)
    cell_f = table1["cell_formulas"]
    assert len(cell_f) == 1
    assert cell_f[0]["formula"]["cell"] == "G2"
    assert cell_f[0]["expression"] == "status = IF(measure_value > 10, 'NG', 'OK')"

    # Verify invalid #REF! (G3)
    invalid = table1["invalid_formulas"]
    assert len(invalid) == 1
    assert invalid[0]["formula"]["cell"] == "G3"
    assert invalid[0]["expression"] is None
    assert len(invalid[0]["errors"]) == 1
    assert invalid[0]["errors"][0]["type"] == "broken_reference"
    assert invalid[0]["errors"][0]["token"] == "#REF!"

    # 3. Test single-table structure shape
    structure_path_single = tmp_path / "structure_single.yaml"
    structure_single = structure_multi["table1"]
    with open(structure_path_single, "w", encoding="utf-8") as f:
        yaml.safe_dump(structure_single, f)

    output_path_single = tmp_path / "relations_single.yaml"

    # Run extractor on single table shape
    extract_relations(xlsx_path, structure_path_single, output_path_single)

    assert output_path_single.exists()
    with open(output_path_single, "r", encoding="utf-8") as f:
        relations_single = yaml.safe_load(f)

    # The single table output should not have a table1 key at root, but directly the categories
    assert "normal_formulas" in relations_single
    assert "aggregate_formulas" in relations_single
    assert "cell_formulas" in relations_single
    assert "invalid_formulas" in relations_single
    assert len(relations_single["normal_formulas"]) == 3


def test_new_relations_requirements(tmp_path):
    import subprocess
    import sys

    # 1. Create workbook with multiple sheets
    xlsx_path = tmp_path / "test_new_requirements.xlsx"
    wb = openpyxl.Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = 10
    ws1["A2"] = 20
    ws1["A3"] = "=SUBTOTAL(9, A1:A2)"

    ws1["B1"] = 15
    ws1["B2"] = 25
    ws1["B3"] = "=SUBTOTAL(9, B1:B2)"

    # C1 is division by zero error
    ws1["C1"] = "=#DIV/0!"

    # D2 is wrong value type error, within side margin of table1
    ws1["D2"] = "=#VALUE! + 1"

    # Sheet 2 - unrelated sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "=SUM(A2:A3)"

    wb.save(xlsx_path)
    wb.close()

    # 2. Create structure
    structure_path = tmp_path / "structure_new.yaml"
    structure = {
        "table1": {
            "name": "Table 1",
            "sheet": "Sheet1",
            "headers": [
                {"id": "col_a", "data_range": "A1:A2"},
                {"id": "col_b", "data_range": "B1:B2"},
                {"id": "subtotal_row", "data_range": "A3:B3"},
                {"id": "err_col", "data_range": "C1:C1"}
            ]
        },
        "table_nonexistent": {
            "name": "Nonexistent Table",
            "sheet": "NonexistentSheet",
            "headers": [
                {"id": "col_a", "data_range": "A1:A2"}
            ]
        }
    }
    with open(structure_path, "w", encoding="utf-8") as f:
        yaml.safe_dump(structure, f)

    output_path = tmp_path / "relations_new.yaml"

    # 3. Run extract_relations
    extract_relations(xlsx_path, structure_path, output_path)

    assert output_path.exists()
    with open(output_path, "r", encoding="utf-8") as f:
        relations = yaml.safe_load(f)

    # Check table1
    assert "table1" in relations
    table1 = relations["table1"]

    # Verify SUBTOTAL in repeated normal formula metadata (A3:B3 is repeated row-wise subtotal)
    normal = table1["normal_formulas"]
    subtotal_rep = next((n for n in normal if n["range"] == "A3:B3"), None)
    assert subtotal_rep is not None
    assert subtotal_rep["agg_function"] == "SUBTOTAL"
    assert "SUBTOTAL" in subtotal_rep["expression"]
    assert subtotal_rep["description"] == "Repeated column-wise SUBTOTAL formula. Each column calculates a value from rows above it."


    # Verify division by zero invalid formula and actual error description
    invalid = table1["invalid_formulas"]
    assert len(invalid) >= 1
    div_zero_item = next((i for i in invalid if i["formula"]["cell"] == "C1"), None)
    assert div_zero_item is not None
    assert div_zero_item["errors"][0]["type"] == "division_by_zero"
    # Description should reflect division_by_zero and not broken reference
    assert "division_by_zero" in div_zero_item["description"]
    assert "broken_reference" not in div_zero_item["description"]

    # Verify wrong value type invalid formula and actual error description
    wrong_val_item = next((i for i in invalid if i["formula"]["cell"] == "D2"), None)
    assert wrong_val_item is not None
    assert wrong_val_item["errors"][0]["type"] == "wrong_value_type"
    assert "wrong_value_type" in wrong_val_item["description"]

    # Verify table_nonexistent did not fallback to Sheet1 and has no relations (empty categories)
    assert "table_nonexistent" in relations
    table_non = relations["table_nonexistent"]
    assert len(table_non["normal_formulas"]) == 0
    assert len(table_non["aggregate_formulas"]) == 0
    assert len(table_non["cell_formulas"]) == 0
    assert len(table_non["invalid_formulas"]) == 0

    # Verify formula on an unrelated sheet (Sheet2!A1) stays unassigned
    for cat in ["normal_formulas", "aggregate_formulas", "cell_formulas", "invalid_formulas"]:
        for item in table1[cat]:
            # check both possible shapes of items
            formula_info = item.get("formula") or item.get("formula_example")
            if formula_info:
                assert formula_info.get("raw") != "=SUM(A2:A3)"

    # 4. Test CLI flag invocation
    cli_output = tmp_path / "cli_output.yaml"
    cmd = [
        sys.executable,
        str(Path(__file__).parent.parent / "TableAgent" / "perception" / "relations" / "extract.py"),
        "--xlsx", str(xlsx_path),
        "--structure", str(structure_path),
        "--output", str(cli_output)
    ]
    res = subprocess.run(cmd, capture_output=True, text=True)
    assert res.returncode == 0, f"CLI output: stdout: {res.stdout}, stderr: {res.stderr}"
    assert cli_output.exists()

    with open(cli_output, "r", encoding="utf-8") as f:
        cli_relations = yaml.safe_load(f)
    assert "table1" in cli_relations


def test_repeated_formula_descriptions():
    from TableAgent.perception.relations.classifier import generate_description

    # Test all row-wise aggregate functions
    for fn in ["SUM", "AVERAGE", "COUNT", "COUNTA", "MAX", "MIN", "SUBTOTAL"]:
        desc = generate_description("normal_formulas", "F{row} = " + fn + "(D{row}:E{row})", is_row_wise=True, agg_fun=fn)
        assert desc == f"Repeated row-wise {fn} formula. Each row calculates a value from a range of cells in the same row."

    # Test all column-wise aggregate functions
    for fn in ["SUM", "AVERAGE", "COUNT", "COUNTA", "MAX", "MIN", "SUBTOTAL"]:
        desc = generate_description("normal_formulas", "{col}8 = " + fn + "({col}2:{col}7)", is_col_wise=True, agg_fun=fn)
        assert desc == f"Repeated column-wise {fn} formula. Each column calculates a value from rows above it."

    # Test non-aggregate behavior remains unchanged
    # Row-wise non-aggregate arithmetic
    desc_arith = generate_description("normal_formulas", "C{row} = B{row} * 2", is_row_wise=True, agg_fun=None)
    assert desc_arith == "Repeated row-wise arithmetic formula."

    # Column-wise non-aggregate
    desc_col = generate_description("normal_formulas", "{col}8 = {col}2 - {col}3", is_col_wise=True, agg_fun=None)
    assert desc_col == "Repeated column-wise formula. Each column calculates a value from rows above it."

