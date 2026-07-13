from __future__ import annotations

import ast
from collections.abc import Iterable, Mapping
from numbers import Number
from typing import Any

import openpyxl
from openpyxl.formula import Tokenizer
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from TableAgent.QA.operators.base_operator import BaseOperator
from TableAgent.schema.range import CellRange
from TableAgent.utils import _lexical_overlap_score


class FormulaRelationOperator(BaseOperator):
    """Discover and deterministically evaluate formula relations."""

    name = "multitab.formula"
    description = (
        "Find stored formula relations and recalculate them with temporary mutations; "
        "the workbook is never modified."
    )
    examples = (
        "operators.list_relations(table_id='salary') -> list[dict]",
        "operators.find_relation('salary amount calculation', table_id='salary') -> list[dict]",
        "operators.evaluate_formula('rel_salary_calc', target_cell='E13', mutations={'C13': 20000000}) -> dict",
    )

    def list_relations(
        self,
        table_id: str | None = None,
        *,
        category: str | None = None,
    ) -> list[dict[str, Any]]:
        output = []
        for relation in getattr(self.env, "relations", []):
            if category and relation.get("category") != category:
                continue
            relation_tables = self._relation_table_ids(relation)
            if table_id and table_id not in relation_tables:
                continue
            item = dict(relation)
            item["table_ids"] = relation_tables
            output.append(item)
        return output

    def find_relation(
        self,
        query: str,
        *,
        table_id: str | None = None,
        top_k: int = 3,
    ) -> list[dict[str, Any]]:
        scored = []
        for relation in self.list_relations(table_id):
            text = " ".join(
                str(relation.get(key, ""))
                for key in ("id", "description", "expression", "pattern", "formula_example", "formula")
            )
            score = _lexical_overlap_score(query, text)
            if str(relation.get("id", "")).lower() in query.lower():
                score += 20.0
            if score > 0:
                scored.append((score, relation))
        scored.sort(key=lambda item: (-item[0], str(item[1].get("id", ""))))
        return [relation for _, relation in scored[:top_k]]

    def evaluate_formula(
        self,
        relation_id: str,
        *,
        target_cell: str | None = None,
        mutations: Mapping[str, Any] | None = None,
        table_id: str | None = None,
    ) -> dict[str, Any]:
        matches = [
            relation
            for relation in self.list_relations(table_id)
            if str(relation.get("id", "")) == relation_id
        ]
        if not matches:
            raise KeyError(f"Unknown formula relation: {relation_id}")
        if len(matches) > 1 and table_id is None:
            raise ValueError(f"Relation id {relation_id!r} is ambiguous; pass table_id")
        relation = matches[0]

        relation_tables = relation.get("table_ids") or self._relation_table_ids(relation)
        resolved_table_id = table_id or (relation_tables[0] if len(relation_tables) == 1 else None)
        sheet_name = self._relation_sheet(relation, resolved_table_id)
        formula, origin_cell = self._relation_formula(relation, sheet_name)
        target = target_cell or self._relation_target_cell(relation) or origin_cell
        if not target:
            raise ValueError(f"Relation {relation_id!r} does not identify a target cell")
        target_sheet, target_coord = _split_cell_reference(target, sheet_name)

        if origin_cell and target_coord != origin_cell:
            formula = Translator(formula, origin=origin_cell).translate_formula(target_coord)

        workbook = openpyxl.load_workbook(self.env.workbook_path, data_only=False)
        try:
            evaluator = _FormulaEvaluator(
                workbook,
                default_sheet=target_sheet,
                mutations=mutations or {},
            )
            value = evaluator.evaluate(formula, sheet_name=target_sheet)
        finally:
            workbook.close()

        return {
            "relation_id": relation_id,
            "table_id": resolved_table_id,
            "target_cell": f"{target_sheet}!{target_coord}",
            "formula": formula,
            "mutations": dict(mutations or {}),
            "value": value,
        }

    def _relation_table_ids(self, relation: Mapping[str, Any]) -> list[str]:
        explicit = relation.get("table_id")
        if explicit in self.env.structures:
            return [str(explicit)]

        relation_ref = self._relation_range_or_cell(relation)
        if not relation_ref:
            return []
        relation_sheet, relation_coord = _split_cell_reference(relation_ref, "")
        min_col, min_row, max_col, max_row = range_boundaries(relation_coord)
        matches = []
        for table_id, structure in self.env.structures.items():
            bounds = _table_bounds(structure)
            if bounds is None:
                continue
            table_sheet, table_range = bounds
            if relation_sheet and table_sheet != relation_sheet:
                continue
            if (
                table_range.start_row <= min_row <= max_row <= table_range.end_row
                and table_range.start_col <= min_col <= max_col <= table_range.end_col
            ):
                matches.append(table_id)
        return matches

    def _relation_sheet(self, relation: Mapping[str, Any], table_id: str | None) -> str:
        relation_ref = self._relation_range_or_cell(relation)
        if relation_ref and "!" in relation_ref:
            return _split_cell_reference(relation_ref, "")[0]
        if table_id and table_id in self.env.structures:
            return str(self.env.structures[table_id].get("sheet", ""))
        return self.env.get_active_sheet_name()

    @staticmethod
    def _relation_range_or_cell(relation: Mapping[str, Any]) -> str | None:
        if relation.get("range"):
            return str(relation["range"])
        formula = relation.get("formula")
        if isinstance(formula, Mapping) and formula.get("cell"):
            return str(formula["cell"])
        return None

    @classmethod
    def _relation_target_cell(cls, relation: Mapping[str, Any]) -> str | None:
        relation_ref = cls._relation_range_or_cell(relation)
        if not relation_ref:
            return None
        sheet, coordinate = _split_cell_reference(relation_ref, "")
        min_col, min_row, _, _ = range_boundaries(coordinate)
        cell = f"{get_column_letter(min_col)}{min_row}"
        return f"{sheet}!{cell}" if sheet else cell

    @classmethod
    def _relation_formula(cls, relation: Mapping[str, Any], sheet_name: str) -> tuple[str, str | None]:
        formula_payload = relation.get("formula")
        if isinstance(formula_payload, Mapping) and formula_payload.get("raw"):
            origin = _origin_coordinate(formula_payload.get("cell"), sheet_name)
            return _ensure_formula(str(formula_payload["raw"])), origin

        example = relation.get("formula_example")
        if isinstance(example, Mapping) and example.get("raw"):
            origin = _origin_coordinate(example.get("cell"), sheet_name)
            return _ensure_formula(str(example["raw"])), origin
        if isinstance(example, str) and example.strip():
            origin = cls._relation_target_cell(relation)
            if origin and "!" in origin:
                _, origin = _split_cell_reference(origin, sheet_name)
            return _ensure_formula(example), origin

        raise ValueError(f"Relation {relation.get('id')!r} does not contain an executable formula example")


class _FormulaEvaluator:
    def __init__(self, workbook: Any, *, default_sheet: str, mutations: Mapping[str, Any]):
        self.workbook = workbook
        self.default_sheet = default_sheet
        self.mutations: dict[tuple[str, str], Any] = {}
        self.stack: set[tuple[str, str]] = set()
        for reference, value in mutations.items():
            sheet, coordinate = _split_cell_reference(str(reference), default_sheet)
            self.mutations[(sheet, coordinate)] = value

    def evaluate(self, formula: str, *, sheet_name: str | None = None) -> Any:
        active_sheet = sheet_name or self.default_sheet
        expression = self._translate_formula(_ensure_formula(formula))
        functions = {
            "__cell": lambda reference: self._cell(reference, active_sheet),
            "__range": lambda reference: self._range(reference, active_sheet),
            "SUM": lambda *values: sum(_numbers(values)),
            "AVERAGE": lambda *values: _average(values),
            "MIN": lambda *values: _min_value(values),
            "MAX": lambda *values: _max_value(values),
            "COUNT": lambda *values: len(_numbers(values)),
            "COUNTA": lambda *values: len([value for value in _flatten(values) if value is not None]),
            "SUBTOTAL": lambda code, *values: _subtotal(code, values),
            "IF": lambda condition, when_true, when_false: when_true if condition else when_false,
            "AND": lambda *values: all(values),
            "OR": lambda *values: any(values),
            "NOT": lambda value: not value,
            "ABS": abs,
            "ROUND": round,
        }
        return _safe_eval(expression, functions)

    def _cell(self, reference: str, current_sheet: str) -> Any:
        sheet_name, coordinate = _split_cell_reference(reference, current_sheet)
        key = (sheet_name, coordinate)
        if key in self.stack:
            raise ValueError(f"Circular formula reference detected at {sheet_name}!{coordinate}")
        value = self.mutations.get(key, self.workbook[sheet_name][coordinate].value)
        if isinstance(value, str) and value.startswith("="):
            self.stack.add(key)
            try:
                return self.evaluate(value, sheet_name=sheet_name)
            finally:
                self.stack.remove(key)
        return value

    def _range(self, reference: str, current_sheet: str) -> list[Any]:
        sheet_name, coordinate = _split_cell_reference(reference, current_sheet)
        min_col, min_row, max_col, max_row = range_boundaries(coordinate)
        return [
            self._cell(
                f"{sheet_name}!{self.workbook[sheet_name].cell(row=row, column=column).coordinate}",
                sheet_name,
            )
            for row in range(min_row, max_row + 1)
            for column in range(min_col, max_col + 1)
        ]

    @staticmethod
    def _translate_formula(formula: str) -> str:
        output = []
        for token in Tokenizer(formula).items:
            if token.type == "OPERAND" and token.subtype == "RANGE":
                function = "__range" if ":" in token.value else "__cell"
                output.append(f"{function}({token.value!r})")
            elif token.type == "OPERAND" and token.subtype == "LOGICAL":
                output.append("True" if token.value.upper() == "TRUE" else "False")
            elif token.type == "FUNC" and token.subtype == "OPEN":
                output.append(f"{token.value[:-1].upper()}(")
            elif token.type == "OPERATOR-INFIX":
                output.append({"^": "**", "=": "==", "<>": "!=", "&": "+"}.get(token.value, token.value))
            elif token.type == "OPERATOR-POSTFIX" and token.value == "%":
                output.append("*0.01")
            elif token.type == "SEP":
                output.append(",")
            else:
                output.append(token.value)
        return "".join(output)


def _safe_eval(expression: str, functions: Mapping[str, Any]) -> Any:
    tree = ast.parse(expression, mode="eval")
    allowed_nodes = (
        ast.Expression,
        ast.Constant,
        ast.BinOp,
        ast.UnaryOp,
        ast.BoolOp,
        ast.Compare,
        ast.Call,
        ast.Name,
        ast.Load,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.FloorDiv,
        ast.Mod,
        ast.Pow,
        ast.USub,
        ast.UAdd,
        ast.Not,
        ast.And,
        ast.Or,
        ast.Eq,
        ast.NotEq,
        ast.Lt,
        ast.LtE,
        ast.Gt,
        ast.GtE,
    )
    for node in ast.walk(tree):
        if not isinstance(node, allowed_nodes):
            raise ValueError(f"Unsupported formula syntax: {type(node).__name__}")
        if isinstance(node, ast.Name) and node.id not in functions:
            raise ValueError(f"Unsupported formula function or name: {node.id}")
        if isinstance(node, ast.Call) and not isinstance(node.func, ast.Name):
            raise ValueError("Only direct formula function calls are supported")
    return eval(compile(tree, "<formula>", "eval"), {"__builtins__": {}}, dict(functions))


def _flatten(values: Iterable[Any]) -> list[Any]:
    flattened = []
    for value in values:
        if isinstance(value, (list, tuple)):
            flattened.extend(_flatten(value))
        else:
            flattened.append(value)
    return flattened


def _numbers(values: Iterable[Any]) -> list[Number]:
    return [value for value in _flatten(values) if isinstance(value, Number) and not isinstance(value, bool)]


def _average(values: Iterable[Any]) -> float:
    numbers = _numbers(values)
    if not numbers:
        raise ValueError("AVERAGE requires at least one numeric value")
    return float(sum(numbers)) / len(numbers)


def _min_value(values: Iterable[Any]) -> Number:
    numbers = _numbers(values)
    if not numbers:
        raise ValueError("MIN requires at least one numeric value")
    return min(numbers)


def _max_value(values: Iterable[Any]) -> Number:
    numbers = _numbers(values)
    if not numbers:
        raise ValueError("MAX requires at least one numeric value")
    return max(numbers)


def _subtotal(code: Any, values: Iterable[Any]) -> Any:
    function_code = int(code) % 100
    if function_code == 1:
        return _average(values)
    if function_code == 2:
        return len(_numbers(values))
    if function_code == 3:
        return len([value for value in _flatten(values) if value is not None])
    if function_code == 4:
        return _max_value(values)
    if function_code == 5:
        return _min_value(values)
    if function_code == 9:
        return sum(_numbers(values))
    raise ValueError(f"Unsupported SUBTOTAL function code: {code}")


def _ensure_formula(value: str) -> str:
    value = value.strip()
    return value if value.startswith("=") else f"={value}"


def _split_cell_reference(reference: str, default_sheet: str) -> tuple[str, str]:
    text = reference.strip()
    if "!" in text:
        sheet_name, coordinate = text.rsplit("!", 1)
        sheet_name = sheet_name.strip("'").replace("''", "'")
    else:
        sheet_name, coordinate = default_sheet, text
    return sheet_name, coordinate.replace("$", "")


def _origin_coordinate(reference: Any, default_sheet: str) -> str | None:
    if not reference:
        return None
    _, coordinate = _split_cell_reference(str(reference), default_sheet)
    return coordinate


def _table_bounds(structure: Mapping[str, Any]) -> tuple[str, CellRange] | None:
    ranges = [
        cell_range
        for header in structure.get("headers", [])
        for cell_range in (header.header_range, header.data_range)
        if cell_range is not None
    ]
    if not ranges:
        return None
    sheet = str(structure.get("sheet") or ranges[0].sheet)
    return sheet, CellRange(
        min(cell_range.start_row for cell_range in ranges),
        min(cell_range.start_col for cell_range in ranges),
        max(cell_range.end_row for cell_range in ranges),
        max(cell_range.end_col for cell_range in ranges),
        sheet,
    )
