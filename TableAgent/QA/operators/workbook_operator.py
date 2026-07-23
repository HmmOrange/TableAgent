from __future__ import annotations
from typing import Union, List, Any
import pandas as pd
from openpyxl.utils import get_column_letter
from TableAgent.schema.range import CellRange
from TableAgent.utils import parse_a1_range, read_excel_range
from TableAgent.QA.operators.base_operator import BaseOperator

class WorkbookOperator(BaseOperator):
    """Operator for reading cell values and converting to structures like DataFrames."""
    name = "workbook"
    description = "Read workbook cells from an A1 range or CellRange; use selective ranges to keep observations compact."
    examples = (
        "operators.read_table_as_dataframe(table_id, has_headers=False) -> pandas.DataFrame",
        "operators.read_range(range_or_a1, sheet='') -> list[list[Any]]",
        "operators.read_range_flat(range_or_a1, sheet='') -> list[Any]",
        "operators.read_range_as_dataframe(range_or_a1, sheet='', has_headers=True) -> pandas.DataFrame",
        "operators.sheet_dimensions(sheet='') -> dict[str, int | str]",
        "operators.read_sheet_as_dataframe(sheet='', min_row=1, max_row=None, min_col=1, max_col=None) -> pandas.DataFrame",
    )

    def read_range(
        self,
        range_or_a1: Union[CellRange, str],
        sheet: str = "",
        *,
        expand_merged: bool = False,
    ) -> List[List[Any]]:
        """Read raw cell values from the workbook range."""
        if isinstance(range_or_a1, str):
            cell_range = parse_a1_range(range_or_a1, sheet)
        else:
            cell_range = range_or_a1
            if sheet:
                cell_range = CellRange(
                    cell_range.start_row,
                    cell_range.start_col,
                    cell_range.end_row,
                    cell_range.end_col,
                    sheet
                )
        
        sheet_name = cell_range.sheet
        if not sheet_name:
            sheet_name = self.env.get_active_sheet_name()
        
        sheet_obj = self.env.get_sheet(sheet_name)
        if not sheet_obj:
            sheet_obj = self.env.get_active_sheet()
            
        return read_excel_range(sheet_obj, cell_range, expand_merged=expand_merged)

    def read_range_flat(self, range_or_a1: Union[CellRange, str], sheet: str = "") -> List[Any]:
        """Read values from range and flatten them into a single-dimensional list."""
        rows = self.read_range(range_or_a1, sheet)
        return [cell for row in rows for cell in row]

    def read_range_as_dataframe(self, range_or_a1: Union[CellRange, str], sheet: str = "", has_headers: bool = True) -> pd.DataFrame:
        """Read range and convert to a pandas DataFrame."""
        data = self.read_range(range_or_a1, sheet)
        if not data:
            return pd.DataFrame()
        if has_headers and len(data) > 1:
            headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(data[0])]
            return pd.DataFrame(data[1:], columns=headers)
        return pd.DataFrame(data)

    def sheet_dimensions(self, sheet: str = "") -> dict[str, int | str]:
        """Return the physical used bounds for a named worksheet."""
        sheet_obj = self.env.get_sheet(sheet) if sheet else self.env.get_active_sheet()
        if sheet_obj is None:
            raise ValueError(f"Workbook sheet not found: {sheet!r}")
        return {
            "sheet": sheet_obj.title,
            "min_row": 1,
            "max_row": int(sheet_obj.max_row),
            "min_col": 1,
            "max_col": int(sheet_obj.max_column),
            "range": f"A1:{get_column_letter(sheet_obj.max_column)}{sheet_obj.max_row}",
        }

    def read_sheet_as_dataframe(
        self,
        sheet: str = "",
        *,
        min_row: int = 1,
        max_row: int | None = None,
        min_col: int = 1,
        max_col: int | None = None,
    ) -> pd.DataFrame:
        """Read physical worksheet cells without relying on structure header grouping."""
        sheet_obj = self.env.get_sheet(sheet) if sheet else self.env.get_active_sheet()
        if sheet_obj is None:
            raise ValueError(f"Workbook sheet not found: {sheet!r}")
        end_row = min(int(max_row or sheet_obj.max_row), int(sheet_obj.max_row))
        end_col = min(int(max_col or sheet_obj.max_column), int(sheet_obj.max_column))
        if min_row < 1 or min_col < 1 or end_row < min_row or end_col < min_col:
            raise ValueError("Invalid worksheet bounds.")
        rows = list(sheet_obj.iter_rows(
            min_row=min_row,
            max_row=end_row,
            min_col=min_col,
            max_col=end_col,
            values_only=True,
        ))
        columns = [get_column_letter(column) for column in range(min_col, end_col + 1)]
        return pd.DataFrame(rows, columns=columns, index=range(min_row, end_row + 1))

if __name__ == "__main__":
    import argparse
    from TableAgent.environment.qa_env import QAEnvironment

    parser = argparse.ArgumentParser(description="Smoke-test workbook range operators.")
    parser.add_argument("--structure", default="sample/structure.yaml")
    parser.add_argument("--workbook", default="sample/QA_sample.xlsx")
    parser.add_argument("--range", default="A1:C5")
    args = parser.parse_args()

    env = QAEnvironment(args.structure, args.workbook)
    op = WorkbookOperator(env)
    values = op.read_range(args.range)
    flat = op.read_range_flat(args.range)
    print(f"range={args.range}")
    print(f"rows={len(values)} cols={len(values[0]) if values else 0}")
    print(f"first_row={values[0] if values else []}")
    print(f"flat_preview={flat[:8]}")
