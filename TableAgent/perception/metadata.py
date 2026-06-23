from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string


@dataclass(frozen=True)
class SheetMetadata:
    sheet_name: str
    used_range: str | None
    merged_ranges: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def to_yaml(self) -> str:
        return yaml.safe_dump(self.to_dict(), sort_keys=False, allow_unicode=True)


class ExStructMetadataExtractor:
    """Run ExStruct once per workbook and derive the layout metadata contract."""

    def __init__(self, mode: str = "light"):
        self.mode = mode

    def extract(self, workbook_path: Path) -> dict[str, Any]:
        import exstruct

        workbook = exstruct.extract(str(workbook_path), mode=self.mode, alpha_col=True)
        if hasattr(workbook, "model_dump"):
            return workbook.model_dump(mode="json")
        return workbook.to_dict()

    def sheet_metadata(
        self,
        workbook_path: Path,
        workbook_payload: dict[str, Any],
        sheet_name: str,
    ) -> SheetMetadata:
        sheet = (workbook_payload.get("sheets") or {}).get(sheet_name) or {}
        workbook_range, workbook_merges = _workbook_sheet_geometry(workbook_path, sheet_name)
        exstruct_range = _used_range_from_rows(sheet.get("rows") or [])
        exstruct_merges = [str(value) for value in sheet.get("merged_ranges") or []]
        merged_ranges = exstruct_merges or workbook_merges
        return SheetMetadata(
            sheet_name=sheet_name,
            used_range=_expand_range_to_merges(exstruct_range, merged_ranges) or workbook_range,
            merged_ranges=merged_ranges,
        )


def _used_range_from_rows(rows: list[dict[str, Any]]) -> str | None:
    coordinates: list[tuple[int, int]] = []
    for row in rows:
        try:
            row_index = int(row["r"])
        except (KeyError, TypeError, ValueError):
            continue
        cells = row.get("c") or {}
        if not isinstance(cells, dict):
            continue
        for column, value in cells.items():
            if value in (None, ""):
                continue
            try:
                column_index = (
                    int(column)
                    if str(column).isdigit()
                    else column_index_from_string(str(column))
                )
            except (TypeError, ValueError):
                continue
            coordinates.append((row_index, column_index))

    if not coordinates:
        return None
    min_row = min(row for row, _ in coordinates)
    max_row = max(row for row, _ in coordinates)
    min_col = min(column for _, column in coordinates)
    max_col = max(column for _, column in coordinates)
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _workbook_sheet_geometry(workbook_path: Path, sheet_name: str) -> tuple[str | None, list[str]]:
    import openpyxl

    workbook = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        worksheet = workbook[sheet_name]
        used_range = worksheet.calculate_dimension()
        if used_range == "A1:A1" and worksheet["A1"].value is None:
            used_range = None
        merged_ranges = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
        return used_range, merged_ranges
    finally:
        workbook.close()


def _expand_range_to_merges(used_range: str | None, merged_ranges: list[str]) -> str | None:
    if not used_range:
        return None
    from openpyxl.utils.cell import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(used_range)
    for merged_range in merged_ranges:
        try:
            merge_min_col, merge_min_row, merge_max_col, merge_max_row = range_boundaries(merged_range)
        except (TypeError, ValueError):
            continue
        if (
            merge_max_row < min_row
            or merge_min_row > max_row
            or merge_max_col < min_col
            or merge_min_col > max_col
        ):
            continue
        min_col = min(min_col, merge_min_col)
        min_row = min(min_row, merge_min_row)
        max_col = max(max_col, merge_max_col)
        max_row = max(max_row, merge_max_row)
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
