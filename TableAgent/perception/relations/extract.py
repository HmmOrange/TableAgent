from __future__ import annotations

import sys
from pathlib import Path
import argparse
from typing import Any
import openpyxl
import yaml

# Add project root to sys.path to allow direct script execution
project_root = str(Path(__file__).resolve().parent.parent.parent.parent)
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from TableAgent.perception.relations.scanner import scan_formulas
from TableAgent.perception.relations.assignment import (
    get_table_bounds,
    resolve_table_sheet,
    assign_formulas_to_table,
)
from TableAgent.perception.relations.classifier import classify_formulas_for_table
from TableAgent.perception.relations.writer import write_relations


def extract_relations(
    xlsx_path: str | Path,
    structure_path: str | Path,
    output_path: str | Path,
    footer_margin_rows: int = 5,
    side_margin_cols: int = 3,
) -> None:
    xlsx_path = Path(xlsx_path)
    structure_path = Path(structure_path)
    output_path = Path(output_path)

    with open(structure_path, "r", encoding="utf-8") as f:
        structure = yaml.safe_load(f) or {}

    is_single_table = False
    if "headers" in structure:
        is_single_table = True
        structure = {"table1": structure}

    workbook = openpyxl.load_workbook(xlsx_path, data_only=False)

    # 1. Scan formula cells across every worksheet used range once.
    scanned_formulas = scan_formulas(workbook)

    relations_output = {}

    for table_key, table_dict in structure.items():
        if not isinstance(table_dict, dict):
            continue

        # 2. Resolve sheet name, strictly adhering to no fallback if nonexistent explicit sheet
        sheet_name = resolve_table_sheet(table_dict, table_key, workbook)

        # 3. Calculate table bounds
        table_min_col, table_min_row, table_max_col, table_max_row = get_table_bounds(table_dict)

        if table_min_col == 0:
            if sheet_name and sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                table_min_col, table_min_row, table_max_col, table_max_row = 1, 1, sheet.max_column, sheet.max_row
            else:
                table_min_col, table_min_row, table_max_col, table_max_row = 1, 1, 1, 1

        # 4. Assign formulas strictly by sheet and range bounds
        assigned_formulas = assign_formulas_to_table(
            scanned_formulas,
            sheet_name,
            table_min_col,
            table_min_row,
            table_max_col,
            table_max_row,
            footer_margin_rows,
            side_margin_cols,
        )

        # 5. Classify the assigned formulas
        normal, aggregate, cell, invalid = classify_formulas_for_table(assigned_formulas, table_dict)

        relations_output[table_key] = {
            "normal_formulas": normal,
            "aggregate_formulas": aggregate,
            "cell_formulas": cell,
            "invalid_formulas": invalid,
        }

    output_dict = relations_output["table1"] if is_single_table else relations_output

    # 6. Write out relations
    write_relations(output_dict, output_path)

    workbook.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract relations from workbook and structure.")
    parser.add_argument("xlsx_path", nargs="?", help="Path to the Excel file")
    parser.add_argument("structure_path", nargs="?", help="Path to the structure YAML file")
    parser.add_argument("output_path", nargs="?", help="Path to write the relations YAML output")
    parser.add_argument("--xlsx", help="Path to the Excel file")
    parser.add_argument("--structure", help="Path to the structure YAML file")
    parser.add_argument("--output", help="Path to write the relations YAML output")
    parser.add_argument("--footer-margin-rows", type=int, default=5, help="Number of footer rows margin")
    parser.add_argument("--side-margin-cols", type=int, default=3, help="Number of side columns margin")

    args = parser.parse_args()

    xlsx_path = args.xlsx or args.xlsx_path
    structure_path = args.structure or args.structure_path
    output_path = args.output or args.output_path

    if not xlsx_path or not structure_path or not output_path:
        parser.error("Missing required arguments. Provide positionals or --xlsx, --structure, --output.")

    extract_relations(
        xlsx_path=xlsx_path,
        structure_path=structure_path,
        output_path=output_path,
        footer_margin_rows=args.footer_margin_rows,
        side_margin_cols=args.side_margin_cols
    )
