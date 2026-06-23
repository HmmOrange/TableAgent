from __future__ import annotations

from typing import Any
import openpyxl
from openpyxl.utils.cell import range_boundaries

# Constants
_UNCERTAIN_RANGE_VALUES = {"unknown", "uncertain", "n/a", "none", "null", "?"}


def get_table_bounds(table_dict: dict) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = float("inf"), float("inf"), float("-inf"), float("-inf")

    def collect_ranges(headers_list):
        ranges = []
        for h in headers_list:
            if not isinstance(h, dict):
                continue
            for key in ["data_range", "header_range", "range"]:
                val = h.get(key)
                if val and str(val).strip() and str(val).lower() not in _UNCERTAIN_RANGE_VALUES:
                    ranges.append(str(val).strip())
            sub = h.get("sub_headers")
            if sub and isinstance(sub, list):
                ranges.extend(collect_ranges(sub))
        return ranges

    ranges = collect_ranges(table_dict.get("headers") or [])
    if not ranges:
        return 0, 0, 0, 0

    for r_str in ranges:
        try:
            c_part = r_str.split("!", 1)[1] if "!" in r_str else r_str
            c_min_col, c_min_row, c_max_col, c_max_row = range_boundaries(c_part)
            min_col = min(min_col, c_min_col)
            min_row = min(min_row, c_min_row)
            max_col = max(max_col, c_max_col)
            max_row = max(max_row, c_max_row)
        except Exception:
            continue

    if min_col == float("inf"):
        return 0, 0, 0, 0
    return int(min_col), int(min_row), int(max_col), int(max_row)


def resolve_table_sheet(table_dict: dict[str, Any], table_key: str, workbook: openpyxl.Workbook) -> str | None:
    """
    Resolves the worksheet name for a table.
    Strictly follows: do not fallback a table with an explicit nonexistent sheet to active sheet.
    """
    explicit_sheet = None
    for key in ["sheet", "sheet_name"]:
        if key in table_dict and table_dict[key]:
            explicit_sheet = str(table_dict[key]).strip()
            break

    if explicit_sheet is not None:
        if explicit_sheet in workbook.sheetnames:
            return explicit_sheet
        # Do not fallback to active/default sheet if it was explicitly specified but does not exist.
        return None

    # Apply heuristics when not explicitly specified
    if len(workbook.sheetnames) == 1:
        return workbook.sheetnames[0]

    table_name = str(table_dict.get("name") or "")
    for s in workbook.sheetnames:
        if table_key.lower() in s.lower() or (table_name and table_name.lower() in s.lower()):
            return s

    return workbook.sheetnames[0]


def assign_formulas_to_table(
    scanned_formulas: list[dict[str, Any]],
    sheet_name: str | None,
    table_min_col: int,
    table_min_row: int,
    table_max_col: int,
    table_max_row: int,
    footer_margin_rows: int,
    side_margin_cols: int,
) -> list[dict[str, Any]]:
    """
    Assigns formula cells to a table strictly by sheet name and range boundaries.
    """
    if not sheet_name:
        return []

    assigned = []
    for f in scanned_formulas:
        if f["sheet"] == sheet_name:
            row, col = f["row"], f["col"]
            if (table_min_row <= row <= table_max_row + footer_margin_rows) and \
               (table_min_col - side_margin_cols <= col <= table_max_col + side_margin_cols):
                assigned.append(f)
    return assigned
