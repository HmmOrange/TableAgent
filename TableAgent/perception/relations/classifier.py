from __future__ import annotations

import re
from typing import Any
import openpyxl
from openpyxl.utils.cell import get_column_letter, range_boundaries

from TableAgent.perception.relations.normalize import (
    ref_regex,
    parse_reference,
    parse_cell_absolute,
    get_normalized_formula,
    get_pattern,
    normalize_expression_spacing,
)

# Constants from assignment (or local definition for self-containment)
_UNCERTAIN_RANGE_VALUES = {"unknown", "uncertain", "n/a", "none", "null", "?"}


def find_header_for_range(
    table_dict: dict,
    min_col: int,
    min_row: int,
    max_col: int,
    max_row: int,
    relaxed: bool = True,
) -> str | None:
    flat_headers = []

    def collect_headers(headers_list):
        for h in headers_list:
            if not isinstance(h, dict):
                continue
            dr = h.get("data_range") or h.get("range")
            if dr and str(dr).strip() and str(dr).lower() not in _UNCERTAIN_RANGE_VALUES:
                try:
                    c_part = str(dr).split("!", 1)[1] if "!" in str(dr) else str(dr)
                    h_min_col, h_min_row, h_max_col, h_max_row = range_boundaries(c_part)
                    flat_headers.append({
                        "id": h.get("id"),
                        "min_col": h_min_col,
                        "min_row": h_min_row,
                        "max_col": h_max_col,
                        "max_row": h_max_row,
                        "orientation": h.get("orientation", "column")
                    })
                except Exception:
                    pass
            sub = h.get("sub_headers")
            if sub and isinstance(sub, list):
                collect_headers(sub)

    collect_headers(table_dict.get("headers") or [])

    matches = []
    for h in flat_headers:
        if h["min_col"] <= min_col <= max_col <= h["max_col"] and h["min_row"] <= min_row <= max_row <= h["max_row"]:
            if h["id"]:
                matches.append(h)

    if matches:
        matches.sort(key=lambda x: (x["max_col"] - x["min_col"] + 1, x["max_row"] - x["min_row"] + 1))
        return matches[0]["id"]

    if not relaxed:
        return None

    for h in flat_headers:
        if h["id"]:
            if h["orientation"] in ["column", "column_group", "mixed"]:
                if h["min_col"] <= min_col <= max_col <= h["max_col"]:
                    matches.append(h)
            elif h["orientation"] in ["row", "row_group"]:
                if h["min_row"] <= min_row <= max_row <= h["max_row"]:
                    matches.append(h)

    if matches:
        matches.sort(key=lambda x: (x["max_col"] - x["min_col"] + 1, x["max_row"] - x["min_row"] + 1))
        return matches[0]["id"]

    return None


def partition_coordinates(
    coords: list[tuple[int, int]]
) -> tuple[list[list[tuple[int, int]]], list[list[tuple[int, int]]], list[tuple[int, int]]]:
    coords_set = set(coords)
    vertical_segments = []

    by_col = {}
    for r, c in coords:
        by_col.setdefault(c, []).append(r)

    for c, rows in by_col.items():
        rows.sort()
        current_block = []
        for r in rows:
            if not current_block or r == current_block[-1] + 1:
                current_block.append(r)
            else:
                if len(current_block) >= 2:
                    vertical_segments.append([(row, c) for row in current_block])
                current_block = [r]
        if len(current_block) >= 2:
            vertical_segments.append([(row, c) for row in current_block])

    for seg in vertical_segments:
        for coord in seg:
            coords_set.discard(coord)

    horizontal_segments = []
    by_row = {}
    for r, c in coords_set:
        by_row.setdefault(r, []).append(c)

    for r, cols in by_row.items():
        cols.sort()
        current_block = []
        for c in cols:
            if not current_block or c == current_block[-1] + 1:
                current_block.append(c)
            else:
                if len(current_block) >= 2:
                    horizontal_segments.append([(r, col) for col in current_block])
                current_block = [c]
        if len(current_block) >= 2:
            horizontal_segments.append([(r, col) for col in current_block])

    for seg in horizontal_segments:
        for coord in seg:
            coords_set.discard(coord)

    singletons = sorted(list(coords_set))
    return vertical_segments, horizontal_segments, singletons


def check_invalid_formula_errors(formula_raw: str) -> list[dict[str, str]]:
    errors = []
    error_mappings = [
        ("#REF!", "broken_reference", "Original referenced cell or range is lost."),
        ("#DIV/0!", "division_by_zero", "Formula attempts to divide by zero."),
        ("#VALUE!", "wrong_value_type", "Formula uses a value of the wrong data type."),
        ("#NAME?", "invalid_name", "Formula contains unrecognized text or function name."),
        ("#NUM!", "number_error", "Formula contains invalid numeric values."),
        ("#N/A", "value_not_available", "A value is not available to the formula."),
        ("#NULL!", "null_error", "Formula refers to an intersection of two ranges that do not intersect."),
    ]
    for token, err_type, msg in error_mappings:
        if token in formula_raw:
            errors.append({
                "type": err_type,
                "token": token,
                "message": msg
            })
    return errors


def translate_formula_to_expression(formula_raw: str, table_dict: dict, r_cell: int, c_cell: int) -> str:
    formula_body = formula_raw.lstrip("=")
    strings = []
    def replace_str(match):
        strings.append(match.group(0))
        return f"__STR_{len(strings)-1}__"
    formula_no_str = re.sub(r'"[^"]*"', replace_str, formula_body)

    matches = list(ref_regex.finditer(formula_no_str))
    parts = list(formula_no_str)

    for m in reversed(matches):
        ref_str = m.group(0)
        sheet_name, min_col, min_row, max_col, max_row = parse_reference(ref_str)
        header_id = find_header_for_range(table_dict, min_col, min_row, max_col, max_row)
        if header_id:
            replacement = header_id
        else:
            replacement = ref_str.split("!", 1)[1] if "!" in ref_str else ref_str

        start, end = m.span()
        parts[start:end] = list(replacement)

    expr_body = "".join(parts)
    for i, s in enumerate(strings):
        single_quoted = s.replace('"', "'")
        expr_body = expr_body.replace(f"__STR_{i}__", single_quoted)

    target_header_id = find_header_for_range(table_dict, c_cell, r_cell, c_cell, r_cell, relaxed=False)
    if target_header_id:
        target_name = target_header_id
    else:
        relaxed_header_id = find_header_for_range(table_dict, c_cell, r_cell, c_cell, r_cell, relaxed=True)
        if relaxed_header_id:
            flat_headers = []

            def collect_headers(headers_list):
                for h in headers_list:
                    if not isinstance(h, dict):
                        continue
                    if h.get("id") == relaxed_header_id:
                        dr = h.get("data_range") or h.get("range")
                        if dr and str(dr).strip() and str(dr).lower() not in _UNCERTAIN_RANGE_VALUES:
                            try:
                                c_part = str(dr).split("!", 1)[1] if "!" in str(dr) else str(dr)
                                _, _, _, h_max_row = range_boundaries(c_part)
                                flat_headers.append(h_max_row)
                            except Exception:
                                pass
                    sub = h.get("sub_headers")
                    if sub and isinstance(sub, list):
                        collect_headers(sub)

            collect_headers(table_dict.get("headers") or [])
            max_h_row = max(flat_headers) if flat_headers else 0
            if r_cell > max_h_row:
                outer_fn = None
                for fn in ["SUM", "AVERAGE", "MIN", "MAX", "COUNT", "COUNTA", "SUBTOTAL"]:
                    if re.search(r"\b" + fn + r"\s*\(", formula_raw, re.IGNORECASE):
                        outer_fn = fn
                        break
                if outer_fn:
                    fn_prefix = outer_fn.lower()
                    if fn_prefix == "sum":
                        fn_prefix = "total"
                    target_name = f"{fn_prefix}_{relaxed_header_id}"
                else:
                    target_name = f"summary_{relaxed_header_id}"
            else:
                target_name = relaxed_header_id
        else:
            target_name = f"{get_column_letter(c_cell)}{r_cell}"

    return f"{target_name} = {expr_body}"


def generate_description(
    category: str,
    pattern_or_formula: str,
    is_row_wise: bool = False,
    is_col_wise: bool = False,
    agg_fun: str | None = None,
    cell_name: str | None = None,
    errors: list[dict[str, str]] | None = None,
) -> str:
    if category == "invalid_formulas":
        if errors:
            err_details = []
            for err in errors:
                err_type = err.get("type", "unknown_error")
                err_msg = err.get("message", "An error was detected.")
                err_details.append(f"{err_type} ({err_msg})")
            return f"Formula contains error(s): {', '.join(err_details)}."
        return "Formula contains a broken reference and should not be executed before repair."

    if category == "normal_formulas":
        if is_row_wise:
            if agg_fun and agg_fun.upper() in ["SUM", "AVERAGE", "COUNT", "COUNTA", "MAX", "MIN", "SUBTOTAL"]:
                return f"Repeated row-wise {agg_fun.upper()} formula. Each row calculates a value from a range of cells in the same row."
            elif any(c in pattern_or_formula for c in ["*", "/", "+", "-"]):
                if "$" in pattern_or_formula:
                    return "Repeated row-wise arithmetic formula using an absolute reference."
                else:
                    return "Repeated row-wise arithmetic formula."
            else:
                return "Repeated row-wise formula."
        elif is_col_wise:
            if agg_fun and agg_fun.upper() in ["SUM", "AVERAGE", "COUNT", "COUNTA", "MAX", "MIN", "SUBTOTAL"]:
                return f"Repeated column-wise {agg_fun.upper()} formula. Each column calculates a value from rows above it."
            else:
                return "Repeated column-wise formula. Each column calculates a value from rows above it."


    if category == "aggregate_formulas":
        if agg_fun and cell_name:
            return f"Aggregate {agg_fun} formula in cell {cell_name} over a column or range."
        elif agg_fun:
            return f"Aggregate {agg_fun} formula over a column or range."
        elif cell_name:
            return f"Aggregate formula in cell {cell_name} over a column or range."
        return "Aggregate formula over a column or range."

    if category == "cell_formulas":
        cell_part = f" in cell {cell_name}" if cell_name else ""
        if "IF" in pattern_or_formula.upper():
            return f"Standalone formula used only in one specific cell{cell_part}."
        elif any(c in pattern_or_formula for c in ["*", "/", "+", "-"]):
            return f"Standalone arithmetic formula{cell_part} that does not repeat across rows or columns."
        else:
            return f"Standalone cell formula{cell_part}."

    return "Formula cell relation."


def classify_formulas_for_table(
    assigned_formulas: list[dict[str, Any]],
    table_dict: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Classifies the assigned formulas of a single table into:
    - normal_formulas
    - aggregate_formulas
    - cell_formulas
    - invalid_formulas
    """
    invalid_list = []
    valid_list = []

    for f in assigned_formulas:
        cell = f["cell"]
        formula_raw = f["raw"]
        errors = check_invalid_formula_errors(formula_raw)
        if errors:
            invalid_list.append((cell, formula_raw, errors))
        else:
            valid_list.append((cell, formula_raw))

    normalized_groups = {}
    for cell, formula_raw in valid_list:
        norm = get_normalized_formula(formula_raw, cell.row, cell.column)
        normalized_groups.setdefault(norm, []).append((cell, formula_raw))

    normal_formulas_out = []
    aggregate_formulas_out = []
    cell_formulas_out = []
    invalid_formulas_out = []

    normal_idx = 1
    agg_idx = 1
    cell_idx = 1
    invalid_idx = 1

    for norm_str, cells_in_group in normalized_groups.items():
        coords = [(c.row, c.column) for c, _ in cells_in_group]
        cell_map = {(c.row, c.column): (c, raw) for c, raw in cells_in_group}

        vertical_segs, horizontal_segs, singletons = partition_coordinates(coords)

        for seg in vertical_segs:
            rows = [r for r, _ in seg]
            cols = [c for _, c in seg]
            min_r, max_r = min(rows), max(rows)
            min_c, max_c = min(cols), max(cols)

            range_str = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
            example_coord = seg[0]
            example_cell, example_raw = cell_map[example_coord]

            pattern_str = get_pattern(example_raw, example_cell.row, example_cell.column, is_vertical=True)
            pattern_str = normalize_expression_spacing(pattern_str)
            agg_fun = None
            for fn in ["SUM", "AVERAGE", "MIN", "MAX", "COUNT", "COUNTA", "SUBTOTAL"]:
                if re.search(r"\b" + fn + r"\s*\(", example_raw, re.IGNORECASE):
                    agg_fun = fn
                    break

            expr = translate_formula_to_expression(example_raw, table_dict, example_cell.row, example_cell.column)
            expr = normalize_expression_spacing(expr)

            col_letter = get_column_letter(min_c)
            if agg_fun == "SUM":
                type_suffix = "row_sum"
            elif agg_fun == "SUBTOTAL":
                type_suffix = "row_subtotal"
            elif any(c in example_raw for c in ["*", "/", "+", "-"]):
                type_suffix = "row_arithmetic"
            else:
                type_suffix = "row_formula"
            fid = f"rel_repeat_{type_suffix}_{col_letter.lower()}_{normal_idx}"
            normal_idx += 1

            desc = generate_description("normal_formulas", pattern_str, is_row_wise=True, agg_fun=agg_fun)
            normal_formulas_out.append({
                "id": fid,
                "expression": expr,
                "description": desc,
                "range": range_str,
                "pattern": pattern_str,
                "agg_function": agg_fun,
                "formula_example": {
                    "cell": f"{col_letter}{example_cell.row}",
                    "raw": example_raw
                }
            })

        for seg in horizontal_segs:
            rows = [r for r, _ in seg]
            cols = [c for _, c in seg]
            min_r, max_r = min(rows), max(rows)
            min_c, max_c = min(cols), max(cols)

            range_str = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
            example_coord = seg[0]
            example_cell, example_raw = cell_map[example_coord]

            pattern_str = get_pattern(example_raw, example_cell.row, example_cell.column, is_vertical=False)
            pattern_str = normalize_expression_spacing(pattern_str)

            agg_fun = None
            for fn in ["SUM", "AVERAGE", "MIN", "MAX", "COUNT", "COUNTA", "SUBTOTAL"]:
                if re.search(r"\b" + fn + r"\s*\(", example_raw, re.IGNORECASE):
                    agg_fun = fn
                    break

            expr = translate_formula_to_expression(example_raw, table_dict, example_cell.row, example_cell.column)
            expr = normalize_expression_spacing(expr)

            fid = f"rel_repeat_column_formula_{min_r}_{normal_idx}"
            normal_idx += 1

            desc = generate_description("normal_formulas", pattern_str, is_col_wise=True, agg_fun=agg_fun)
            normal_formulas_out.append({
                "id": fid,
                "expression": expr,
                "description": desc,
                "range": range_str,
                "pattern": pattern_str,
                "agg_function": agg_fun,
                "formula_example": {
                    "cell": f"{get_column_letter(example_cell.column)}{example_cell.row}",
                    "raw": example_raw
                }
            })

        for coord in singletons:
            cell, raw = cell_map[coord]
            agg_fun = None
            for fn in ["SUM", "AVERAGE", "MIN", "MAX", "COUNT", "COUNTA", "SUBTOTAL"]:
                if re.search(r"\b" + fn + r"\s*\(", raw, re.IGNORECASE):
                    agg_fun = fn
                    break

            expr = translate_formula_to_expression(raw, table_dict, cell.row, cell.column)
            expr = normalize_expression_spacing(expr)
            cell_name = f"{get_column_letter(cell.column)}{cell.row}"

            if agg_fun is not None:
                fid = f"rel_aggregate_{agg_fun.lower()}_{agg_idx}"
                agg_idx += 1
                desc = generate_description("aggregate_formulas", raw, cell_name=cell_name, agg_fun=agg_fun)
                aggregate_formulas_out.append({
                    "id": fid,
                    "expression": expr,
                    "description": desc,
                    "formula": {
                        "cell": cell_name,
                        "raw": raw
                    }
                })
            else:
                if "IF" in raw.upper():
                    fid = f"rel_single_cell_formula_{cell_idx}"
                elif any(c in raw for c in ["*", "/", "+", "-"]):
                    fid = f"rel_single_arithmetic_formula_{cell_idx}"
                else:
                    fid = f"rel_single_cell_formula_{cell_idx}"
                cell_idx += 1

                desc = generate_description("cell_formulas", raw, cell_name=cell_name)
                cell_formulas_out.append({
                    "id": fid,
                    "expression": expr,
                    "description": desc,
                    "formula": {
                        "cell": cell_name,
                        "raw": raw
                    }
                })

    for cell, raw, errors in invalid_list:
        cell_name = f"{get_column_letter(cell.column)}{cell.row}"
        fid = f"rel_invalid_ref_{invalid_idx}"
        invalid_idx += 1
        desc = generate_description("invalid_formulas", raw, errors=errors)

        invalid_formulas_out.append({
            "id": fid,
            "expression": None,
            "description": desc,
            "formula": {
                "cell": cell_name,
                "raw": raw
            },
            "errors": errors
        })

    return normal_formulas_out, aggregate_formulas_out, cell_formulas_out, invalid_formulas_out
