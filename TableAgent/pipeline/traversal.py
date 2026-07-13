from __future__ import annotations

import heapq
from dataclasses import dataclass
from enum import IntEnum

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


class Direction(IntEnum):
    STAY = 0
    RIGHT = 1
    DOWN = 2
    LEFT = 3
    UP = 4

    @classmethod
    def parse(cls, value: str | "Direction") -> "Direction | None":
        if isinstance(value, cls):
            return value
        try:
            return cls[str(value).strip().upper()]
        except KeyError:
            return None


@dataclass(frozen=True)
class Viewport:
    row: int
    column: int
    rows: int
    columns: int

    @property
    def a1_range(self) -> str:
        end_row = self.row + self.rows - 1
        end_column = self.column + self.columns - 1
        return (
            f"{get_column_letter(self.column)}{self.row}:"
            f"{get_column_letter(end_column)}{end_row}"
        )

    def clipped_a1_range(self, table_range: str | None) -> str:
        """Return the visible part of this viewport inside the worksheet used range."""
        if not table_range:
            return self.a1_range
        min_col, min_row, max_col, max_row = range_boundaries(table_range)
        start_col = max(self.column, min_col)
        start_row = max(self.row, min_row)
        end_col = min(self.column + self.columns - 1, max_col)
        end_row = min(self.row + self.rows - 1, max_row)
        return (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )

    @property
    def key(self) -> tuple[int, int]:
        return self.row, self.column

    def shifted(self, direction: Direction, distance: int) -> "Viewport":
        row = self.row
        column = self.column
        if direction == Direction.RIGHT:
            column += distance
        elif direction == Direction.DOWN:
            row += distance
        elif direction == Direction.LEFT:
            column = max(1, column - distance)
        elif direction == Direction.UP:
            row = max(1, row - distance)
        return Viewport(row=row, column=column, rows=self.rows, columns=self.columns)


@dataclass(frozen=True)
class TraversalTask:
    direction: Direction
    viewport: Viewport


class DirectionQueue:
    def __init__(self):
        self._heap: list[tuple[int, int, TraversalTask]] = []
        self._queued: set[tuple[Direction, int, int]] = set()
        self._sequence = 0

    def push(self, task: TraversalTask) -> bool:
        key = (task.direction, *task.viewport.key)
        if key in self._queued:
            return False
        heapq.heappush(self._heap, (int(task.direction), self._sequence, task))
        self._queued.add(key)
        self._sequence += 1
        return True

    def pop(self) -> TraversalTask:
        _, _, task = heapq.heappop(self._heap)
        self._queued.remove((task.direction, *task.viewport.key))
        return task

    def __bool__(self) -> bool:
        return bool(self._heap)

    def __len__(self) -> int:
        return len(self._heap)


def initial_viewport(
    table_range: str | None,
    *,
    rows: int,
    columns: int,
) -> Viewport:
    if table_range:
        min_col, min_row, _, _ = range_boundaries(table_range)
    else:
        min_col, min_row = 1, 1
    return Viewport(row=min_row, column=min_col, rows=rows, columns=columns)


def corner_viewports(
    table_range: str | None,
    *,
    rows: int,
    columns: int,
) -> list[tuple[Direction, Viewport]]:
    if not table_range:
        return [(Direction.STAY, initial_viewport(table_range, rows=rows, columns=columns))]

    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    left_col = min_col
    right_col = max(min_col, max_col - columns + 1)
    top_row = min_row
    bottom_row = max(min_row, max_row - rows + 1)

    candidates = [
        (Direction.STAY, Viewport(row=top_row, column=left_col, rows=rows, columns=columns)),
        (Direction.RIGHT, Viewport(row=top_row, column=right_col, rows=rows, columns=columns)),
        (Direction.DOWN, Viewport(row=bottom_row, column=left_col, rows=rows, columns=columns)),
        (Direction.RIGHT, Viewport(row=bottom_row, column=right_col, rows=rows, columns=columns)),
    ]
    seen_ranges: set[str] = set()
    unique: list[tuple[Direction, Viewport]] = []
    for direction, viewport in candidates:
        clipped_range = viewport.clipped_a1_range(table_range)
        if clipped_range in seen_ranges:
            continue
        seen_ranges.add(clipped_range)
        unique.append((direction, viewport))
    return unique


def frontier_directions(table_range: str | None, viewport: Viewport) -> list[Direction]:
    if not table_range:
        return []
    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    directions = []
    if viewport.column + viewport.columns - 1 < max_col:
        directions.append(Direction.RIGHT)
    if viewport.row + viewport.rows - 1 < max_row:
        directions.append(Direction.DOWN)
    if viewport.column > min_col:
        directions.append(Direction.LEFT)
    if viewport.row > min_row:
        directions.append(Direction.UP)
    return directions
