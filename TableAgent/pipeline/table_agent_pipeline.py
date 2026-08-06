from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import TYPE_CHECKING, Any, Callable

import openpyxl
from TableAgent.prompts.answer import ANSWER_SYSTEM_PROMPT, ANSWER_USER_PROMPT_TEMPLATE
from TableAgent.prompts.reranker import RERANKER_SYSTEM_PROMPT, RERANKER_USER_PROMPT_TEMPLATE

from TableAgent.configs import TableAgentConfig
from TableAgent.configs.models_config import available_models
from TableAgent.llm import BaseLLM
from TableAgent.pipeline.base import BasePipeline
from TableAgent.QA.agents.answer_agent import QAAgent
from TableAgent.QA.runner import TableQARunner
from TableAgent.perception.metadata import SheetMetadata
from TableAgent.run_logging import Logger
from TableAgent.schema import EvalSample
from TableAgent.structure.layout.agent import LayoutAgent
from TableAgent.pipeline.common import has_workbook_sources
from TableAgent.pipeline.prompting import PromptBuilder
from TableAgent.pipeline.retrieval import SourceRetriever
from TableAgent.pipeline.retrieval.embeddings import OpenAICompatibleEmbeddingClient
from TableAgent.pipeline.pipeline_qa import PipelineQAMixin
from TableAgent.pipeline.pipeline_run import PipelineRunMixin
from TableAgent.pipeline.pipeline_source_qa import PipelineSourceQAMixin
from TableAgent.pipeline.source_preparer import SourcePreparer
from TableAgent.structure.layout.workflow import TableLayoutWorkflow
from TableAgent.pipeline.structure_cache import StructureCache, StructureCacheRecord
from TableAgent.rendering.workbook import WorkbookRenderer
from TableAgent.structure.verification import DeterministicVerifier

if TYPE_CHECKING:
    from TableAgent.pipeline.retrieval import TableRetrieverContract

logger = Logger(__name__)

class TableAgentPipeline(
    PipelineRunMixin,
    PipelineSourceQAMixin,
    PipelineQAMixin,
    BasePipeline,
):
    name = "table_agent"
    prepare_samples_before_run = True
    answer_system_prompt = ANSWER_SYSTEM_PROMPT
    answer_user_prompt_template = ANSWER_USER_PROMPT_TEMPLATE
    reranker_system_prompt = RERANKER_SYSTEM_PROMPT
    reranker_user_prompt_template = RERANKER_USER_PROMPT_TEMPLATE

    def __init__(
        self,
        llm_client: BaseLLM | None,
        layout_vlm_client: BaseLLM | None,
        config: dict[str, Any] | None = None,
        table_retriever: TableRetrieverContract | None = None,
        embedding_client: Any | None = None,
    ):
        self.llm = llm_client
        self.layout_vlm = layout_vlm_client
        self.settings = TableAgentConfig.from_config(config)
        self._artifact_dir = self.settings.artifact_dir
        self._artifact_dir.mkdir(parents=True, exist_ok=True)
        if self.settings.phase in {"qa", "all"} and self.llm is None:
            raise ValueError(f"TableAgent phase '{self.settings.phase}' requires an answer LLM client")
        if self.settings.phase in {"structure", "all"} and self.layout_vlm is None:
            raise ValueError(f"TableAgent phase '{self.settings.phase}' requires a layout VLM client")
        self.prompts = PromptBuilder(self.settings, self)
        self.workbook_renderer = WorkbookRenderer(self.settings, logger)
        self.layout_agent = LayoutAgent(self.layout_vlm) if self.layout_vlm is not None else None
        self.verifier = DeterministicVerifier(
            data_only=self.settings.structure_data_only,
        )
        self.qa_agent = QAAgent(self.llm, self.answer_system_prompt) if self.llm is not None else None
        self.table_retriever = table_retriever
        self.layout_workflow = (
            TableLayoutWorkflow(
                self.settings,
                self.workbook_renderer,
                self.layout_agent,
                self.verifier,
                progress_callback=self._progress,
            )
            if self.layout_agent is not None
            else None
        )
        configured_models = available_models(config or {})
        if (
            embedding_client is None
            and self.settings.retrieval_embedding_provider not in {None, "mock"}
            and self.settings.retrieval_embedding_provider in configured_models
        ):
            embedding_client = OpenAICompatibleEmbeddingClient.from_config(
                config or {},
                self.settings.retrieval_embedding_provider,
            )
        embedding_model = str(getattr(embedding_client, "model", "") or "")
        if self.settings.embed_retrieval_cards and (
            embedding_client is None or not embedding_model
        ):
            raise ValueError(
                "Embedding export requires a configured real embedding provider and model"
            )
        self.source_preparer = SourcePreparer(
            self.settings,
            self._analyze_source_sheet,
            progress_callback=self._progress,
            embedding_client=embedding_client,
            embedding_model=embedding_model,
        )
        self.source_retriever = SourceRetriever(
            self.settings,
            self.llm,
            self,
            self.prompts,
            embedding_client=embedding_client,
        )
        self.structure_cache = StructureCache(
            self.settings,
            self.layout_workflow,
            self._metadata_for_workbook_sheet,
        )
        self._verified_samples: dict[str, StructureCacheRecord] = {}
        self._prepared_source_samples: set[str] = set()
        self._progress_callback: Callable[[str], None] | None = None
        self._apply_generation_cap()

    def prepare_samples(self, samples: list[EvalSample], logger: Any | None = None) -> None:
        if self.settings.phase == "qa":
            missing = []
            for sample in samples:
                if (
                    self.settings.should_retrieve(sample)
                    and has_workbook_sources(sample)
                    and (
                        self.source_retriever.load_perfect_candidates(sample)
                        if self.settings.perfect_retrieval
                        else self.source_retriever.load_candidates(sample)
                    )
                ):
                    continue
                record = self.structure_cache.load(sample)
                if record is None or not record.valid:
                    missing.append(sample.sample_id)
            if missing:
                raise RuntimeError(
                    "Missing or stale TableAgent structure caches for: "
                    + ", ".join(missing[:20])
                    + ". Run with --table-agent-phase structure or all first."
                )
            return
        records = self.verify_samples(samples, force=self.settings.phase == "all")
        failed = [record for record in records if not record.valid]
        if failed:
            raise RuntimeError(f"TableAgent verification failed for {len(failed)} cache entries")
        if self.settings.phase == "all":
            self._prepared_source_samples.update(
                sample.sample_id
                for sample in samples
                if has_workbook_sources(sample) and self.settings.should_retrieve(sample)
            )

    def filter_samples(self, samples: list[EvalSample]) -> list[EvalSample]:
        """Skip samples whose explicitly configured perfect-retrieval source is unavailable."""
        if not self.settings.perfect_retrieval:
            return samples

        filtered = []
        for sample in samples:
            try:
                self.source_retriever.select_perfect(sample)
            except RuntimeError as exc:
                if "Perfect retrieval excludes sheet" in str(exc):
                    continue
                raise
            filtered.append(sample)
        return filtered

    def verify_samples(self, samples: list[EvalSample], *, force: bool = True) -> list[StructureCacheRecord]:
        source_samples = [
            sample
            for sample in samples
            if has_workbook_sources(sample) and self.settings.should_retrieve(sample)
        ]
        standard_samples = [
            sample
            for sample in samples
            if not has_workbook_sources(sample) or not self.settings.should_retrieve(sample)
        ]
        records = []
        for sample in standard_samples:
            record = self.structure_cache.prepare(sample, force=force)
            records.append(record)
            self._progress(
                "structure_done",
                sample=sample.sample_id,
                workbook=record.workbook_path.name,
                sheet=record.sheet_name,
            )
        self._verified_samples.update({sample.sample_id: record for sample, record in zip(standard_samples, records)})
        if source_samples:
            if not self.settings.perfect_retrieval:
                self.source_preparer.prepare(source_samples, regenerate_invalid=force, force=force)
            seen: set[Path] = set()
            for sample in source_samples:
                candidates = (
                    self.source_retriever.load_perfect_candidates(sample)
                    if self.settings.perfect_retrieval
                    else self.source_retriever.load_candidates(sample)
                )
                if not candidates:
                    failure_dir = self.settings.source_artifact_dir or self.settings.structure_cache_dir
                    key = hashlib.sha256(sample.sample_id.encode("utf-8")).hexdigest()[:24]
                    records.append(StructureCacheRecord(
                        key=key,
                        directory=failure_dir,
                        workbook_path=Path(str(sample.table_path).split(";")[0]),
                        sheet_name="",
                        structure_path=failure_dir / "structure.yaml",
                        manifest_path=failure_dir / "metadata.json",
                        status="not_good",
                        cache_hit=False,
                    ))
                    continue
                for candidate in candidates:
                    if candidate.directory in seen:
                        continue
                    seen.add(candidate.directory)
                    key = hashlib.sha256(str(candidate.directory.resolve()).encode("utf-8")).hexdigest()[:24]
                    verification = self._prepared_verification(candidate.directory)
                    records.append(StructureCacheRecord(
                        key=key,
                        directory=candidate.directory,
                        workbook_path=candidate.workbook_path,
                        sheet_name=candidate.sheet_name,
                        structure_path=candidate.directory / "structure.yaml",
                        manifest_path=candidate.directory / "metadata.json",
                        status=str(verification.get("status") or "good"),
                        cache_hit=not force,
                    ))
        return records

    @staticmethod
    def structure_progress_totals(samples: list[EvalSample]) -> dict[str, Any]:
        """Count the structure work units shown by the CLI progress bar."""
        standard_samples = [sample for sample in samples if not has_workbook_sources(sample)]
        source_samples = [sample for sample in samples if has_workbook_sources(sample)]
        selected_sheets = set(SourcePreparer.selected_sheet_names(source_samples))
        sheets_per_file = {f"sample:{sample.sample_id}": 1 for sample in standard_samples}
        files_per_key = {key: 1 for key in sheets_per_file}

        for source_path in SourcePreparer._source_paths(source_samples):
            try:
                workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
                try:
                    sheet_count = max(
                        len([name for name in workbook.sheetnames if not selected_sheets or name in selected_sheets]),
                        1,
                    )
                finally:
                    workbook.close()
            except Exception:
                sheet_count = 1
            key = f"book:{source_path.name}"
            sheets_per_file[key] = sheets_per_file.get(key, 0) + sheet_count
            files_per_key[key] = files_per_key.get(key, 0) + 1

        return {
            "files": sum(files_per_key.values()),
            "sheets": sum(sheets_per_file.values()),
            "sheets_per_file": sheets_per_file,
            "files_per_key": files_per_key,
        }

    def set_progress_callback(self, callback: Callable[[str], None] | None) -> None:
        self._progress_callback = callback

    def _progress(self, stage: str, **fields: Any) -> None:
        if self._progress_callback is None:
            return
        labels = {
            "prepare": "prepare",
            "prepare_extract": "prepare:extract",
            "prepare_metadata": "prepare:metadata",
            "prepare_cached": "prepare:cached",
            "prepare_error": "prepare:error",
            "prepare_layout": "prepare:layout",
            "prepare_done": "prepare:done",
            "retrieval": "retrieve",
            "rerank": "rerank",
            "render": "render",
            "layout": "layout",
            "verify": "verify",
            "structure_done": "structure:done",
            "qa": "qa",
            "answer": "answer",
            "done": "done",
        }
        parts = [labels.get(stage, stage)]
        if stage in {"prepare_layout", "prepare_done", "render", "layout", "verify"}:
            ordered_fields = [
                ("range", "range"),
                ("iteration", "iter"),
                ("direction", "dir"),
                ("workbook", "book"),
                ("sheet", "sheet"),
                ("sample", "sample"),
            ]
        else:
            ordered_fields = [
                ("sample", "sample"),
                ("workbook", "book"),
                ("sheet", "sheet"),
                ("table", "table"),
                ("range", "range"),
                ("iteration", "iter"),
                ("direction", "dir"),
            ]
        for key, label in ordered_fields:
            value = fields.get(key)
            if value is None or value == "":
                continue
            text = str(value)
            parts.append(f"{label}={text}")
        self._progress_callback(" | ".join(parts))

    def set_run_id(self, run_id: int) -> Path:
        if run_id < 1:
            raise ValueError("run_id must be at least 1")
        if self.settings.run_artifact_dir is not None:
            repeat_dir = self.settings.repeat_dir_template.format(run_id=run_id)
            self._artifact_dir = self.settings.run_artifact_dir / repeat_dir
        else:
            self._artifact_dir = self.settings.artifact_dir
        self._artifact_dir.mkdir(parents=True, exist_ok=True)
        return self._artifact_dir

    def _analyze_source_sheet(
        self,
        source_path: Path,
        sheet_name: str,
        metadata: SheetMetadata,
        sheet_dir: Path,
    ) -> str:
        if self.layout_workflow is None:
            raise RuntimeError("Source verification requires a layout VLM client")
        self._progress("prepare", workbook=source_path.name, sheet=sheet_name, range=metadata.used_range)
        result = self.layout_workflow.run(
            workbook_path=source_path,
            sheet_name=sheet_name,
            metadata=metadata,
            output_dir=sheet_dir,
        )
        metadata_path = sheet_dir / "metadata.json"
        if metadata_path.is_file():
            metadata_payload = json.loads(metadata_path.read_text(encoding="utf-8"))
            metadata_payload["verification"] = result.verification
            metadata_path.write_text(json.dumps(metadata_payload, ensure_ascii=False), encoding="utf-8")
        return result.structure_text

    @staticmethod
    def _metadata_for_workbook_sheet(workbook_path: Path, sheet_name: str) -> SheetMetadata:
        workbook = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
        try:
            worksheet = workbook[sheet_name]
            used_range = worksheet.calculate_dimension()
            merged_ranges = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
            if used_range == "A1:A1" and worksheet["A1"].value is None:
                used_range = None
            return SheetMetadata(sheet_name, used_range, merged_ranges)
        finally:
            workbook.close()

    def _apply_generation_cap(self) -> None:
        if self.settings.generation_max_tokens is None:
            return
        if hasattr(self.llm, "max_tokens"):
            self.llm.max_tokens = self.settings.generation_max_tokens
        if hasattr(self.layout_vlm, "max_tokens"):
            self.layout_vlm.max_tokens = self.settings.generation_max_tokens

    @staticmethod
    def _client_config(client: Any) -> dict[str, Any]:
        return {
            "model_name": getattr(client, "model_name", None),
            "temperature": getattr(client, "temperature", None),
            "max_tokens": getattr(client, "max_tokens", None),
        }
