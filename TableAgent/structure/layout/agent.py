from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from TableAgent.prompts.structure import (
    LAYOUT_MAS_SYSTEM_PROMPT,
    LAYOUT_MAS_USER_PROMPT_TEMPLATE,
)
from TableAgent.agents.base import AgentMessage, BaseTableAgent
from TableAgent.llm import BaseLLM, LLMResponse

from TableAgent.structure.layout.parsing import (
    _is_valid_structure,
    extract_layout_structure,
)

@dataclass(frozen=True)
class LayoutResult:
    structure_text: str
    changelog: str
    directions: list[str]
    changed: bool
    response: LLMResponse
    discarded: str


class LayoutAgent(BaseTableAgent):
    name = "LayoutAgent"
    profile = "Spreadsheet layout VLM"
    goal = "Incrementally extract table headers and data ranges from coordinate-labelled viewports."

    def __init__(self, vlm: BaseLLM):
        super().__init__()
        self.vlm = vlm

    def run(
        self,
        *,
        metadata_text: str,
        structure_text: str,
        image_path: Path,
        viewport_range: str,
        direction: str,
        feedback: str,
        iteration: int,
        iteration_dir: Path,
    ) -> LayoutResult:
        feedback_block = f"\nDeterministic verifier feedback:\n{feedback}\n" if feedback else ""
        prompt = LAYOUT_MAS_USER_PROMPT_TEMPLATE.format(
            metadata_text=metadata_text,
            viewport_range=viewport_range,
            direction=direction,
            structure_text=structure_text or "{}",
            feedback_block=feedback_block,
        )
        iteration_dir.joinpath("layout_prompt.txt").write_text(prompt, encoding="utf-8")
        response = self.vlm.generate_with_image(
            prompt=prompt,
            image_path=image_path,
            system_prompt=LAYOUT_MAS_SYSTEM_PROMPT,
        )
        iteration_dir.joinpath("layout_response.txt").write_text(response.content, encoding="utf-8")
        updated, discarded, directions, model_changelog = extract_layout_structure(response.content)
        directions = _normalize_remaining_directions(directions, direction)
        if not _is_valid_structure(updated):
            updated = structure_text
        else:
            updated = _union_existing_data_ranges(structure_text, updated)
        changed = bool(updated.strip()) and _canonical_yaml(updated) != _canonical_yaml(structure_text)
        changelog = model_changelog or ("Structure updated." if changed else "No change.")
        if not changed:
            changelog = "No change."
        self.remember(AgentMessage(
            sent_from=self.name,
            sent_to="deterministic_verifier",
            content=changelog,
            iteration=iteration,
            metadata={"viewport": viewport_range, "directions": directions, "changed": changed},
        ))
        return LayoutResult(updated, changelog, directions, changed, response, discarded)


def _normalize_remaining_directions(directions: list[str], current_direction: str) -> list[str]:
    current = str(current_direction).strip().lower()
    opposites = {
        "right": "left",
        "left": "right",
        "down": "up",
        "up": "down",
    }
    blocked = {current}
    opposite = opposites.get(current)
    if opposite is not None:
        blocked.add(opposite)

    normalized: list[str] = []
    for direction in directions:
        value = str(direction).strip().lower()
        if value not in opposites or value in blocked or value in normalized:
            continue
        normalized.append(value)
        if len(normalized) == 2:
            break
    return normalized


def _canonical_yaml(text: str) -> Any:
    if not text.strip():
        return None
    try:
        return yaml.safe_load(text)
    except yaml.YAMLError:
        return text.strip()


def _union_existing_data_ranges(previous_text: str, updated_text: str) -> str:
    try:
        previous = yaml.safe_load(previous_text) if previous_text.strip() else None
        updated = yaml.safe_load(updated_text) if updated_text.strip() else None
    except yaml.YAMLError:
        return updated_text
    if not isinstance(previous, dict) or not isinstance(updated, dict):
        return updated_text

    previous_headers = {
        _header_identity(path, header): header
        for path, header in _iter_structure_headers(previous)
    }
    changed = False
    for path, header in _iter_structure_headers(updated):
        prior = previous_headers.get(_header_identity(path, header))
        if not prior:
            continue
        old_range = prior.get("data_range")
        new_range = header.get("data_range")
        orientation = str(header.get("orientation") or prior.get("orientation") or "column").lower()
        unioned = _union_data_range(old_range, new_range, orientation)
        if unioned and unioned != new_range:
            header["data_range"] = unioned
            changed = True
    if not changed:
        return updated_text
    return yaml.safe_dump(updated, sort_keys=False, allow_unicode=True).strip()


def _iter_structure_headers(structure: dict[str, Any]):
    for table_key, table in structure.items():
        if not isinstance(table, dict):
            continue
        headers = table.get("headers") or []
        if not isinstance(headers, list):
            continue
        for index, header in enumerate(headers):
            yield from _iter_header_tree(header, f"{table_key}.headers[{index}]")


def _iter_header_tree(header: Any, path: str):
    if not isinstance(header, dict):
        return
    yield path, header
    sub_headers = header.get("sub_headers") or []
    if not isinstance(sub_headers, list):
        return
    for index, child in enumerate(sub_headers):
        yield from _iter_header_tree(child, f"{path}.sub_headers[{index}]")


def _header_identity(path: str, header: dict[str, Any]) -> tuple[str, str, str]:
    return (
        path,
        str(header.get("label") or "").strip().casefold(),
        str(header.get("header_range") or header.get("range") or "").strip().upper(),
    )


def _union_data_range(old_range: Any, new_range: Any, orientation: str) -> str | None:
    if not old_range or not new_range:
        return None
    try:
        old_box = range_boundaries(str(old_range))
        new_box = range_boundaries(str(new_range))
    except (TypeError, ValueError):
        return None

    if orientation == "row":
        if old_box[1] != new_box[1] or old_box[3] != new_box[3]:
            return None
    else:
        if old_box[0] != new_box[0] or old_box[2] != new_box[2]:
            return None
    if not _boxes_overlap_or_touch(old_box, new_box):
        return None
    return _box_to_range((
        min(old_box[0], new_box[0]),
        min(old_box[1], new_box[1]),
        max(old_box[2], new_box[2]),
        max(old_box[3], new_box[3]),
    ))


def _boxes_overlap_or_touch(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> bool:
    return not (
        left[2] + 1 < right[0]
        or right[2] + 1 < left[0]
        or left[3] + 1 < right[1]
        or right[3] + 1 < left[1]
    )


def _box_to_range(box: tuple[int, int, int, int]) -> str:
    min_col, min_row, max_col, max_row = box
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


__all__ = ["LayoutAgent", "LayoutResult"]

