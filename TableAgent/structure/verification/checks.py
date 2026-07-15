from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Iterator

import openpyxl
import yaml
from openpyxl.utils.cell import get_column_letter, range_boundaries


def verify_structure(workbook_path: str | Path, sheet_name: str, structure_path: str | Path) -> dict[str, Any]:
    structure = yaml.safe_load(Path(structure_path).read_text(encoding="utf-8")) or {}
    workbook = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
    errors: list[str] = []
    actions: list[str] = []
    null_fields: list[str] = []
    try:
        worksheet = workbook[sheet_name]
        used_box = _range_box(worksheet.calculate_dimension())
        for path, value in _walk_ranges(structure):
            try:
                min_col, min_row, max_col, max_row = _range_box(value)
            except (TypeError, ValueError):
                errors.append(f"{path} is not a valid A1 range: {value}")
                continue
            if min_row < 1 or min_col < 1 or max_row > 1048576 or max_col > 16384:
                errors.append(f"{path} is outside Excel worksheet bounds: {value}")
        for path, header in _walk_headers(structure):
            _check_header(worksheet, path, header, used_box, errors, actions, null_fields)
    finally:
        workbook.close()

    feedback_parts = []
    if actions:
        feedback_parts.append("Deterministic verifier repaired workbook-backed fields: " + "; ".join(actions))
    if errors:
        feedback_parts.append(
            "Deterministic verifier could not repair these fields. "
            "LayoutAgent must fill only these exact null fields with concrete A1 ranges/text: "
            + "; ".join(errors)
        )
    return {
        "status": "not_good" if errors or null_fields else "good",
        "errors": errors,
        "actions": actions,
        "null_fields": list(dict.fromkeys(null_fields)),
        "feedback": " ".join(feedback_parts) if feedback_parts else "Structure verified by workbook-backed code.",
        "repaired_structure_yaml": yaml.safe_dump(structure, sort_keys=False, allow_unicode=True).strip(),
    }


def _walk_ranges(value: Any, path: str = "") -> Iterator[tuple[str, Any]]:
    if isinstance(value, dict):
        for key, child in value.items():
            child_path = f"{path}.{key}" if path else key
            if key in {"range", "header_range", "data_range"} and child is not None:
                yield child_path, child
            else:
                yield from _walk_ranges(child, child_path)
    elif isinstance(value, list):
        for index, child in enumerate(value):
            yield from _walk_ranges(child, f"{path}[{index}]")


def _norm(value: Any) -> str:
    text = str(value or "").replace("\\n", "\n")
    return re.sub(r"\s+", "", text).casefold()


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\\n", "\n")).strip()


def _range_box(value: Any) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(str(value))
    if min_col > max_col or min_row > max_row:
        raise ValueError(f"invalid range order: {value}")
    return min_col, min_row, max_col, max_row


def _box_to_range(box: tuple[int, int, int, int]) -> str:
    min_col, min_row, max_col, max_row = box
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _box_cells(box: tuple[int, int, int, int]) -> set[tuple[int, int]]:
    min_col, min_row, max_col, max_row = box
    return {(row, col) for row in range(min_row, max_row + 1) for col in range(min_col, max_col + 1)}


def _intersects(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> bool:
    return not (left[2] < right[0] or right[2] < left[0] or left[3] < right[1] or right[3] < left[1])


def _contains(outer: tuple[int, int, int, int], inner: tuple[int, int, int, int]) -> bool:
    return outer[0] <= inner[0] and outer[1] <= inner[1] and outer[2] >= inner[2] and outer[3] >= inner[3]


def _merged_box_for(worksheet: Any, box: tuple[int, int, int, int]) -> tuple[int, int, int, int] | None:
    cells = _box_cells(box)
    for merged in worksheet.merged_cells.ranges:
        merged_box = (merged.min_col, merged.min_row, merged.max_col, merged.max_row)
        if cells & _box_cells(merged_box):
            return merged_box
    return None


def _next_text_box_right(worksheet: Any, box: tuple[int, int, int, int], used_box: tuple[int, int, int, int]):
    row = box[1]
    for col in range(box[2] + 1, used_box[2] + 1):
        value = worksheet.cell(row=row, column=col).value
        if value is not None and str(value).strip():
            candidate = (col, row, col, row)
            return _merged_box_for(worksheet, candidate) or candidate
    return None


def _effective_value(worksheet: Any, row: int, col: int) -> Any:
    value = worksheet.cell(row=row, column=col).value
    if value is not None:
        return value
    for merged in worksheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return worksheet.cell(row=merged.min_row, column=merged.min_col).value
    return None


def _cell_texts(worksheet: Any, box: tuple[int, int, int, int]) -> list[str]:
    texts = []
    seen = set()
    for row, col in sorted(_box_cells(box)):
        value = _effective_value(worksheet, row, col)
        if value is not None and str(value).strip():
            text = _clean_text(value)
            if text not in seen:
                texts.append(text)
                seen.add(text)
    return texts


def _set_null(header: dict[str, Any], path: str, field_name: str, null_fields: list[str]) -> None:
    header[field_name] = None
    null_fields.append(f"{path}.{field_name}")


def _header_at_path(root: dict[str, Any], child_path: str) -> dict[str, Any] | None:
    child: Any = root
    for token in re.findall(r"sub_headers\[(\d+)\]", child_path):
        sub_headers = child.get("sub_headers", [])
        index = int(token)
        if not isinstance(sub_headers, list) or index >= len(sub_headers):
            return None
        child = sub_headers[index]
    return child if isinstance(child, dict) else None


def _repair_label(header: dict[str, Any], texts: list[str], actions: list[str], path: str) -> None:
    if not texts:
        return
    extracted = _clean_text(" ".join(texts))
    current = str(header.get("label") or "")
    if extracted and (_norm(extracted) != _norm(current) or _clean_text(current) != current):
        actions.append(f"{path}.label corrected to workbook text {extracted!r}")
        header["label"] = extracted


def _repair_data_box(header_box, data_box, orientation, used_box):
    if orientation == "row":
        return header_box[2] + 1, header_box[1], data_box[2] if data_box else used_box[2], header_box[3]
    return header_box[0], header_box[3] + 1, header_box[2], data_box[3] if data_box else used_box[3]


def _walk_headers(structure: dict[str, Any]):
    for table_key, table in structure.items():
        if not isinstance(table, dict):
            continue
        headers = table.get("headers") or []
        if not isinstance(headers, list):
            continue
        for index, header in enumerate(headers):
            yield from _walk_header(header, f"{table_key}.headers[{index}]")


def _walk_header(header: Any, path: str):
    if not isinstance(header, dict):
        return
    yield path, header
    sub_headers = header.get("sub_headers") or []
    if isinstance(sub_headers, list):
        for index, child in enumerate(sub_headers):
            yield from _walk_header(child, f"{path}.sub_headers[{index}]")


def _header_text_boxes_in_band(worksheet: Any, band_box):
    boxes = []
    seen = set()
    min_col, min_row, max_col, max_row = band_box
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            value = _effective_value(worksheet, row, col)
            if value is None or not str(value).strip():
                continue
            box = _merged_box_for(worksheet, (col, row, col, row)) or (col, row, col, row)
            key = (box, _clean_text(value))
            if key not in seen:
                boxes.append((box, _clean_text(value)))
                seen.add(key)
    return boxes


def _visible_subheader_gaps(worksheet, parent_box, child_header_boxes, data_box, orientation):
    if not child_header_boxes:
        return []
    boxes = [item[2] for item in child_header_boxes]
    if orientation == "row":
        start_col, end_col = min(box[0] for box in boxes), max(box[2] for box in boxes)
        if data_box is not None:
            end_col = min(end_col, data_box[0] - 1)
        band_box = (start_col, parent_box[1], end_col, parent_box[3])
    else:
        start_row, end_row = min(box[1] for box in boxes), max(box[3] for box in boxes)
        if data_box is not None:
            end_row = min(end_row, data_box[1] - 1)
        band_box = (parent_box[0], start_row, parent_box[2], end_row)
    if band_box[0] > band_box[2] or band_box[1] > band_box[3]:
        return []
    return [
        (_box_to_range(text_box), text)
        for text_box, text in _header_text_boxes_in_band(worksheet, band_box)
        if not any(_contains(child_box, text_box) for child_box in boxes)
    ]


def _check_header(worksheet, path, header, used_box, errors, actions, null_fields):
    label = str(header.get("label") or "").strip()
    header_range = header.get("header_range") or header.get("range")
    data_range = header.get("data_range")
    orientation = str(header.get("orientation") or "column").strip().lower()
    if orientation not in {"row", "column"}:
        errors.append(f"{path}.orientation must be row or column: {orientation}")

    header_box = None
    data_box = None
    if header_range is not None:
        try:
            header_box = _range_box(header_range)
        except (TypeError, ValueError):
            _set_null(header, path, "header_range", null_fields)
            if data_range is not None:
                _set_null(header, path, "data_range", null_fields)
            errors.append(f"{path}.header_range is not a valid A1 range: {header_range}")
        else:
            original_header_box = header_box
            merged_box = _merged_box_for(worksheet, header_box)
            if merged_box and merged_box != header_box:
                merged_texts = _cell_texts(worksheet, merged_box)
                direct_value = worksheet.cell(row=original_header_box[1], column=original_header_box[0]).value
                if direct_value is None and label and merged_texts and not any(_norm(label) in _norm(text) for text in merged_texts):
                    next_box = _next_text_box_right(worksheet, merged_box, used_box)
                    if next_box is not None:
                        actions.append(f"{path}.header_range moved from blank merged follower {header_range} to next workbook header {_box_to_range(next_box)}")
                        header_box = next_box
                    else:
                        header_box = merged_box
                else:
                    actions.append(f"{path}.header_range expanded from {header_range} to {_box_to_range(merged_box)} using workbook merged cells")
                    header_box = merged_box
                header["header_range"] = _box_to_range(header_box)
                header_range = header["header_range"]
            if not _contains(used_box, header_box):
                errors.append(f"{path}.header_range is outside used range: {header_range}")
            texts = _cell_texts(worksheet, header_box)
            if not texts:
                _set_null(header, path, "header_range", null_fields)
                if data_range is not None:
                    _set_null(header, path, "data_range", null_fields)
                errors.append(f"{path}.header_range contains no visible header text: {header_range}")
            else:
                _repair_label(header, texts, actions, path)
            if len(texts) > 1:
                _set_null(header, path, "header_range", null_fields)
                if data_range is not None:
                    _set_null(header, path, "data_range", null_fields)
                errors.append(f"{path}.header_range contains multiple unrelated texts {texts!r}: {header_range}")

    if data_range is not None:
        try:
            data_box = _range_box(data_range)
        except (TypeError, ValueError):
            _set_null(header, path, "data_range", null_fields)
            errors.append(f"{path}.data_range is not a valid A1 range: {data_range}")
        else:
            if not _contains(used_box, data_box):
                _set_null(header, path, "data_range", null_fields)
                errors.append(f"{path}.data_range is outside used range: {data_range}")
            if header_box is not None and _intersects(data_box, header_box):
                repaired = _repair_data_box(header_box, data_box, orientation, used_box)
                if _contains(used_box, repaired):
                    actions.append(f"{path}.data_range shifted from {data_range} to {_box_to_range(repaired)} to exclude header cells")
                    data_box = repaired
                    header["data_range"] = _box_to_range(data_box)
                    data_range = header["data_range"]
                else:
                    _set_null(header, path, "data_range", null_fields)
                    errors.append(f"{path}.data_range overlaps its header_range: {data_range} vs {header_range}")
            if header_box is not None and data_box is not None:
                if orientation == "column" and (data_box[0] != header_box[0] or data_box[2] != header_box[2]):
                    repaired = (header_box[0], data_box[1], header_box[2], data_box[3])
                    actions.append(f"{path}.data_range realigned from {data_range} to {_box_to_range(repaired)} to match header columns")
                    data_box = repaired
                    header["data_range"] = _box_to_range(data_box)
                    data_range = header["data_range"]
                if orientation == "row" and (data_box[1] != header_box[1] or data_box[3] != header_box[3]):
                    repaired = (data_box[0], header_box[1], data_box[2], header_box[3])
                    actions.append(f"{path}.data_range realigned from {data_range} to {_box_to_range(repaired)} to match header rows")
                    data_box = repaired
                    header["data_range"] = _box_to_range(data_box)
                    data_range = header["data_range"]
            if data_box is not None and not _cell_texts(worksheet, data_box):
                if header_box is not None:
                    repaired = _repair_data_box(header_box, data_box, orientation, used_box)
                    if _contains(used_box, repaired) and _cell_texts(worksheet, repaired):
                        actions.append(f"{path}.data_range repaired from {data_range} to {_box_to_range(repaired)} using the verified header span")
                        data_box = repaired
                        header["data_range"] = _box_to_range(data_box)
                        data_range = header["data_range"]
                    else:
                        _set_null(header, path, "data_range", null_fields)
                        errors.append(f"{path}.data_range contains no visible data: {data_range}")
                else:
                    _set_null(header, path, "data_range", null_fields)
                    errors.append(f"{path}.data_range contains no visible data: {data_range}")

    sub_headers = header.get("sub_headers") or []
    child_header_boxes = []
    child_data_boxes = []
    if isinstance(sub_headers, list):
        for index, child in enumerate(sub_headers):
            child_path = f"{path}.sub_headers[{index}]"
            if not isinstance(child, dict):
                errors.append(f"{child_path} must be a mapping")
                continue
            child_header_range = child.get("header_range") or child.get("range")
            child_data_range = child.get("data_range")
            if child_data_range is None:
                null_fields.append(f"{child_path}.data_range")
                errors.append(f"{child_path}.data_range is required when a sub-header is declared")
            if child_header_range is not None:
                try:
                    child_header_boxes.append((child_path, child_header_range, _range_box(child_header_range)))
                except (TypeError, ValueError):
                    pass
            if child_data_range is not None:
                try:
                    child_data_boxes.append((child_path, child_data_range, _range_box(child_data_range)))
                except (TypeError, ValueError):
                    pass

    if data_box is not None:
        for child_path, child_header_range, child_header_box in child_header_boxes:
            if _intersects(data_box, child_header_box):
                repaired = _repair_data_box(child_header_box, data_box, orientation, used_box)
                if _contains(used_box, repaired):
                    actions.append(f"{path}.data_range shifted from {data_range} to {_box_to_range(repaired)} to exclude {child_path}.header_range")
                    data_box = repaired
                    header["data_range"] = _box_to_range(data_box)
                    data_range = header["data_range"]
                else:
                    _set_null(header, path, "data_range", null_fields)
                    errors.append(f"{path}.data_range overlaps {child_path}.header_range: {data_range} vs {child_header_range}")
        for child_path, child_data_range, child_data_box in child_data_boxes:
            if not _contains(data_box, child_data_box):
                child = _header_at_path(header, child_path)
                if child is not None:
                    _set_null(child, child_path, "data_range", null_fields)
                errors.append(f"{child_path}.data_range is not contained by parent {path}.data_range: {child_data_range} vs {data_range}")

    if header_box is not None and child_header_boxes:
        for child_path, child_header_range, child_header_box in child_header_boxes:
            invalid = False
            if orientation == "column":
                invalid = child_header_box[0] < header_box[0] or child_header_box[2] > header_box[2] or child_header_box[1] <= header_box[3]
            else:
                invalid = child_header_box[1] < header_box[1] or child_header_box[3] > header_box[3] or child_header_box[0] <= header_box[2]
            if invalid:
                child = _header_at_path(header, child_path)
                if child is not None:
                    _set_null(child, child_path, "header_range", null_fields)
                errors.append(f"{child_path}.header_range is outside the parent header hierarchy: {child_header_range} vs {header_range}")
        gaps = _visible_subheader_gaps(worksheet, header_box, child_header_boxes, data_box, orientation)
        if gaps:
            formatted = ", ".join(f"{cell}={text!r}" for cell, text in gaps[:8])
            errors.append(f"{path}.sub_headers do not cover visible layered header cells under parent {header_range}: {formatted}")
