from __future__ import annotations

import json
import math
import shutil
import subprocess
import tempfile
import threading
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from TableAgent.configs import TableAgentConfig
from TableAgent.rendering.converter import WorkbookConversion, sample_to_xlsx
from TableAgent.rendering.image_utils import (
    _generate_image_tiles,
    _resize_image_file_to_fit,
)
from TableAgent.schema import EvalSample


_PDFIUM_LOCK = threading.Lock()


@dataclass(frozen=True)
class RenderResult:
    image_path: Path
    html_path: Path | None
    width: int | None
    height: int | None
    browser_path: Path


class WorkbookRenderer:
    def __init__(self, settings: TableAgentConfig, logger: Any):
        self.settings = settings
        self.logger = logger

    def sample_to_image(self, sample: EvalSample, sample_dir: Path):
        workbook_path = sample_dir / "table.xlsx"
        source_path = Path(sample.table_path) if sample.table_path else None
        if source_path and source_path.is_file() and source_path.suffix.lower() == ".xlsx":
            shutil.copy2(source_path, workbook_path)
            from openpyxl import load_workbook

            source_workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            try:
                sheet_names = list(source_workbook.sheetnames)
            finally:
                source_workbook.close()
            workbook = WorkbookConversion(workbook_path, "xlsx", sheet_names)
        else:
            workbook = sample_to_xlsx(sample, workbook_path)
        render_result = self.render_workbook_range(
            workbook_path,
            workbook.sheet_names[0] if workbook.sheet_names else None,
            None,
            sample_dir / "table.png",
        )
        tiles = self.postprocess_image(render_result.image_path)
        metadata = {"render": _read_render_metadata(render_result.image_path)}
        if tiles:
            metadata["image_tiles"] = tiles
        (sample_dir / "metadata.json").write_text(json.dumps(metadata, ensure_ascii=False), encoding="utf-8")
        return workbook, render_result

    def source_to_image(self, source_path: Path, sheet_name: str, image_path: Path, html_path: Path) -> list[dict[str, Any]]:
        if not image_path.is_file():
            self.render_workbook_range(source_path, sheet_name, None, image_path)
        return self.postprocess_image(image_path)

    def source_viewport_to_image(
        self,
        source_path: Path,
        sheet_name: str,
        cell_range: str,
        image_path: Path,
    ) -> RenderResult:
        result = self.render_workbook_range(source_path, sheet_name, cell_range, image_path)
        self.postprocess_image(image_path)
        return result

    def render_workbook_range(
        self,
        workbook_path: Path,
        sheet_name: str | None,
        cell_range: str | None,
        image_path: Path,
    ) -> RenderResult:
        _render_xlsx_range_with_libreoffice(
            workbook_path,
            sheet_name,
            cell_range,
            image_path,
            libreoffice_path=self.settings.libreoffice_path,
            resolution=self.settings.libreoffice_image_resolution,
            timeout_seconds=self.settings.render_timeout_seconds,
            show_coordinates=self.settings.workbook_show_coordinates,
        )
        result = _image_render_result(image_path, browser_path=Path("libreoffice"))
        _write_render_metadata(
            workbook_path=workbook_path,
            sheet_name=sheet_name,
            cell_range=cell_range,
            image_path=image_path,
            result=result,
            show_coordinates=self.settings.workbook_show_coordinates,
        )
        return result

    def postprocess_image(self, image_path: Path) -> list[dict[str, Any]]:
        tiles = []
        if self.settings.image_tile_size is not None:
            tiles = _generate_image_tiles(
                image_path,
                self.settings.image_tile_size,
                self.settings.image_tile_overlap,
                logger=self.logger,
            )
        _resize_image_file_to_fit(
            image_path,
            self.settings.max_image_dimension,
            self.settings.max_image_pixels,
            logger=self.logger,
        )
        return tiles


def _render_xlsx_range_with_libreoffice(
    workbook_path: Path,
    sheet_name: str | None,
    cell_range: str | None,
    image_path: Path,
    *,
    libreoffice_path: Path | str | None = None,
    resolution: int = 192,
    timeout_seconds: float = 60,
    show_coordinates: bool = True,
) -> None:
    try:
        import openpyxl
        import pypdfium2
        from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
    except ImportError as exc:
        raise RuntimeError(
            "LibreOffice workbook rendering requires openpyxl and pypdfium2. "
            "Install it before running TableAgent workbook image rendering."
        ) from exc

    image_path.parent.mkdir(parents=True, exist_ok=True)
    soffice = _resolve_libreoffice_path(libreoffice_path)
    with tempfile.TemporaryDirectory(prefix="tableagent_lo_") as tmp:
        tmpdir = Path(tmp)
        prepared_path = tmpdir / workbook_path.name
        pdf_path = prepared_path.with_suffix(".pdf")
        profile_dir = tmpdir / "lo_profile"
        workbook = openpyxl.load_workbook(workbook_path)
        try:
            worksheet = workbook.worksheets[0] if sheet_name is None else workbook[sheet_name]
            workbook.active = workbook.worksheets.index(worksheet)
            for sheet in workbook.worksheets:
                sheet.sheet_state = "visible" if sheet is worksheet else "hidden"
            if cell_range:
                worksheet.print_area = cell_range
            worksheet.print_options.headings = bool(show_coordinates)
            worksheet.print_options.gridLines = True
            worksheet.page_margins.left = 0
            worksheet.page_margins.right = 0
            worksheet.page_margins.top = 0
            worksheet.page_margins.bottom = 0
            worksheet.sheet_properties.pageSetUpPr.fitToPage = True
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = 1
            _ensure_visible_dimensions(worksheet, RowDimension, ColumnDimension)
            _expand_dimensions_to_fit(worksheet)
            workbook.save(prepared_path)
        finally:
            workbook.close()

        command = [
            str(soffice),
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--nodefault",
            "--nolockcheck",
            f"-env:UserInstallation={_libreoffice_file_url(profile_dir)}",
            "--convert-to",
            "pdf:calc_pdf_Export",
            "--outdir",
            str(tmpdir),
            str(prepared_path),
        ]
        completed = subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            timeout=max(1, float(timeout_seconds)),
        )
        if completed.returncode != 0 or not pdf_path.is_file():
            raise RuntimeError(
                "LibreOffice failed to export workbook to PDF: "
                f"{completed.stderr.strip() or completed.stdout.strip()}"
            )

        # PDFium is not thread-safe, so concurrent sample workers must not
        # enter any part of its document/page/bitmap lifecycle together.
        with _PDFIUM_LOCK:
            pdf = pypdfium2.PdfDocument(str(pdf_path))
            try:
                page = pdf[0]
                try:
                    render_resolution = max(384 if cell_range else 96, int(resolution))
                    bitmap = page.render(scale=render_resolution / 72)
                    try:
                        bitmap.to_pil().save(image_path)
                    finally:
                        bitmap.close()
                finally:
                    page.close()
            finally:
                pdf.close()
        _trim_image_file_to_content(image_path)


def _resolve_libreoffice_path(libreoffice_path: Path | str | None) -> Path:
    candidates = [
        str(libreoffice_path or "").strip(),
        shutil.which("soffice") or "",
        shutil.which("libreoffice") or "",
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for candidate in candidates:
        if candidate and Path(candidate).is_file():
            return Path(candidate)
    raise FileNotFoundError(
        "LibreOffice executable not found. Set table_agent.libreoffice_path "
        "or add soffice/libreoffice to PATH."
    )


def _libreoffice_file_url(path: Path) -> str:
    return path.resolve().as_uri()


def _ensure_visible_dimensions(worksheet: Any, row_dimension_cls: Any, column_dimension_cls: Any) -> None:
    from openpyxl.utils import get_column_letter

    for row_index in range(1, worksheet.max_row + 1):
        if row_index not in worksheet.row_dimensions:
            worksheet.row_dimensions[row_index] = row_dimension_cls(worksheet, index=row_index)
    for column_index in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        if column_letter not in worksheet.column_dimensions:
            worksheet.column_dimensions[column_letter] = column_dimension_cls(worksheet, index=column_letter)


def _expand_dimensions_to_fit(worksheet: Any) -> None:
    """Expand worksheet dimensions so cell text is present in the exported image."""
    from openpyxl.utils import get_column_letter

    required_widths: dict[int, float] = {}
    merged_by_cell: dict[tuple[int, int], Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                merged_by_cell[(row, column)] = merged_range

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None or not str(cell.value).strip():
                continue
            lines = str(cell.value).splitlines() or [""]
            required = max(_display_width(line) for line in lines) + 2
            merged_range = merged_by_cell.get((cell.row, cell.column))
            if merged_range is None:
                required_widths[cell.column] = max(required_widths.get(cell.column, 0), required)
                continue
            current = sum(
                float(worksheet.column_dimensions[get_column_letter(column)].width or 13)
                for column in range(merged_range.min_col, merged_range.max_col + 1)
            )
            required_widths[merged_range.min_col] = max(
                required_widths.get(merged_range.min_col, 0),
                float(worksheet.column_dimensions[get_column_letter(merged_range.min_col)].width or 13)
                + max(0, required - current),
            )

    for column, required in required_widths.items():
        letter = get_column_letter(column)
        dimension = worksheet.column_dimensions[letter]
        dimension.width = max(float(dimension.width or 13), required)

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None or not str(cell.value).strip():
                continue
            lines = str(cell.value).splitlines() or [""]
            merged_range = merged_by_cell.get((cell.row, cell.column))
            columns = (
                range(merged_range.min_col, merged_range.max_col + 1)
                if merged_range is not None
                else range(cell.column, cell.column + 1)
            )
            available_width = sum(
                float(worksheet.column_dimensions[get_column_letter(column)].width or 13)
                for column in columns
            )
            wrap = bool(cell.alignment.wrap_text)
            wrapped_lines = sum(
                max(1, math.ceil(_display_width(line) / max(1, available_width)))
                for line in lines
            ) if wrap else len(lines)
            if wrapped_lines <= 1:
                continue
            font_size = float(cell.font.sz or 11)
            required_height = wrapped_lines * font_size * 1.25 + 2
            dimension = worksheet.row_dimensions[cell.row]
            dimension.height = max(float(dimension.height or 15), required_height)


def _display_width(value: str) -> int:
    return sum(2 if unicodedata.east_asian_width(char) in {"W", "F"} else 1 for char in value)


def _image_render_result(image_path: Path, *, browser_path: Path) -> RenderResult:
    from PIL import Image

    with Image.open(image_path) as image:
        width, height = image.size
    return RenderResult(
        image_path=image_path,
        html_path=None,
        width=width,
        height=height,
        browser_path=browser_path,
    )


def _trim_image_file_to_content(image_path: Path, *, padding: int = 6) -> None:
    from PIL import Image, ImageChops

    with Image.open(image_path) as image:
        rgb = image.convert("RGB")
        background = Image.new("RGB", rgb.size, "white")
        bbox = ImageChops.difference(rgb, background).getbbox()
        if bbox is None:
            return
        left = max(0, bbox[0] - padding)
        top = max(0, bbox[1] - padding)
        right = min(rgb.width, bbox[2] + padding)
        bottom = min(rgb.height, bbox[3] + padding)
        rgb.crop((left, top, right, bottom)).save(image_path)


def _write_render_metadata(
    *,
    workbook_path: Path,
    sheet_name: str | None,
    cell_range: str | None,
    image_path: Path,
    result: RenderResult,
    show_coordinates: bool,
) -> None:
    from openpyxl.utils.cell import range_boundaries

    payload: dict[str, Any] = {
        "workbook_path": str(workbook_path.resolve()),
        "sheet_name": sheet_name,
        "cell_range": cell_range,
        "image_width": result.width,
        "image_height": result.height,
        "renderer": "libreoffice",
        "show_coordinates": show_coordinates,
    }
    if cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        payload["bounds"] = {
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
        }
    image_path.with_suffix(".metadata.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _read_render_metadata(image_path: Path) -> dict[str, Any] | None:
    metadata_path = image_path.with_suffix(".metadata.json")
    if not metadata_path.is_file():
        return None
    return json.loads(metadata_path.read_text(encoding="utf-8"))
