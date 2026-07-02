from __future__ import annotations

import json
import os
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import yaml
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from TableAgent.prompts import (
    LAYOUT_MAS_SYSTEM_PROMPT,
    LAYOUT_MAS_USER_PROMPT_TEMPLATE,
    VERIFICATION_MAS_SYSTEM_PROMPT,
    VERIFICATION_MAS_USER_PROMPT_TEMPLATE,
)
from utils.llm.base import BaseLLM, LLMResponse

from TableAgent.perception.structure import (
    _is_valid_structure,
    _parse_yaml_mapping,
    extract_layout_structure,
)


@dataclass(frozen=True)
class AgentMessage:
    sent_from: str
    sent_to: str
    content: str
    iteration: int
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass
class AgentMemory:
    messages: list[AgentMessage] = field(default_factory=list)

    def add(self, message: AgentMessage) -> None:
        self.messages.append(message)


class BaseTableAgent:
    name = "Agent"
    profile = ""
    goal = ""

    def __init__(self):
        self.memory = AgentMemory()

    def remember(self, message: AgentMessage) -> None:
        self.memory.add(message)


@dataclass(frozen=True)
class LayoutResult:
    structure_text: str
    changelog: str
    directions: list[str]
    changed: bool
    response: LLMResponse
    discarded: str


class LayoutAgent(BaseTableAgent):
    name = "LayoutAgent"
    profile = "Spreadsheet layout VLM"
    goal = "Incrementally extract table headers and data ranges from coordinate-labelled viewports."

    def __init__(self, vlm: BaseLLM):
        super().__init__()
        self.vlm = vlm

    def run(
        self,
        *,
        metadata_text: str,
        structure_text: str,
        image_path: Path,
        viewport_range: str,
        direction: str,
        feedback: str,
        iteration: int,
        iteration_dir: Path,
    ) -> LayoutResult:
        feedback_block = f"\nVerificationAgent feedback:\n{feedback}\n" if feedback else ""
        prompt = LAYOUT_MAS_USER_PROMPT_TEMPLATE.format(
            metadata_text=metadata_text,
            viewport_range=viewport_range,
            direction=direction,
            structure_text=structure_text or "{}",
            feedback_block=feedback_block,
        )
        iteration_dir.joinpath("layout_prompt.txt").write_text(prompt, encoding="utf-8")
        response = self.vlm.generate_with_image(
            prompt=prompt,
            image_path=image_path,
            system_prompt=LAYOUT_MAS_SYSTEM_PROMPT,
        )
        iteration_dir.joinpath("layout_response.txt").write_text(response.content, encoding="utf-8")
        updated, discarded, directions, model_changelog = extract_layout_structure(response.content)
        directions = _normalize_remaining_directions(directions, direction)
        if not _is_valid_structure(updated):
            updated = structure_text
        else:
            updated = _union_existing_data_ranges(structure_text, updated)
        changed = bool(updated.strip()) and _canonical_yaml(updated) != _canonical_yaml(structure_text)
        changelog = model_changelog or ("Structure updated." if changed else "No change.")
        if not changed:
            changelog = "No change."
        self.remember(AgentMessage(
            sent_from=self.name,
            sent_to="VerificationAgent",
            content=changelog,
            iteration=iteration,
            metadata={"viewport": viewport_range, "directions": directions, "changed": changed},
        ))
        return LayoutResult(updated, changelog, directions, changed, response, discarded)


def _normalize_remaining_directions(directions: list[str], current_direction: str) -> list[str]:
    current = str(current_direction).strip().lower()
    opposites = {
        "right": "left",
        "left": "right",
        "down": "up",
        "up": "down",
    }
    blocked = {current}
    opposite = opposites.get(current)
    if opposite is not None:
        blocked.add(opposite)

    normalized: list[str] = []
    for direction in directions:
        value = str(direction).strip().lower()
        if value not in opposites or value in blocked or value in normalized:
            continue
        normalized.append(value)
        if len(normalized) == 2:
            break
    return normalized


@dataclass(frozen=True)
class VerificationResult:
    status: str
    feedback: str
    null_fields: list[str]
    report: dict[str, Any]
    response: LLMResponse
    structure_text: str

    @property
    def is_good(self) -> bool:
        return self.status == "good"


class VerificationAgent(BaseTableAgent):
    name = "VerificationAgent"
    profile = "Spreadsheet structure verifier"
    goal = "Verify header and data ranges with executable checks and semantic review."

    def __init__(self, llm: BaseLLM):
        super().__init__()
        self.llm = llm

    def run(
        self,
        *,
        workbook_path: Path,
        sheet_name: str,
        metadata_text: str,
        structure_text: str,
        changelog: str,
        viewport_range: str,
        iteration: int,
        iteration_dir: Path,
    ) -> VerificationResult:
        verifier_path = iteration_dir / "verification.py"
        verifier_path.write_text(_VERIFIER_CODE, encoding="utf-8")
        report = _execute_verifier(verifier_path, workbook_path, sheet_name, iteration_dir / "structure_after.yaml")
        if not _is_valid_structure(structure_text):
            report = {
                "status": "not_good",
                "errors": ["Candidate structure is empty or invalid."],
            }
        repaired_structure_text = str(report.get("repaired_structure_yaml") or structure_text)
        if repaired_structure_text != structure_text:
            (iteration_dir / "structure_after.yaml").write_text(repaired_structure_text, encoding="utf-8")
            (iteration_dir / "structure_repaired.yaml").write_text(repaired_structure_text, encoding="utf-8")
            structure_text = repaired_structure_text
        report_text = json.dumps(
            {key: value for key, value in report.items() if key != "repaired_structure_yaml"},
            ensure_ascii=False,
            indent=2,
        )
        (iteration_dir / "verification_output.json").write_text(report_text, encoding="utf-8")
        prompt = VERIFICATION_MAS_USER_PROMPT_TEMPLATE.format(
            metadata_text=metadata_text,
            viewport_range=viewport_range,
            structure_text=structure_text,
            changelog=changelog,
            verification_report=report_text,
        )
        (iteration_dir / "verification_prompt.txt").write_text(prompt, encoding="utf-8")
        response = self.llm.generate(prompt=prompt, system_prompt=VERIFICATION_MAS_SYSTEM_PROMPT)
        (iteration_dir / "verification_response.yaml").write_text(response.content, encoding="utf-8")
        parsed = _parse_yaml_mapping(response.content)
        parsed_status = parsed.get("status")
        status = str(parsed_status or "not_good").strip().lower()
        feedback = str(parsed.get("feedback") or response.content).strip()
        if parsed_status is None and report.get("status") == "good":
            status = "good"
            feedback = str(report.get("feedback") or "Deterministic verification passed.")
        null_fields = report.get("null_fields") or parsed.get("null_fields") or []
        if not isinstance(null_fields, list):
            null_fields = []
        semantic_structure_text, semantic_discarded, _, _ = extract_layout_structure(response.content)
        semantic_updated = False
        if semantic_structure_text and _is_valid_structure(semantic_structure_text):
            semantic_structure_text = _union_existing_data_ranges(structure_text, semantic_structure_text)
            semantic_updated = _canonical_yaml(semantic_structure_text) != _canonical_yaml(structure_text)
            if semantic_updated:
                structure_text = semantic_structure_text
                (iteration_dir / "structure_after.yaml").write_text(structure_text, encoding="utf-8")
                (iteration_dir / "structure_semantic.yaml").write_text(structure_text, encoding="utf-8")
                if semantic_discarded:
                    (iteration_dir / "verification_discarded.txt").write_text(semantic_discarded, encoding="utf-8")
                semantic_report = _execute_verifier(verifier_path, workbook_path, sheet_name, iteration_dir / "structure_after.yaml")
                semantic_report_text = json.dumps(
                    {key: value for key, value in semantic_report.items() if key != "repaired_structure_yaml"},
                    ensure_ascii=False,
                    indent=2,
                )
                (iteration_dir / "verification_output_after_semantic.json").write_text(
                    semantic_report_text,
                    encoding="utf-8",
                )
                if semantic_report.get("status") == "good":
                    report = semantic_report
                    null_fields = report.get("null_fields") or parsed.get("null_fields") or []
                    if not isinstance(null_fields, list):
                        null_fields = []
        semantic_accepts_update = (
            semantic_updated
            and status == "good"
            and not report.get("tool_error")
        )
        if report.get("status") != "good" and not semantic_accepts_update:
            status = "not_good"
            feedback = str(report.get("feedback") or "; ".join(report.get("errors") or [feedback]))
        if status not in {"good", "not_good"}:
            status = "not_good"
        self.remember(AgentMessage(
            sent_from=self.name,
            sent_to="LayoutAgent" if status == "not_good" else "orchestrator",
            content=feedback,
            iteration=iteration,
            metadata={"viewport": viewport_range, "status": status},
        ))
        return VerificationResult(status, feedback, [str(value) for value in null_fields], report, response, structure_text)


class QAAgent(BaseTableAgent):
    name = "QAAgent"
    profile = "Table question answering agent"
    goal = "Produce the final answer after layout traversal is complete."

    def __init__(self, llm: BaseLLM, system_prompt: str):
        super().__init__()
        self.llm = llm
        self.system_prompt = system_prompt

    def run(self, *, prompt: str, image_path: Path | None = None, fallback_prompt: str | None = None) -> LLMResponse:
        generate_with_image = getattr(self.llm, "generate_with_image", None)
        if image_path is not None and callable(generate_with_image):
            return generate_with_image(prompt=prompt, image_path=image_path, system_prompt=self.system_prompt)
        return self.llm.generate(prompt=fallback_prompt or prompt, system_prompt=self.system_prompt)


def _canonical_yaml(text: str) -> Any:
    if not text.strip():
        return None
    try:
        return yaml.safe_load(text)
    except yaml.YAMLError:
        return text.strip()


def _union_existing_data_ranges(previous_text: str, updated_text: str) -> str:
    try:
        previous = yaml.safe_load(previous_text) if previous_text.strip() else None
        updated = yaml.safe_load(updated_text) if updated_text.strip() else None
    except yaml.YAMLError:
        return updated_text
    if not isinstance(previous, dict) or not isinstance(updated, dict):
        return updated_text

    previous_headers = {
        _header_identity(path, header): header
        for path, header in _iter_structure_headers(previous)
    }
    changed = False
    for path, header in _iter_structure_headers(updated):
        prior = previous_headers.get(_header_identity(path, header))
        if not prior:
            continue
        old_range = prior.get("data_range")
        new_range = header.get("data_range")
        orientation = str(header.get("orientation") or prior.get("orientation") or "column").lower()
        unioned = _union_data_range(old_range, new_range, orientation)
        if unioned and unioned != new_range:
            header["data_range"] = unioned
            changed = True
    if not changed:
        return updated_text
    return yaml.safe_dump(updated, sort_keys=False, allow_unicode=True).strip()


def _iter_structure_headers(structure: dict[str, Any]):
    for table_key, table in structure.items():
        if not isinstance(table, dict):
            continue
        headers = table.get("headers") or []
        if not isinstance(headers, list):
            continue
        for index, header in enumerate(headers):
            yield from _iter_header_tree(header, f"{table_key}.headers[{index}]")


def _iter_header_tree(header: Any, path: str):
    if not isinstance(header, dict):
        return
    yield path, header
    sub_headers = header.get("sub_headers") or []
    if not isinstance(sub_headers, list):
        return
    for index, child in enumerate(sub_headers):
        yield from _iter_header_tree(child, f"{path}.sub_headers[{index}]")


def _header_identity(path: str, header: dict[str, Any]) -> tuple[str, str, str]:
    return (
        path,
        str(header.get("label") or "").strip().casefold(),
        str(header.get("header_range") or header.get("range") or "").strip().upper(),
    )


def _union_data_range(old_range: Any, new_range: Any, orientation: str) -> str | None:
    if not old_range or not new_range:
        return None
    try:
        old_box = range_boundaries(str(old_range))
        new_box = range_boundaries(str(new_range))
    except (TypeError, ValueError):
        return None

    if orientation == "row":
        if old_box[1] != new_box[1] or old_box[3] != new_box[3]:
            return None
    else:
        if old_box[0] != new_box[0] or old_box[2] != new_box[2]:
            return None
    if not _boxes_overlap_or_touch(old_box, new_box):
        return None
    return _box_to_range((
        min(old_box[0], new_box[0]),
        min(old_box[1], new_box[1]),
        max(old_box[2], new_box[2]),
        max(old_box[3], new_box[3]),
    ))


def _boxes_overlap_or_touch(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> bool:
    return not (
        left[2] + 1 < right[0]
        or right[2] + 1 < left[0]
        or left[3] + 1 < right[1]
        or right[3] + 1 < left[1]
    )


def _box_to_range(box: tuple[int, int, int, int]) -> str:
    min_col, min_row, max_col, max_row = box
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _execute_verifier(
    verifier_path: Path,
    workbook_path: Path,
    sheet_name: str,
    structure_path: Path,
) -> dict[str, Any]:
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"
    try:
        result = subprocess.run(
            [sys.executable, str(verifier_path), str(workbook_path), sheet_name, str(structure_path)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
            timeout=30,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        return {
            "status": "not_good",
            "errors": [f"Verifier execution failed: {exc}"],
            "tool_error": True,
            "feedback": f"Deterministic verifier tool failed before validating the structure: {exc}",
        }
    if result.returncode != 0:
        return {
            "status": "not_good",
            "errors": [result.stderr.strip() or "Verifier failed."],
            "tool_error": True,
            "feedback": "Deterministic verifier tool failed before validating the structure.",
        }
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        return {
            "status": "not_good",
            "errors": ["Verifier returned invalid JSON."],
            "tool_error": True,
            "feedback": "Deterministic verifier returned invalid JSON and did not validate the structure.",
        }


_VERIFIER_CODE = '''from __future__ import annotations

import json
import re
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except AttributeError:
    pass

import openpyxl
import yaml
from openpyxl.utils.cell import get_column_letter, range_boundaries


def walk(value, path=""):
    if isinstance(value, dict):
        for key, child in value.items():
            child_path = f"{path}.{key}" if path else key
            if key in {"range", "header_range", "data_range"} and child is not None:
                yield child_path, child
            else:
                yield from walk(child, child_path)
    elif isinstance(value, list):
        for index, child in enumerate(value):
            yield from walk(child, f"{path}[{index}]")


def norm(value):
    text = str(value or "").replace("\\\\n", "\\n")
    return re.sub(r"\\s+", "", text).casefold()


def clean_text(value):
    return re.sub(r"\\s+", " ", str(value or "").replace("\\\\n", "\\n")).strip()


def range_box(value):
    min_col, min_row, max_col, max_row = range_boundaries(str(value))
    if min_col > max_col or min_row > max_row:
        raise ValueError(f"invalid range order: {value}")
    return min_col, min_row, max_col, max_row


def box_to_range(box):
    min_col, min_row, max_col, max_row = box
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def box_cells(box):
    min_col, min_row, max_col, max_row = box
    return {
        (row, col)
        for row in range(min_row, max_row + 1)
        for col in range(min_col, max_col + 1)
    }


def intersects(left, right):
    return not (
        left[2] < right[0]
        or right[2] < left[0]
        or left[3] < right[1]
        or right[3] < left[1]
    )


def contains(outer, inner):
    return outer[0] <= inner[0] and outer[1] <= inner[1] and outer[2] >= inner[2] and outer[3] >= inner[3]


def merged_box_for(worksheet, box):
    cells = box_cells(box)
    for merged in worksheet.merged_cells.ranges:
        merged_box = (merged.min_col, merged.min_row, merged.max_col, merged.max_row)
        if cells & box_cells(merged_box):
            return merged_box
    return None


def next_text_box_right(worksheet, box, used_box):
    row = box[1]
    for col in range(box[2] + 1, used_box[2] + 1):
        value = worksheet.cell(row=row, column=col).value
        if value is not None and str(value).strip():
            candidate = (col, row, col, row)
            return merged_box_for(worksheet, candidate) or candidate
    return None


def effective_value(worksheet, row, col):
    value = worksheet.cell(row=row, column=col).value
    if value is not None:
        return value
    for merged in worksheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return worksheet.cell(row=merged.min_row, column=merged.min_col).value
    return None


def cell_texts(worksheet, box):
    texts = []
    seen = set()
    for row, col in sorted(box_cells(box)):
        value = effective_value(worksheet, row, col)
        if value is not None and str(value).strip():
            text = clean_text(value)
            if text not in seen:
                texts.append(text)
                seen.add(text)
    return texts


def set_null(header, path, field_name, null_fields):
    header[field_name] = None
    null_fields.append(f"{path}.{field_name}")


def repair_label(header, texts, actions, path):
    if not texts:
        return
    extracted = clean_text(" ".join(texts))
    current = str(header.get("label") or "")
    if extracted and (norm(extracted) != norm(current) or clean_text(current) != current):
        actions.append(f"{path}.label corrected to workbook text {extracted!r}")
        header["label"] = extracted


def repair_data_box(header_box, data_box, orientation, used_box):
    if orientation == "row":
        min_col = header_box[2] + 1
        max_col = data_box[2] if data_box else used_box[2]
        return min_col, header_box[1], max_col, header_box[3]
    min_row = header_box[3] + 1
    max_row = data_box[3] if data_box else used_box[3]
    return header_box[0], min_row, header_box[2], max_row


def walk_headers(structure):
    for table_key, table in structure.items():
        if not isinstance(table, dict):
            continue
        headers = table.get("headers") or []
        if not isinstance(headers, list):
            continue
        for index, header in enumerate(headers):
            yield from walk_header(header, f"{table_key}.headers[{index}]")


def walk_header(header, path):
    if not isinstance(header, dict):
        return
    yield path, header
    sub_headers = header.get("sub_headers") or []
    if not isinstance(sub_headers, list):
        return
    for index, child in enumerate(sub_headers):
        yield from walk_header(child, f"{path}.sub_headers[{index}]")


def check_header(worksheet, path, header, used_box, errors, actions, null_fields):
    label = str(header.get("label") or "").strip()
    header_range = header.get("header_range") or header.get("range")
    data_range = header.get("data_range")
    orientation = str(header.get("orientation") or "column").strip().lower()
    if orientation not in {"row", "column"}:
        errors.append(f"{path}.orientation must be row or column: {orientation}")

    header_box = None
    data_box = None
    if header_range is not None:
        try:
            header_box = range_box(header_range)
        except (TypeError, ValueError):
            set_null(header, path, "header_range", null_fields)
            if data_range is not None:
                set_null(header, path, "data_range", null_fields)
            errors.append(f"{path}.header_range is not a valid A1 range: {header_range}")
        else:
            original_header_box = header_box
            merged_box = merged_box_for(worksheet, header_box)
            if merged_box and merged_box != header_box:
                merged_texts = cell_texts(worksheet, merged_box)
                direct_value = worksheet.cell(row=original_header_box[1], column=original_header_box[0]).value
                if (
                    direct_value is None
                    and label
                    and merged_texts
                    and not any(norm(label) in norm(text) for text in merged_texts)
                ):
                    next_box = next_text_box_right(worksheet, merged_box, used_box)
                    if next_box is not None:
                        actions.append(
                            f"{path}.header_range moved from blank merged follower {header_range} "
                            f"to next workbook header {box_to_range(next_box)}"
                        )
                        header_box = next_box
                    else:
                        header_box = merged_box
                else:
                    actions.append(
                        f"{path}.header_range expanded from {header_range} to {box_to_range(merged_box)} "
                        "using workbook merged cells"
                    )
                    header_box = merged_box
                header["header_range"] = box_to_range(header_box)
                header_range = header["header_range"]
            if not contains(used_box, header_box):
                errors.append(f"{path}.header_range is outside used range: {header_range}")
            texts = cell_texts(worksheet, header_box)
            normalized_label = norm(label)
            if not texts:
                set_null(header, path, "header_range", null_fields)
                if data_range is not None:
                    set_null(header, path, "data_range", null_fields)
                errors.append(f"{path}.header_range contains no visible header text: {header_range}")
            elif normalized_label and not any(normalized_label in norm(text) for text in texts):
                repair_label(header, texts, actions, path)
            else:
                repair_label(header, texts, actions, path)
            if len(texts) > 1:
                set_null(header, path, "header_range", null_fields)
                if data_range is not None:
                    set_null(header, path, "data_range", null_fields)
                errors.append(f"{path}.header_range contains multiple unrelated texts {texts!r}: {header_range}")

    if data_range is not None:
        try:
            data_box = range_box(data_range)
        except (TypeError, ValueError):
            set_null(header, path, "data_range", null_fields)
            errors.append(f"{path}.data_range is not a valid A1 range: {data_range}")
        else:
            if not contains(used_box, data_box):
                set_null(header, path, "data_range", null_fields)
                errors.append(f"{path}.data_range is outside used range: {data_range}")
            if header_box is not None and intersects(data_box, header_box):
                repaired = repair_data_box(header_box, data_box, orientation, used_box)
                if contains(used_box, repaired):
                    actions.append(
                        f"{path}.data_range shifted from {data_range} to {box_to_range(repaired)} "
                        "to exclude header cells"
                    )
                    data_box = repaired
                    header["data_range"] = box_to_range(data_box)
                    data_range = header["data_range"]
                else:
                    set_null(header, path, "data_range", null_fields)
                    errors.append(f"{path}.data_range overlaps its header_range: {data_range} vs {header_range}")
            if header_box is not None and data_box is not None:
                if orientation == "column" and (data_box[0] != header_box[0] or data_box[2] != header_box[2]):
                    repaired = (header_box[0], data_box[1], header_box[2], data_box[3])
                    actions.append(
                        f"{path}.data_range realigned from {data_range} to {box_to_range(repaired)} "
                        "to match header columns"
                    )
                    data_box = repaired
                    header["data_range"] = box_to_range(data_box)
                    data_range = header["data_range"]
                if orientation == "row" and (data_box[1] != header_box[1] or data_box[3] != header_box[3]):
                    repaired = (data_box[0], header_box[1], data_box[2], header_box[3])
                    actions.append(
                        f"{path}.data_range realigned from {data_range} to {box_to_range(repaired)} "
                        "to match header rows"
                    )
                    data_box = repaired
                    header["data_range"] = box_to_range(data_box)
                    data_range = header["data_range"]
            data_texts = cell_texts(worksheet, data_box)
            if not data_texts:
                if header_box is not None:
                    repaired = repair_data_box(header_box, data_box, orientation, used_box)
                    if contains(used_box, repaired) and cell_texts(worksheet, repaired):
                        actions.append(
                            f"{path}.data_range repaired from {data_range} to {box_to_range(repaired)} "
                            "using the verified header span"
                        )
                        data_box = repaired
                        header["data_range"] = box_to_range(data_box)
                        data_range = header["data_range"]
                    else:
                        set_null(header, path, "data_range", null_fields)
                        errors.append(f"{path}.data_range contains no visible data: {data_range}")
                else:
                    set_null(header, path, "data_range", null_fields)
                    errors.append(f"{path}.data_range contains no visible data: {data_range}")

    sub_headers = header.get("sub_headers") or []
    child_header_boxes = []
    child_data_boxes = []
    if isinstance(sub_headers, list):
        for index, child in enumerate(sub_headers):
            child_path = f"{path}.sub_headers[{index}]"
            if not isinstance(child, dict):
                errors.append(f"{child_path} must be a mapping")
                continue
            child_header_range = child.get("header_range") or child.get("range")
            child_data_range = child.get("data_range")
            if child_header_range is not None:
                try:
                    child_header_box = range_box(child_header_range)
                except (TypeError, ValueError):
                    continue
                child_header_boxes.append((child_path, child_header_range, child_header_box))
            if child_data_range is not None:
                try:
                    child_data_box = range_box(child_data_range)
                except (TypeError, ValueError):
                    continue
                child_data_boxes.append((child_path, child_data_range, child_data_box))

    if data_box is not None:
        for child_path, child_header_range, child_header_box in child_header_boxes:
            if intersects(data_box, child_header_box):
                repaired = repair_data_box(child_header_box, data_box, orientation, used_box)
                if contains(used_box, repaired):
                    actions.append(
                        f"{path}.data_range shifted from {data_range} to {box_to_range(repaired)} "
                        f"to exclude {child_path}.header_range"
                    )
                    data_box = repaired
                    header["data_range"] = box_to_range(data_box)
                    data_range = header["data_range"]
                else:
                    set_null(header, path, "data_range", null_fields)
                    errors.append(
                        f"{path}.data_range overlaps {child_path}.header_range: "
                        f"{data_range} vs {child_header_range}"
                    )
        for child_path, child_data_range, child_data_box in child_data_boxes:
            if not contains(data_box, child_data_box):
                child = header
                for token in re.findall(r"sub_headers\\[(\\d+)\\]", child_path):
                    child = child.get("sub_headers", [])[int(token)]
                set_null(child, child_path, "data_range", null_fields)
                errors.append(
                    f"{child_path}.data_range is not contained by parent {path}.data_range: "
                    f"{child_data_range} vs {data_range}"
                )

    if header_box is not None and child_header_boxes:
        for child_path, child_header_range, child_header_box in child_header_boxes:
            if orientation == "column":
                if child_header_box[0] < header_box[0] or child_header_box[2] > header_box[2]:
                    child = header.get("sub_headers", [])[int(re.search(r"\\[(\\d+)\\]$", child_path).group(1))]
                    set_null(child, child_path, "header_range", null_fields)
                    errors.append(
                        f"{child_path}.header_range columns are outside parent {path}.header_range: "
                        f"{child_header_range} vs {header_range}"
                    )
                if child_header_box[1] <= header_box[3]:
                    child = header.get("sub_headers", [])[int(re.search(r"\\[(\\d+)\\]$", child_path).group(1))]
                    set_null(child, child_path, "header_range", null_fields)
                    errors.append(
                        f"{child_path}.header_range must be below parent {path}.header_range: "
                        f"{child_header_range} vs {header_range}"
                    )
            else:
                if child_header_box[1] < header_box[1] or child_header_box[3] > header_box[3]:
                    child = header.get("sub_headers", [])[int(re.search(r"\\[(\\d+)\\]$", child_path).group(1))]
                    set_null(child, child_path, "header_range", null_fields)
                    errors.append(
                        f"{child_path}.header_range rows are outside parent {path}.header_range: "
                        f"{child_header_range} vs {header_range}"
                    )
                if child_header_box[0] <= header_box[2]:
                    child = header.get("sub_headers", [])[int(re.search(r"\\[(\\d+)\\]$", child_path).group(1))]
                    set_null(child, child_path, "header_range", null_fields)
                    errors.append(
                        f"{child_path}.header_range must be right of parent {path}.header_range: "
                        f"{child_header_range} vs {header_range}"
                    )


workbook_path, sheet_name, structure_path = sys.argv[1:4]
with open(structure_path, "r", encoding="utf-8") as handle:
    structure = yaml.safe_load(handle) or {}

workbook = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
errors = []
actions = []
null_fields = []
try:
    worksheet = workbook[sheet_name]
    used_box = range_box(worksheet.calculate_dimension())
    for path, value in walk(structure):
        try:
            min_col, min_row, max_col, max_row = range_box(value)
        except (TypeError, ValueError):
            errors.append(f"{path} is not a valid A1 range: {value}")
            continue
        if min_row < 1 or min_col < 1 or max_row > 1048576 or max_col > 16384:
            errors.append(f"{path} is outside Excel worksheet bounds: {value}")
    for path, header in walk_headers(structure):
        check_header(worksheet, path, header, used_box, errors, actions, null_fields)
finally:
    workbook.close()

feedback_parts = []
if actions:
    feedback_parts.append("VerificationAgent repaired provable workbook-backed fields: " + "; ".join(actions))
if errors:
    feedback_parts.append(
        "VerificationAgent could not repair these fields from workbook code alone. "
        "LayoutAgent must fill only these exact null fields with concrete A1 ranges/text: "
        + "; ".join(errors)
    )

print(json.dumps({
    "status": "not_good" if errors or null_fields else "good",
    "errors": errors,
    "actions": actions,
    "null_fields": list(dict.fromkeys(null_fields)),
    "feedback": " ".join(feedback_parts) if feedback_parts else "Structure verified by workbook-backed code.",
    "repaired_structure_yaml": yaml.safe_dump(structure, sort_keys=False, allow_unicode=True).strip(),
}, ensure_ascii=False))
'''
