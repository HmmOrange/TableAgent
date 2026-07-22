from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from datasets.base import EvalSample


@dataclass(frozen=True)
class WorkbookConversion:
    path: Path
    source_format: str
    sheet_names: list[str]


def sample_to_xlsx(sample: EvalSample, output_path: str | Path) -> WorkbookConversion:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    source_format = _write_sample(workbook, sample)
    workbook.save(output)

    return WorkbookConversion(
        path=output,
        source_format=source_format,
        sheet_names=workbook.sheetnames,
    )


def _write_sample(workbook: Workbook, sample: EvalSample) -> str:
    raw_tables = sample.raw.get("tables")
    if isinstance(raw_tables, list) and raw_tables:
        tables = [_table_payload_to_rows(table) for table in raw_tables]
        _write_combined_tables(workbook, tables)
        for index, table in enumerate(tables, start=1):
            sheet = workbook.create_sheet(_sheet_title(f"Table {index}"))
            _write_rows(sheet, table.rows, table.merged_regions)
        return "json:tables"

    content = sample.table_content.strip()
    if content.startswith("<"):
        table = _html_to_rows(content)
        sheet = workbook.create_sheet(_sheet_title(sample.table_id or "Table"))
        _write_rows(sheet, table.rows, table.merged_regions)
        return "html"

    if "\\begin{tabular}" in content:
        rows = latex_table_to_rows(sample.table_content)
        sheet = workbook.create_sheet(_sheet_title(sample.table_id or "Table"))
        _write_rows(sheet, rows, [])
        return "latex"

    try:
        payload = json.loads(sample.table_content)
    except json.JSONDecodeError:
        rows = _plain_text_to_rows(sample.table_content)
        sheet = workbook.create_sheet(_sheet_title(sample.table_id or "Table"))
        _write_rows(sheet, rows, [])
        return "text"

    source_format = _write_json_payload(workbook, payload, title=sample.table_id or "Table")
    return source_format


@dataclass(frozen=True)
class _TableRows:
    rows: list[list[Any]]
    merged_regions: list[dict[str, int]]
    title: str = ""


def _table_payload_to_rows(payload: Any) -> _TableRows:
    if isinstance(payload, str):
        if "<table" in payload.lower():
            return _html_to_rows(payload)
        return _TableRows(_plain_text_to_rows(payload), [])
    if isinstance(payload, list):
        return _TableRows(_rectangularize(payload), [])
    if isinstance(payload, dict):
        if "texts" in payload:
            return _hitab_json_to_rows(payload)
        for key in ("rows", "data", "table", "values"):
            value = payload.get(key)
            if isinstance(value, str) and "<table" in value.lower():
                return _html_to_rows(value)
            if isinstance(value, list):
                return _TableRows(_rectangularize(value), [])
        return _TableRows([[key, value] for key, value in payload.items()], [])
    return _TableRows([[payload]], [])


def _write_json_payload(workbook: Workbook, payload: Any, *, title: str) -> str:
    if isinstance(payload, dict) and "texts" in payload:
        table = _hitab_json_to_rows(payload)
        sheet = workbook.create_sheet(_sheet_title(table.title or title))
        _write_rows(sheet, table.rows, table.merged_regions)
        return "hitab-json"

    if isinstance(payload, dict) and isinstance(payload.get("tables"), list):
        tables = [_table_payload_to_rows(table) for table in payload["tables"]]
        _write_combined_tables(workbook, tables)
        for index, table in enumerate(tables, start=1):
            sheet = workbook.create_sheet(_sheet_title(f"Table {index}"))
            _write_rows(sheet, table.rows, table.merged_regions)
        return "json:tables"

    table = _table_payload_to_rows(payload)
    sheet = workbook.create_sheet(_sheet_title(table.title or title))
    _write_rows(sheet, table.rows, table.merged_regions)
    return "json"


def _hitab_json_to_rows(payload: dict[str, Any]) -> _TableRows:
    return _TableRows(
        rows=_rectangularize(payload.get("texts") or []),
        merged_regions=[
            {
                "first_row": int(region.get("first_row", 0)),
                "last_row": int(region.get("last_row", region.get("first_row", 0))),
                "first_column": int(region.get("first_column", 0)),
                "last_column": int(region.get("last_column", region.get("first_column", 0))),
            }
            for region in payload.get("merged_regions") or []
        ],
        title=str(payload.get("title") or ""),
    )


def _html_to_rows(content: str) -> _TableRows:
    soup = BeautifulSoup(content, "html.parser")
    table = soup.find("table")
    if table is None:
        return _TableRows(_plain_text_to_rows(soup.get_text("\n", strip=True)), [])

    caption = table.find("caption")
    title = caption.get_text(" ", strip=True) if caption else ""
    rows: list[list[Any]] = []
    occupied: set[tuple[int, int]] = set()
    merged_regions: list[dict[str, int]] = []

    for row_index, tr in enumerate(table.find_all("tr")):
        while len(rows) <= row_index:
            rows.append([])
        col_index = 0
        for cell in tr.find_all(["td", "th"], recursive=False):
            while (row_index, col_index) in occupied:
                _ensure_cell(rows, row_index, col_index)
                col_index += 1

            rowspan = max(1, _safe_int(cell.get("rowspan"), 1))
            colspan = max(1, _safe_int(cell.get("colspan"), 1))
            _ensure_cell(rows, row_index, col_index)
            rows[row_index][col_index] = cell.get_text(" ", strip=True)

            if rowspan > 1 or colspan > 1:
                merged_regions.append(
                    {
                        "first_row": row_index,
                        "last_row": row_index + rowspan - 1,
                        "first_column": col_index,
                        "last_column": col_index + colspan - 1,
                    }
                )

            for row_offset in range(rowspan):
                for col_offset in range(colspan):
                    target = (row_index + row_offset, col_index + col_offset)
                    _ensure_cell(rows, target[0], target[1])
                    if row_offset or col_offset:
                        occupied.add(target)

            col_index += colspan

    return _TableRows(_rectangularize(rows), merged_regions, title=title)


def _write_combined_tables(workbook: Workbook, tables: list[_TableRows]) -> None:
    sheet = workbook.create_sheet("Combined")
    cursor = 1
    for index, table in enumerate(tables, start=1):
        sheet.cell(row=cursor, column=1, value=table.title or f"Table {index}")
        cursor += 1
        _write_rows(sheet, table.rows, table.merged_regions, start_row=cursor)
        cursor += len(table.rows) + 2


def _write_rows(
    sheet: Worksheet,
    rows: list[list[Any]],
    merged_regions: list[dict[str, int]],
    *,
    start_row: int = 1,
) -> None:
    rows = _rectangularize(rows)
    for row_index, row in enumerate(rows, start=start_row):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=_cell_value(value))

    for region in merged_regions:
        first_row = start_row + int(region["first_row"])
        last_row = start_row + int(region["last_row"])
        first_col = 1 + int(region["first_column"])
        last_col = 1 + int(region["last_column"])
        if first_row != last_row or first_col != last_col:
            sheet.merge_cells(
                start_row=first_row,
                start_column=first_col,
                end_row=last_row,
                end_column=last_col,
            )


def _plain_text_to_rows(content: str) -> list[list[str]]:
    lines = [line for line in content.splitlines() if line.strip()]
    if any("|" in line for line in lines):
        return [
            [cell.strip() for cell in line.strip().strip("|").split("|")]
            for line in lines
            if not re.fullmatch(r"\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?\s*", line)
        ]
    return [[line] for line in lines]


def latex_table_to_rows(content: str) -> list[list[str]]:
    """Parse RealHiTBench's generated LaTeX tabular data into plain rows."""
    body = _outer_tabular_body(content)
    if body is None:
        return _plain_text_to_rows(content)

    rows: list[list[str]] = []
    for raw_cells in _split_latex_tabular(body):
        row: list[str] = []
        for raw_cell in raw_cells:
            span = _latex_multicolumn_span(raw_cell)
            row.append(_clean_latex_cell(raw_cell))
            row.extend([""] * (span - 1))
        if any(cell.strip() for cell in row):
            rows.append(row)
    return _rectangularize(rows)


def _outer_tabular_body(content: str) -> str | None:
    match = re.search(r"\\begin\{tabular\}(?:\[[^\]]*\])?\{", content)
    if match is None:
        return None

    column_spec_end = _matching_brace(content, match.end() - 1)
    if column_spec_end is None:
        return None
    body_start = column_spec_end + 1
    depth = 1
    position = body_start
    begin_token = r"\begin{tabular}"
    end_token = r"\end{tabular}"
    while position < len(content):
        next_begin = content.find(begin_token, position)
        next_end = content.find(end_token, position)
        if next_end == -1:
            return None
        if next_begin != -1 and next_begin < next_end:
            depth += 1
            position = next_begin + len(begin_token)
            continue
        depth -= 1
        if depth == 0:
            return content[body_start:next_end]
        position = next_end + len(end_token)
    return None


def _matching_brace(text: str, start: int) -> int | None:
    depth = 0
    for index in range(start, len(text)):
        if text[index] == "{" and (index == 0 or text[index - 1] != "\\"):
            depth += 1
        elif text[index] == "}" and (index == 0 or text[index - 1] != "\\"):
            depth -= 1
            if depth == 0:
                return index
    return None


def _split_latex_tabular(body: str) -> list[list[str]]:
    rows: list[list[str]] = []
    row: list[str] = []
    cell: list[str] = []
    nested_depth = 0
    position = 0
    begin_token = r"\begin{tabular}"
    end_token = r"\end{tabular}"

    while position < len(body):
        if body.startswith(begin_token, position):
            nested_depth += 1
            column_match = re.match(r"\\begin\{tabular\}(?:\[[^\]]*\])?\{", body[position:])
            if column_match is None:
                position += len(begin_token)
                continue
            column_start = position + column_match.end() - 1
            column_end = _matching_brace(body, column_start)
            position = (column_end + 1) if column_end is not None else position + column_match.end()
            continue
        if body.startswith(end_token, position):
            nested_depth = max(0, nested_depth - 1)
            position += len(end_token)
            continue
        if body.startswith(r"\\", position):
            if nested_depth == 0:
                row.append("".join(cell))
                rows.append(row)
                row = []
                cell = []
            else:
                cell.append(" ")
            position += 2
            continue
        if body[position] == "&" and nested_depth == 0 and (position == 0 or body[position - 1] != "\\"):
            row.append("".join(cell))
            cell = []
            position += 1
            continue
        cell.append(body[position])
        position += 1

    if cell or row:
        row.append("".join(cell))
        rows.append(row)
    return rows


def _latex_multicolumn_span(cell: str) -> int:
    match = re.search(r"\\multicolumn\{(\d+)\}", cell)
    return max(1, int(match.group(1))) if match else 1


def _clean_latex_cell(cell: str) -> str:
    text = re.sub(r"\\(?:toprule|midrule|bottomrule|hline|cline\{[^}]*\})", "", cell)
    text = re.sub(r"\\rowcolor(?:\[[^\]]*\])?\{[^}]*\}", "", text)
    text = re.sub(r"\\cellcolor(?:\[[^\]]*\])?\{[^}]*\}", "", text)
    text = re.sub(r"\\color(?:\[[^\]]*\])?\{[^}]*\}", "", text)

    text = _unwrap_latex_commands(text)

    replacements = {
        r"\&": "&",
        r"\%": "%",
        r"\$": "$",
        r"\#": "#",
        r"\_": "_",
        r"\{": "{",
        r"\}": "}",
        "{[}": "[",
        "{]}": "]",
        "~": " ",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = re.sub(r"\\[A-Za-z]+(?:\[[^\]]*\])?", "", text)
    text = text.replace("{", "").replace("}", "")
    return " ".join(text.split())


def _unwrap_latex_commands(text: str) -> str:
    arities = {
        "multicolumn": 3,
        "multirow": 3,
        "textbf": 1,
        "textit": 1,
        "emph": 1,
        "underline": 1,
        "textsuperscript": 1,
        "textsubscript": 1,
    }
    output: list[str] = []
    position = 0
    while position < len(text):
        if text[position] != "\\":
            output.append(text[position])
            position += 1
            continue

        command_match = re.match(r"\\([A-Za-z]+)", text[position:])
        if command_match is None:
            output.append(text[position])
            position += 1
            continue
        command = command_match.group(1)
        position += command_match.end()
        arity = arities.get(command)
        if arity is None:
            continue

        arguments: list[str] = []
        for _ in range(arity):
            while position < len(text) and text[position].isspace():
                position += 1
            if position >= len(text) or text[position] != "{":
                break
            end = _matching_brace(text, position)
            if end is None:
                break
            arguments.append(_unwrap_latex_commands(text[position + 1:end]))
            position = end + 1
        if arguments:
            output.append(arguments[-1])
    return "".join(output)


def _rectangularize(rows: Any) -> list[list[Any]]:
    normalized = []
    for row in rows or []:
        if isinstance(row, dict):
            normalized.append([f"{key}: {value}" for key, value in row.items()])
        elif isinstance(row, (list, tuple)):
            normalized.append(list(row))
        else:
            normalized.append([row])
    width = max((len(row) for row in normalized), default=0)
    return [row + [""] * (width - len(row)) for row in normalized]


def _ensure_cell(rows: list[list[Any]], row_index: int, col_index: int) -> None:
    while len(rows) <= row_index:
        rows.append([])
    while len(rows[row_index]) <= col_index:
        rows[row_index].append("")


def _safe_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _sheet_title(value: str) -> str:
    title = re.sub(r"[\[\]:*?/\\]", "_", str(value).strip())[:31]
    return title or "Table"


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)
