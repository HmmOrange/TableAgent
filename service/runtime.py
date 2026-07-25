from __future__ import annotations

import hashlib
import json
import shutil
import tempfile
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Literal

import openpyxl
import pandas as pd

from service.clients import create_model_client
from TableAgent.artifacts import (
    SummaryGenerator,
    build_workbook_metadata,
    build_workbook_schema,
    copy_artifact_tree,
    sheet_artifact_dir,
    write_sheet_retrieval_cards,
    write_workbook_retrieval_cards,
    workbook_artifact_dir,
)
from TableAgent.configs import load_config
from TableAgent.pipeline import TableAgentPipeline
from TableAgent.pipeline.base import PipelineOutput
from TableAgent.pipeline.common import safe_name
from TableAgent.pipeline.retrieval.embeddings import (
    OpenAICompatibleEmbeddingClient,
)
from TableAgent.artifacts.retrieval_cards import (
    DEFAULT_RETRIEVAL_CARD_EMBEDDING_MODEL,
)
from TableAgent.schema import EvalSample


Stage = Literal["structure", "qa", "all"]
SUPPORTED_WORKBOOK_EXTENSIONS = {".xls", ".xlsm", ".xlsx", ".xltm", ".xltx"}


class TableAgentService:
    """Reusable entry point for running TableAgent over workbook and query batches."""

    def __init__(
        self,
        config: dict[str, Any],
        *,
        llm_client: Any | None = None,
        layout_vlm_client: Any | None = None,
        llm_profile: str | None = None,
        vlm_profile: str | None = None,
        embedding_client: Any | None = None,
        root_dir: str | Path | None = None,
        pipeline_factory: Callable[..., TableAgentPipeline] = TableAgentPipeline,
    ):
        self.config = dict(config)
        service_config = self.config.get("service") or {}
        if not isinstance(service_config, dict):
            raise ValueError("service configuration must be a mapping")
        configured_root = root_dir or service_config.get("root_dir", "outputs")
        self.root_dir = Path(configured_root).expanduser().resolve()
        self.max_workers = max(1, int(service_config.get("max_workers", 1)))
        self.max_upload_bytes = max(1, int(service_config.get("max_upload_mb", 100))) * 1024 * 1024
        self.api_key = str(service_config["api_key"]) if service_config.get("api_key") else None
        self.allow_local_paths = bool(service_config.get("allow_local_paths", False))
        self.allowed_input_roots = tuple(
            Path(value).expanduser().resolve()
            for value in service_config.get("allowed_input_roots", [])
        )
        self._llm_client = llm_client
        self._layout_vlm_client = layout_vlm_client
        self._embedding_client = embedding_client
        self._embedding_model: str | None = None
        self.llm_profile = llm_profile or "table_agent"
        self.vlm_profile = vlm_profile or "table_agent"
        self.pipeline_factory = pipeline_factory

    @classmethod
    def from_config(
        cls,
        path: str | Path = "config.yaml",
        **kwargs: Any,
    ) -> "TableAgentService":
        return cls(load_config(path), **kwargs)

    def run(
        self,
        *,
        stage: Stage = "all",
        queries: Iterable[str] = (),
        workbooks: Iterable[str | Path],
        job_id: str | None = None,
        embed: bool = False,
        sheets: Iterable[str] = (),
        qa_max_replans: int | None = None,
        persist: bool = True,
    ) -> dict[str, Any]:
        stage = _validate_stage(stage)
        query_list = _validate_queries(queries, required=stage in {"qa", "all"})
        if qa_max_replans is not None and qa_max_replans < 0:
            raise ValueError("qa_max_replans must be greater than or equal to 0")
        workbook_list = [Path(value).expanduser().resolve() for value in workbooks]
        if not workbook_list:
            raise ValueError("At least one workbook is required")
        run_id = safe_name(job_id or new_job_id())
        if run_id in {".", ".."}:
            raise ValueError("Invalid job id")
        selected_sheets = _normalize_sheet_filters(sheets)

        run_dir = self._run_dir(run_id)
        if persist:
            self.root_dir.mkdir(parents=True, exist_ok=True)
            if run_dir.exists():
                raise ValueError(f"Run directory already exists: {run_id}")
            run_dir.mkdir(parents=False)

        try:
            with tempfile.TemporaryDirectory(prefix=f"table-agent-{run_id}-") as workspace_text:
                workspace_dir = Path(workspace_text)
                output_dir = run_dir if persist else workspace_dir / "output"
                output_dir.mkdir(parents=True, exist_ok=True)
                normalized = self._normalize_workbooks(workbook_list, workspace_dir / "normalized")
                self._validate_sheet_filters(normalized, selected_sheets)

                table_path = ";".join(str(item["path"]) for item in normalized)
                workbook_identities = self._workbook_identities(normalized)
                base_sample = self._sample(
                    sample_id=f"{run_id}-structure",
                    question=query_list[0] if query_list else "Generate workbook structure",
                    table_path=table_path,
                    workbook_names=[item["name"] for item in normalized],
                    selected_sheets=selected_sheets,
                    workbook_identities=workbook_identities,
                )
                structures: list[dict[str, Any]] = []
                answers: list[dict[str, Any]] = []
                source_dir = workspace_dir / "structure"

                pipeline = self.pipeline_factory(
                    llm_client=self._answer_client(),
                    layout_vlm_client=self._layout_client(),
                    config=self._pipeline_config(
                        "structure",
                        output_dir,
                        source_dir,
                        embed=False,
                    ),
                )
                records = pipeline.verify_samples([base_sample], force=True)
                structures = self._structure_results(
                    records,
                    normalized,
                    output_dir,
                    include_artifact_paths=persist,
                )
                structures = self._complete_structure_results(structures, normalized, selected_sheets)
                failed = [record for record in structures if record["status"] != "good"]
                if failed:
                    raise RuntimeError(f"Structure generation failed for {len(failed)} workbook sheet(s)")

                if stage in {"qa", "all"}:
                    samples = [
                        self._sample(
                            sample_id=f"{run_id}-query-{index}",
                            question=query,
                            table_path=table_path,
                            workbook_names=[item["name"] for item in normalized],
                            selected_sheets=selected_sheets,
                            workbook_identities=workbook_identities,
                        )
                        for index, query in enumerate(query_list, start=1)
                    ]
                    pipeline = self.pipeline_factory(
                        llm_client=self._answer_client(),
                        layout_vlm_client=None,
                        config=self._pipeline_config(
                            "qa",
                            output_dir,
                            source_dir,
                            embed=False,
                            qa_max_replans=qa_max_replans,
                        ),
                    )
                    pipeline.prepare_samples(samples)
                    for sample in samples:
                        output = pipeline.run(sample)
                        answers.append(self._answer_result(sample.question, output, normalized))

                schema_artifacts, metadata_artifacts, retrieval_records = self._build_workbook_artifacts(
                    normalized,
                    output_dir,
                    embed=embed,
                    selected_sheets=selected_sheets,
                    include_artifact_paths=persist,
                )

                result = {
                    "job_id": run_id,
                    "stage": stage,
                    "workbooks": [item["name"] for item in normalized],
                    "structures": structures,
                    "schema_artifacts": schema_artifacts,
                    "metadata_artifacts": metadata_artifacts,
                    "retrieval_records": retrieval_records,
                    "answers": answers,
                    "artifacts": self._artifact_paths(output_dir) if persist else [],
                }
                if persist:
                    (output_dir / "run.json").write_text(
                        json.dumps(result, ensure_ascii=False, indent=2, default=str) + "\n",
                        encoding="utf-8",
                    )
                    result["artifacts"] = self._artifact_paths(output_dir)
                return result
        except Exception:
            if persist and run_dir.exists():
                shutil.rmtree(run_dir)
            raise

    def run_indexed_qa(
        self,
        *,
        query: str,
        workbooks: Iterable[str | Path],
        artifacts: Iterable[dict[str, Any]],
        qa_max_replans: int | None = None,
        qa_enable_final_review: bool | None = None,
        mode: str = "thinking",
    ) -> dict[str, Any]:
        """Answer from ingestion-time verified structures without running layout extraction."""
        normalized_query = str(query).strip()
        if not normalized_query:
            raise ValueError("At least one non-empty query is required for indexed QA")
        if qa_max_replans is not None and qa_max_replans < 0:
            raise ValueError("qa_max_replans must be greater than or equal to 0")
        if mode not in {"instant", "thinking"}:
            raise ValueError("Indexed QA mode must be instant or thinking")
        workbook_list = [Path(value).expanduser().resolve() for value in workbooks]
        if not workbook_list:
            raise ValueError("At least one workbook is required")

        artifact_list = [dict(item) for item in artifacts if isinstance(item, dict)]
        if not artifact_list:
            raise ValueError("Indexed QA requires at least one retrieval artifact")

        if self.pipeline_factory is TableAgentPipeline:
            return self._run_indexed_qa_hybrid(
                normalized_query=normalized_query,
                workbook_list=workbook_list,
                artifact_list=artifact_list,
                qa_max_replans=qa_max_replans,
                qa_enable_final_review=qa_enable_final_review,
                mode=mode,
            )

        run_id = new_job_id()
        with tempfile.TemporaryDirectory(prefix=f"table-agent-indexed-{safe_name(run_id)}-") as workspace_text:
            workspace_dir = Path(workspace_text)
            output_dir = workspace_dir / "output"
            source_dir = workspace_dir / "indexed"
            output_dir.mkdir(parents=True, exist_ok=True)
            normalized = self._normalize_workbooks(workbook_list, workspace_dir / "normalized")
            by_name = {str(item["name"]): item for item in normalized}

            grouped: dict[str, list[dict[str, Any]]] = {}
            for artifact in sorted(
                artifact_list,
                key=lambda item: float(item.get("score") or 0.0),
                reverse=True,
            ):
                workbook_name = str(
                    artifact.get("upload_name")
                    or artifact.get("document_name")
                    or artifact.get("workbook")
                    or ""
                )
                if workbook_name in by_name:
                    grouped.setdefault(workbook_name, []).append(artifact)

            if not grouped:
                raise ValueError("Indexed artifacts do not match any uploaded workbook")

            pipeline = self.pipeline_factory(
                llm_client=self._answer_client_for_mode(mode),
                layout_vlm_client=None,
                config=self._pipeline_config(
                    "qa",
                    output_dir,
                    source_dir,
                    embed=False,
                    qa_max_replans=qa_max_replans,
                    mode=mode,
                ),
            )
            group_answers: list[dict[str, Any]] = []
            total_prompt_tokens = 0
            total_completion_tokens = 0

            for group_index, (workbook_name, candidates) in enumerate(grouped.items(), start=1):
                structures: list[tuple[str, Path]] = []
                seen_sheets: set[str] = set()
                group_dir = source_dir / f"group-{group_index}"
                for candidate in candidates:
                    sheet_name = str(
                        candidate.get("sheet") or candidate.get("sheet_name") or ""
                    ).strip()
                    structure_text = str(candidate.get("structure_yaml") or "").strip()
                    if not structure_text or not sheet_name or sheet_name in seen_sheets:
                        continue
                    seen_sheets.add(sheet_name)
                    structure_path = group_dir / f"{len(structures) + 1:03d}-{safe_name(sheet_name)}.yaml"
                    structure_path.parent.mkdir(parents=True, exist_ok=True)
                    structure_path.write_text(structure_text + "\n", encoding="utf-8")
                    structures.append((sheet_name, structure_path))

                if not structures:
                    continue

                cards = "\n\n".join(
                    str(candidate.get("retrieval_card") or "").strip()
                    for candidate in candidates
                    if str(candidate.get("retrieval_card") or "").strip()
                )
                fallback_prompt = (
                    "Answer the spreadsheet question using only the indexed, verified TableAgent context. "
                    "Do not invent values that are absent from the context.\n\n"
                    f"Question: {normalized_query}\n\n"
                    f"Retrieved cards:\n{cards}"
                )
                answer_response, qa_info = pipeline._run_verified_qa(
                    question=normalized_query,
                    structure_path=structures[0][1],
                    workbook_path=Path(by_name[workbook_name]["path"]),
                    qa_artifact_dir=output_dir / "qa" / f"group-{group_index}",
                    fallback_prompt=fallback_prompt,
                    fallback_text_prompt=fallback_prompt,
                    related_structure_paths=[path for _, path in structures[1:]],
                    enable_final_answer_review=(
                        True
                        if qa_enable_final_review is None
                        else qa_enable_final_review
                    ),
                )
                total_prompt_tokens += int(getattr(answer_response, "prompt_tokens", 0) or 0)
                total_completion_tokens += int(getattr(answer_response, "completion_tokens", 0) or 0)
                public_qa_info = dict(qa_info)
                public_qa_info.pop("artifacts", None)
                group_answers.append(
                    {
                        "workbook": workbook_name,
                        "sheets": [sheet for sheet, _ in structures],
                        "answer": answer_response.content,
                        "qa": public_qa_info,
                        "candidates": [
                            {
                                "id": candidate.get("id"),
                                "score": float(candidate.get("score") or 0.0),
                                "sheet": candidate.get("sheet") or candidate.get("sheet_name"),
                                "table_id": candidate.get("table_id") or "",
                                "retrieval_type": candidate.get("retrieval_type") or "data",
                                "retrieval_level": candidate.get("retrieval_level") or "table",
                            }
                            for candidate in candidates
                        ],
                    }
                )

            if not group_answers:
                raise RuntimeError("Indexed artifacts did not contain any usable verified structures")

            if len(group_answers) == 1:
                final_answer = str(group_answers[0]["answer"])
            else:
                evidence = "\n\n".join(
                    f"Workbook: {item['workbook']}\nSheets: {', '.join(item['sheets'])}\nAnswer: {item['answer']}"
                    for item in group_answers
                )
                synthesis_prompt = (
                    "Combine the independently verified workbook answers into one direct answer to the user. "
                    "Preserve disagreements and workbook attribution, and do not add facts not present in the evidence.\n\n"
                    f"Question: {normalized_query}\n\nEvidence:\n{evidence}"
                )
                synthesis_response = pipeline.qa_agent.run(prompt=synthesis_prompt)
                final_answer = synthesis_response.content
                total_prompt_tokens += int(getattr(synthesis_response, "prompt_tokens", 0) or 0)
                total_completion_tokens += int(getattr(synthesis_response, "completion_tokens", 0) or 0)

            selected_workbooks = [str(item["workbook"]) for item in group_answers]
            selected_sheets = list(
                dict.fromkeys(
                    sheet
                    for item in group_answers
                    for sheet in item["sheets"]
                )
            )
            return {
                "job_id": run_id,
                "stage": "qa",
                "workbooks": selected_workbooks,
                "structures": [],
                "schema_artifacts": [],
                "metadata_artifacts": [],
                "retrieval_records": [],
                "answers": [
                    {
                        "query": normalized_query,
                        "answer": final_answer,
                        "workbook": selected_workbooks[0] if len(selected_workbooks) == 1 else "",
                        "workbooks": selected_workbooks,
                        "sheets": selected_sheets,
                        "retrieval": {
                            "mode": "indexed_vector_multi_candidate",
                            "candidate_count": sum(len(item["candidates"]) for item in group_answers),
                            "workbook_count": len(group_answers),
                            "groups": [
                                {"workbook": item["workbook"], "candidates": item["candidates"]}
                                for item in group_answers
                            ],
                        },
                        "qa": {
                            "success": True,
                            "fallback_used": any(item["qa"].get("fallback_used") for item in group_answers),
                            "replan_count": sum(int(item["qa"].get("replan_count", 0) or 0) for item in group_answers),
                            "per_workbook": [
                                {"workbook": item["workbook"], **item["qa"]}
                                for item in group_answers
                            ],
                        },
                        "token_usage": {
                            "prompt": total_prompt_tokens,
                            "completion": total_completion_tokens,
                        },
                    }
                ],
                "artifacts": [],
            }

    def select_indexed_artifact(
        self,
        *,
        query: str,
        artifacts: Iterable[dict[str, Any]],
        mode: str = "thinking",
    ) -> dict[str, Any]:
        """Select an indexed artifact before the caller downloads any workbook."""
        normalized_query = str(query).strip()
        if not normalized_query:
            raise ValueError("A non-empty query is required for indexed retrieval")
        artifact_list = [dict(item) for item in artifacts if isinstance(item, dict)]
        if not artifact_list:
            raise ValueError("Indexed retrieval requires at least one artifact")
        if mode not in {"instant", "thinking"}:
            raise ValueError("Indexed retrieval mode must be instant or thinking")

        workbook_names = list(
            dict.fromkeys(
                str(
                    artifact.get("upload_name")
                    or artifact.get("document_name")
                    or artifact.get("workbook")
                    or ""
                ).strip()
                for artifact in artifact_list
            )
        )
        workbook_paths = {
            name: Path(name)
            for name in workbook_names
            if name
        }
        eligible_artifacts = [
            artifact
            for artifact in artifact_list
            if str(
                artifact.get("upload_name")
                or artifact.get("document_name")
                or artifact.get("workbook")
                or ""
            ).strip()
            in workbook_paths
            and str(
                artifact.get("sheet") or artifact.get("sheet_name") or ""
            ).strip()
            and (
                str(artifact.get("retrieval_card") or "").strip()
                or isinstance(artifact.get("metadata"), dict)
            )
        ]
        if not eligible_artifacts:
            raise ValueError("Indexed artifacts did not contain usable retrieval context")

        with tempfile.TemporaryDirectory(prefix="table-agent-select-") as workspace_text:
            workspace_dir = Path(workspace_text)
            pipeline = self.pipeline_factory(
                llm_client=self._answer_client_for_mode(mode),
                layout_vlm_client=None,
                config=self._pipeline_config(
                    "qa",
                    workspace_dir / "output",
                    workspace_dir / "indexed",
                    embed=False,
                    retrieval_rerank_with_llm=(mode == "thinking"),
                ),
            )
            responses = []
            candidate = pipeline.source_retriever.select_indexed(
                question=normalized_query,
                artifacts=eligible_artifacts,
                workbook_paths=workbook_paths,
                responses=responses,
                fit_context=pipeline._fit_context,
            )
        if candidate is None:
            raise RuntimeError("TableAgent hybrid retrieval found no usable candidate")

        selected_artifact = next(
            (
                artifact
                for artifact in eligible_artifacts
                if str(artifact.get("id") or "") == candidate.artifact_id
            ),
            None,
        )
        if selected_artifact is None:
            raise RuntimeError("Selected TableAgent artifact is missing from the candidate set")
        retrieval = self._indexed_retrieval_payload(
            candidate,
            eligible_artifacts,
        )
        rejection_reason = self._indexed_rejection_reason(
            candidate,
            retrieval,
            mode=mode,
        )
        if rejection_reason:
            return {
                "status": "no_evidence",
                "selected_artifact_id": None,
                "document_id": None,
                "workbook": None,
                "sheet": None,
                "retrieval": {
                    **retrieval,
                    "document_id": None,
                    "workbook": None,
                    "sheet": None,
                    "table_id": None,
                    "table_name": None,
                    "selected_artifact_id": None,
                    "rejection_reason": rejection_reason,
                    "audit": [
                        {**row, "selected": False}
                        for row in retrieval.get("audit", [])
                    ],
                },
            }
        return {
            "status": "selected",
            "selected_artifact_id": candidate.artifact_id,
            "document_id": str(selected_artifact.get("document_id") or ""),
            "workbook": candidate.workbook_path.name,
            "sheet": candidate.sheet_name,
            "retrieval": retrieval,
        }

    @staticmethod
    def _indexed_rejection_reason(
        candidate: Any,
        retrieval: dict[str, Any],
        *,
        mode: str,
    ) -> str | None:
        reranker = retrieval.get("reranker")
        if isinstance(reranker, dict) and reranker.get("status") == "need_more":
            return "The TableAgent reranker found no sufficiently relevant indexed table."
        if mode != "instant":
            return None

        matched_count = len(candidate.matched_terms)
        term_count = matched_count + len(candidate.missing_terms)
        coverage = matched_count / term_count if term_count else 0.0
        if (
            candidate.lexical_score >= 2
            or coverage >= 0.5
            or (candidate.embedding_used and candidate.embedding_score >= 0.5)
        ):
            return None
        return "No indexed table passed the instant-mode relevance threshold."

    def _run_indexed_qa_hybrid(
        self,
        *,
        normalized_query: str,
        workbook_list: list[Path],
        artifact_list: list[dict[str, Any]],
        qa_max_replans: int | None,
        qa_enable_final_review: bool | None,
        mode: str,
    ) -> dict[str, Any]:
        """Run one QA pass after TableAgent's lexical/entity/embedding selection."""
        run_id = new_job_id()
        with tempfile.TemporaryDirectory(
            prefix=f"table-agent-indexed-{safe_name(run_id)}-"
        ) as workspace_text:
            workspace_dir = Path(workspace_text)
            output_dir = workspace_dir / "output"
            source_dir = workspace_dir / "indexed"
            output_dir.mkdir(parents=True, exist_ok=True)
            normalized = self._normalize_workbooks(
                workbook_list,
                workspace_dir / "normalized",
            )
            workbook_paths = {
                str(item["name"]): Path(item["path"])
                for item in normalized
            }
            eligible_artifacts = [
                artifact
                for artifact in artifact_list
                if str(
                    artifact.get("upload_name")
                    or artifact.get("document_name")
                    or artifact.get("workbook")
                    or ""
                ).strip()
                in workbook_paths
                and str(
                    artifact.get("sheet") or artifact.get("sheet_name") or ""
                ).strip()
                and (
                    str(artifact.get("retrieval_card") or "").strip()
                    or isinstance(artifact.get("metadata"), dict)
                )
            ]
            pipeline = self.pipeline_factory(
                llm_client=self._answer_client_for_mode(mode),
                layout_vlm_client=None,
                config=self._pipeline_config(
                    "qa",
                    output_dir,
                    source_dir,
                    embed=False,
                    qa_max_replans=qa_max_replans,
                    mode=mode,
                ),
            )
            responses = []
            candidate = pipeline.source_retriever.select_indexed(
                question=normalized_query,
                artifacts=eligible_artifacts,
                workbook_paths=workbook_paths,
                responses=responses,
                fit_context=pipeline._fit_context,
            )
            if candidate is None:
                raise RuntimeError(
                    "Indexed artifacts did not contain any usable verified structures"
                )

            selected_workbook = candidate.workbook_path.name
            selected_artifact = next(
                (
                    artifact
                    for artifact in eligible_artifacts
                    if str(artifact.get("id") or "") == candidate.artifact_id
                    and str(
                        artifact.get("upload_name")
                        or artifact.get("document_name")
                        or artifact.get("workbook")
                        or ""
                    ).strip()
                    == selected_workbook
                    and str(
                        artifact.get("sheet") or artifact.get("sheet_name") or ""
                    ).strip()
                    == candidate.sheet_name
                ),
                None,
            )
            if selected_artifact is None:
                raise RuntimeError(
                    "The selected indexed sheet artifact is missing from the QA request"
                )

            selected_workbook_artifacts = [
                artifact
                for artifact in eligible_artifacts
                if str(
                    artifact.get("upload_name")
                    or artifact.get("document_name")
                    or artifact.get("workbook")
                    or ""
                ).strip()
                == selected_workbook
            ]
            primary_structure = str(
                selected_artifact.get("structure_yaml") or ""
            ).strip()
            if not primary_structure:
                primary_structure = next(
                    (
                        str(artifact.get("structure_yaml") or "").strip()
                        for artifact in selected_workbook_artifacts
                        if str(
                            artifact.get("sheet")
                            or artifact.get("sheet_name")
                            or ""
                        ).strip()
                        == candidate.sheet_name
                        and str(artifact.get("structure_yaml") or "").strip()
                    ),
                    "",
                )

            structure_path = source_dir / f"001-{safe_name(candidate.sheet_name)}.yaml"
            if primary_structure:
                structure_path.parent.mkdir(parents=True, exist_ok=True)
                structure_path.write_text(primary_structure + "\n", encoding="utf-8")

            related_structure_paths: list[Path] = []
            related_sheet_names: list[str] = []
            seen_related_sheets = {candidate.sheet_name}
            for artifact in selected_workbook_artifacts:
                sheet_name = str(
                    artifact.get("sheet") or artifact.get("sheet_name") or ""
                ).strip()
                structure_text = str(artifact.get("structure_yaml") or "").strip()
                if (
                    not sheet_name
                    or not structure_text
                    or sheet_name in seen_related_sheets
                ):
                    continue
                seen_related_sheets.add(sheet_name)
                related_path = source_dir / (
                    f"{len(related_structure_paths) + 2:03d}-{safe_name(sheet_name)}.yaml"
                )
                related_path.parent.mkdir(parents=True, exist_ok=True)
                related_path.write_text(structure_text + "\n", encoding="utf-8")
                related_structure_paths.append(related_path)
                related_sheet_names.append(sheet_name)

            selected_card = str(selected_artifact.get("retrieval_card") or "").strip()
            fallback_prompt = (
                "Answer the spreadsheet question using only the indexed, verified "
                "TableAgent context. Do not invent values absent from the context.\n\n"
                f"Question: {normalized_query}\n\n"
                f"Selected retrieval card:\n{selected_card}"
            )
            answer_response, qa_info = pipeline._run_verified_qa(
                question=normalized_query,
                structure_path=structure_path,
                workbook_path=candidate.workbook_path,
                qa_artifact_dir=output_dir / "qa",
                fallback_prompt=fallback_prompt,
                fallback_text_prompt=fallback_prompt,
                related_structure_paths=related_structure_paths,
                enable_final_answer_review=(
                    True
                    if qa_enable_final_review is None
                    else qa_enable_final_review
                ),
            )
            public_qa_info = dict(qa_info)
            public_qa_info.pop("artifacts", None)
            retrieval_payload = self._indexed_retrieval_payload(
                candidate,
                eligible_artifacts,
            )
            prompt_tokens = sum(
                int(getattr(response, "prompt_tokens", 0) or 0)
                for response in responses
            ) + int(getattr(answer_response, "prompt_tokens", 0) or 0)
            completion_tokens = sum(
                int(getattr(response, "completion_tokens", 0) or 0)
                for response in responses
            ) + int(getattr(answer_response, "completion_tokens", 0) or 0)
            return {
                "job_id": run_id,
                "stage": "qa",
                "workbooks": [selected_workbook],
                "structures": [],
                "schema_artifacts": [],
                "metadata_artifacts": [],
                "retrieval_records": [],
                "answers": [
                    {
                        "query": normalized_query,
                        "answer": answer_response.content,
                        "workbook": selected_workbook,
                        "workbooks": [selected_workbook],
                        "sheets": [candidate.sheet_name, *related_sheet_names],
                        "retrieval": retrieval_payload,
                        "qa": public_qa_info,
                        "token_usage": {
                            "prompt": prompt_tokens,
                            "completion": completion_tokens,
                        },
                    }
                ],
                "artifacts": [],
            }

    @staticmethod
    def _indexed_retrieval_payload(
        candidate: Any,
        eligible_artifacts: list[dict[str, Any]],
    ) -> dict[str, Any]:
        retrieval_trace = candidate.retrieval_trace[-1] if candidate.retrieval_trace else {}
        selected_artifact = next(
            (
                item
                for item in eligible_artifacts
                if str(item.get("id") or "") == candidate.artifact_id
            ),
            {},
        )
        return {
            "mode": "table_agent_hybrid",
            "query_type": retrieval_trace.get("query_type", "data"),
            "candidate_count": len(eligible_artifacts),
            "workbook_count": len(
                {
                    str(
                        item.get("upload_name")
                        or item.get("document_name")
                        or item.get("workbook")
                        or ""
                    ).strip()
                    for item in eligible_artifacts
                }
            ),
            "document_id": str(selected_artifact.get("document_id") or ""),
            "workbook": candidate.workbook_path.name,
            "sheet": candidate.sheet_name,
            "table_id": candidate.table_id,
            "table_name": candidate.table_name,
            "selected_artifact_id": candidate.artifact_id,
            "embedding_used": candidate.embedding_used,
            "audit": [
                {
                    **row,
                    "selected": row.get("artifact_id") == candidate.artifact_id,
                }
                for row in candidate.retrieval_audit
            ],
            "reranker": retrieval_trace,
        }

    def delete_runs(
        self,
        run_ids: Iterable[str] = (),
        *,
        all_runs: bool = False,
    ) -> dict[str, list[str]]:
        selected = [str(run_id).strip() for run_id in run_ids if str(run_id).strip()]
        if all_runs and selected:
            raise ValueError("Choose specific run IDs or --delete-all-jobs, not both")

        if all_runs:
            targets = (
                sorted(path for path in self.root_dir.iterdir() if self._is_run_dir(path))
                if self.root_dir.is_dir()
                else []
            )
        else:
            targets = [self._run_dir(run_id) for run_id in selected]

        deleted: list[str] = []
        missing: list[str] = []
        for target in targets:
            if not target.exists():
                missing.append(target.name)
                continue
            if not self._is_run_dir(target):
                raise ValueError(f"Not a TableAgent run directory: {target.name}")
            shutil.rmtree(target)
            deleted.append(target.name)
        return {"deleted": deleted, "missing": missing}

    def _run_dir(self, run_id: str) -> Path:
        if run_id != safe_name(run_id) or run_id in {".", ".."}:
            raise ValueError("Invalid job id")
        path = (self.root_dir / run_id).resolve()
        if path.parent != self.root_dir:
            raise ValueError("Invalid job id")
        return path

    @staticmethod
    def _is_run_dir(path: Path) -> bool:
        if path.is_symlink() or not path.is_dir():
            return False
        try:
            payload = json.loads((path / "run.json").read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return False
        return isinstance(payload, dict) and str(payload.get("job_id", "")) == path.name

    def validate_local_workbook(self, value: str | Path) -> Path:
        if not self.allow_local_paths:
            raise PermissionError("Server-side workbook paths are disabled; upload the workbook instead")
        path = Path(value).expanduser().resolve()
        if self.allowed_input_roots and not any(path.is_relative_to(root) for root in self.allowed_input_roots):
            raise PermissionError(f"Workbook path is outside the configured allowed_input_roots: {path}")
        self._validate_workbook(path)
        return path

    def _answer_client(self) -> Any:
        if self._llm_client is None:
            self._llm_client = create_model_client(
                self.config,
                kind="llm",
                profile=self.llm_profile,
            )
        return self._llm_client

    def _answer_client_for_mode(self, mode: str) -> Any:
        client = self._answer_client()
        mode_client = copy(client)
        if not hasattr(client, "extra_body"):
            return mode_client
        extra_body = dict(getattr(client, "extra_body", {}) or {})
        template_kwargs = dict(extra_body.get("chat_template_kwargs") or {})
        template_kwargs["enable_thinking"] = False
        extra_body["chat_template_kwargs"] = template_kwargs
        mode_client.extra_body = extra_body
        return mode_client

    def _layout_client(self) -> Any:
        if self._layout_vlm_client is None:
            self._layout_vlm_client = create_model_client(
                self.config,
                kind="vlm",
                profile=self.vlm_profile,
            )
        return self._layout_vlm_client

    def _pipeline_config(
        self,
        phase: Stage,
        output_dir: Path,
        source_dir: Path,
        *,
        embed: bool = False,
        qa_max_replans: int | None = None,
        retrieval_rerank_with_llm: bool | None = None,
        mode: str | None = None,
    ) -> dict[str, Any]:
        agent_config = dict(self.config.get("table_agent") or {})
        agent_config.update(
            {
                "phase": phase,
                "artifact_dir": str(output_dir / "artifacts"),
                "source_artifact_dir": str(source_dir),
                "structure_cache_dir": str(source_dir / "cache"),
                "cache_namespace": "service",
                "embed_retrieval_cards": bool(embed),
            }
        )
        for key in ("models", "vlm_models", "llm_providers", "embedding"):
            if key in self.config:
                agent_config[key] = self.config[key]
        if qa_max_replans is not None:
            agent_config["qa_max_replans"] = qa_max_replans
        if mode == "instant":
            agent_config["qa_max_retries"] = int(
                agent_config.get("qa_instant_max_retries", 1)
            )
            agent_config["generation_max_tokens"] = int(
                agent_config.get("qa_instant_generation_max_tokens", 2048)
            )
        if retrieval_rerank_with_llm is not None:
            agent_config["retrieval_rerank_with_llm"] = retrieval_rerank_with_llm
        return agent_config

    def _retrieval_embedding_backend(self) -> tuple[Any, str]:
        if self._embedding_client is not None:
            model = str(
                self._embedding_model
                or getattr(self._embedding_client, "model", "")
            ).strip()
            if not model:
                raise ValueError(
                    "The configured embedding client must expose a non-empty model name"
                )
            self._embedding_model = model
            return self._embedding_client, model

        table_agent_config = self.config.get("table_agent") or {}
        provider = (
            table_agent_config.get("retrieval_embedding_provider")
            if isinstance(table_agent_config, dict)
            else None
        )
        if not provider or str(provider).strip().lower() == "mock":
            raise ValueError(
                "--embed requires table_agent.retrieval_embedding_provider to name "
                "a configured real embedding model; mock embeddings are not allowed"
            )
        self._embedding_client = OpenAICompatibleEmbeddingClient.from_config(
            self.config,
            str(provider),
        )
        self._embedding_model = str(self._embedding_client.model).strip()
        if not self._embedding_model:
            raise ValueError("The configured retrieval embedding model name is empty")
        return self._embedding_client, self._embedding_model

    def _build_workbook_artifacts(
        self,
        normalized: list[dict[str, Any]],
        job_dir: Path,
        *,
        embed: bool,
        selected_sheets: tuple[str, ...],
        include_artifact_paths: bool,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
        schema_artifacts: list[dict[str, Any]] = []
        metadata_artifacts: list[dict[str, Any]] = []
        retrieval_records: list[dict[str, Any]] = []
        embedding_client = None
        embedding_model = DEFAULT_RETRIEVAL_CARD_EMBEDDING_MODEL
        if embed:
            embedding_client, embedding_model = self._retrieval_embedding_backend()

        for item in normalized:
            job_workbook_dir = workbook_artifact_dir(
                job_dir / "workbooks",
                item["name"],
                sources=False,
            )
            sheet_names = self._selected_sheet_names(item["path"], selected_sheets)
            structure_paths = []
            for sheet_name in sheet_names:
                structure_path = sheet_artifact_dir(job_workbook_dir, sheet_name) / "structure.yaml"
                if not structure_path.is_file():
                    continue
                structure_paths.append((sheet_name, structure_path))

            job_retrieval_records: list[dict[str, Any]] = []
            for sheet_name, structure_path in structure_paths:
                job_sheet_dir = sheet_artifact_dir(job_workbook_dir, sheet_name)
                job_retrieval_records.extend(
                    write_sheet_retrieval_cards(
                        job_sheet_dir,
                        Path(item["name"]),
                        sheet_name,
                        include_embeddings=embed,
                        embedding_client=embedding_client,
                        embedding_model=embedding_model,
                    )
                )
            schema_path = job_workbook_dir / "schema.yaml"
            missing = [name for name in sheet_names if not any(name == value[0] for value in structure_paths)]
            if missing:
                raise RuntimeError(
                    f"Missing valid structures for workbook '{item['name']}': {', '.join(missing)}"
                )
            build_workbook_schema(
                structure_paths,
                schema_path,
                SummaryGenerator(self._answer_client()),
            )
            schema_record: dict[str, Any] = {"workbook": item["name"]}
            if include_artifact_paths:
                schema_record["artifact"] = schema_path.relative_to(job_dir).as_posix()
            else:
                schema_record["schema"] = schema_path.read_text(encoding="utf-8")
            schema_artifacts.append(schema_record)

            metadata_path = job_workbook_dir / "metadata.json"
            build_workbook_metadata(
                item["source_path"],
                item["name"],
                metadata_path,
                schema_path=schema_path,
                summarizer=SummaryGenerator(self._answer_client()),
            )
            schema_text = schema_path.read_text(encoding="utf-8")
            workbook_metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            structure_by_sheet = {
                sheet_name: structure_path.read_text(encoding="utf-8")
                for sheet_name, structure_path in structure_paths
            }
            indexed_sheet_records = []
            for record in job_retrieval_records:
                sheet_name = str(record.get("sheet") or "").strip()
                indexed_sheet_records.append(
                    {
                        **record,
                        "artifact_version": 1,
                        "document_name": item["name"],
                        "workbook_sha256": item["sha256"],
                        "schema_yaml": schema_text,
                        "workbook_metadata": workbook_metadata,
                        **(
                            {"structure_yaml": structure_by_sheet[sheet_name]}
                            if sheet_name in structure_by_sheet
                            else {}
                        ),
                    }
                )
            workbook_retrieval_records: list[dict[str, Any]] = []
            if indexed_sheet_records:
                workbook_retrieval_records = write_workbook_retrieval_cards(
                    job_workbook_dir,
                    item["name"],
                    indexed_sheet_records,
                    include_embeddings=embed,
                    embedding_client=embedding_client,
                    embedding_model=embedding_model,
                )
            for record in workbook_retrieval_records:
                sheet_name = str(record.get("sheet") or "").strip()
                retrieval_records.append(
                    {
                        **record,
                        "artifact_version": 1,
                        "document_name": item["name"],
                        "workbook_sha256": item["sha256"],
                        "schema_yaml": schema_text,
                        "workbook_metadata": workbook_metadata,
                        **(
                            {"structure_yaml": structure_by_sheet[sheet_name]}
                            if sheet_name in structure_by_sheet
                            else {}
                        ),
                    }
                )
            metadata_record: dict[str, Any] = {"workbook": item["name"]}
            if include_artifact_paths:
                metadata_record["artifact"] = metadata_path.relative_to(job_dir).as_posix()
            else:
                metadata_record["metadata"] = workbook_metadata
            metadata_artifacts.append(metadata_record)

        return schema_artifacts, metadata_artifacts, retrieval_records

    @staticmethod
    def _sample(
        *,
        sample_id: str,
        question: str,
        table_path: str,
        workbook_names: list[str],
        selected_sheets: tuple[str, ...],
        workbook_identities: dict[str, dict[str, str]],
    ) -> EvalSample:
        return EvalSample(
            index=0,
            sample_id=sample_id,
            table_id="workbook_set",
            table_content="",
            question=question,
            answer=[],
            sample_path="service/qa/request.json",
            table_path=table_path,
            raw={
                "source": "table-agent-service",
                "workbooks": workbook_names,
                "selected_sheets": list(selected_sheets),
                "workbook_identities": workbook_identities,
            },
        )

    def _normalize_workbook(self, source: Path, workspace_dir: Path) -> dict[str, Any]:
        self._validate_workbook(source)
        digest = _sha256(source)
        if source.suffix.lower() == ".xlsx":
            destination = source
        else:
            workspace_dir.mkdir(parents=True, exist_ok=True)
            destination = workspace_dir / f"{digest[:24]}.xlsx"
            if source.suffix.lower() == ".xls":
                sheets = pd.read_excel(source, sheet_name=None)
                with pd.ExcelWriter(destination, engine="openpyxl") as writer:
                    for sheet_name, frame in sheets.items():
                        frame.to_excel(writer, sheet_name=str(sheet_name)[:31], index=False)
            else:
                workbook = openpyxl.load_workbook(source, data_only=False, keep_vba=False)
                try:
                    workbook.save(destination)
                finally:
                    workbook.close()
        return {
            "name": source.name,
            "path": destination,
            "source_path": source,
            "sha256": digest,
        }

    def _normalize_workbooks(self, sources: list[Path], workspace_dir: Path) -> list[dict[str, Any]]:
        normalized = []
        seen_hashes: set[str] = set()
        for source in sources:
            item = self._normalize_workbook(source, workspace_dir)
            if item["sha256"] in seen_hashes:
                continue
            seen_hashes.add(item["sha256"])
            normalized.append(item)
        name_counts: dict[str, int] = {}
        for item in normalized:
            name_counts[item["name"]] = name_counts.get(item["name"], 0) + 1
        for item in normalized:
            if name_counts[item["name"]] > 1:
                item["name"] = f"{item['name']} ({item['sha256'][:8]})"
        return normalized

    @staticmethod
    def _validate_workbook(path: Path) -> None:
        if not path.is_file():
            raise FileNotFoundError(f"Workbook not found: {path}")
        if path.suffix.lower() not in SUPPORTED_WORKBOOK_EXTENSIONS:
            supported = ", ".join(sorted(SUPPORTED_WORKBOOK_EXTENSIONS))
            raise ValueError(f"Unsupported workbook extension '{path.suffix}'; expected one of: {supported}")

    @staticmethod
    def _workbook_identities(normalized: list[dict[str, Any]]) -> dict[str, dict[str, str]]:
        return {
            str(item["path"].resolve()): {
                "name": str(item["name"]),
                "sha256": str(item["sha256"]),
            }
            for item in normalized
        }

    @staticmethod
    def _selected_sheet_names(path: Path, selected_sheets: tuple[str, ...]) -> list[str]:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            names = list(workbook.sheetnames)
        finally:
            workbook.close()
        return [name for name in names if not selected_sheets or name in selected_sheets]

    def _validate_sheet_filters(
        self,
        normalized: list[dict[str, Any]],
        selected_sheets: tuple[str, ...],
    ) -> None:
        if not selected_sheets:
            return
        missing: list[str] = []
        for item in normalized:
            workbook = openpyxl.load_workbook(item["path"], read_only=True, data_only=True)
            try:
                available = set(workbook.sheetnames)
            finally:
                workbook.close()
            absent = [name for name in selected_sheets if name not in available]
            if absent:
                missing.append(f"{item['name']}: {', '.join(absent)}")
        if missing:
            raise ValueError("Requested sheet(s) not found: " + "; ".join(missing))

    def _structure_results(
        self,
        records: Iterable[Any],
        normalized: list[dict[str, Any]],
        job_dir: Path,
        *,
        include_artifact_paths: bool,
    ) -> list[dict[str, Any]]:
        results = []
        for record in records:
            workbook_name = _workbook_name(record.workbook_path, normalized)
            artifact = None
            structure_text = None
            if record.structure_path.is_file():
                structure_text = record.structure_path.read_text(encoding="utf-8")
                item = next(
                    (value for value in normalized if value["name"] == workbook_name),
                    None,
                )
                if item is not None:
                    target_dir = sheet_artifact_dir(
                        workbook_artifact_dir(
                            job_dir / "workbooks",
                            workbook_name,
                            sources=False,
                        ),
                        record.sheet_name,
                    )
                    copy_artifact_tree(record.structure_path.parent, target_dir)
                    if include_artifact_paths:
                        artifact = (target_dir / "structure.yaml").relative_to(job_dir).as_posix()
            results.append(
                {
                    "workbook": workbook_name,
                    "sheet": record.sheet_name,
                    "status": record.status,
                    "structure": structure_text,
                    "artifact": artifact if include_artifact_paths else None,
                }
            )
        return results

    @staticmethod
    def _complete_structure_results(
        results: list[dict[str, Any]],
        normalized: list[dict[str, Any]],
        selected_sheets: tuple[str, ...] = (),
    ) -> list[dict[str, Any]]:
        by_identity = {(item["workbook"], item["sheet"]): item for item in results}
        completed = []
        for item in normalized:
            workbook = openpyxl.load_workbook(item["path"], read_only=True, data_only=True)
            try:
                sheet_names = [
                    name for name in workbook.sheetnames
                    if not selected_sheets or name in selected_sheets
                ]
            finally:
                workbook.close()
            for sheet_name in sheet_names:
                identity = (item["name"], sheet_name)
                completed.append(
                    by_identity.get(identity)
                    or {
                        "workbook": item["name"],
                        "sheet": sheet_name,
                        "status": "not_good",
                        "structure": None,
                        "artifact": None,
                    }
                )
        return completed

    @staticmethod
    def _answer_result(
        query: str,
        output: PipelineOutput,
        normalized: list[dict[str, Any]],
    ) -> dict[str, Any]:
        metadata = output.metadata or {}
        qa = dict(metadata.get("qa") or {})
        qa.pop("artifacts", None)
        return {
            "query": query,
            "answer": output.predicted_answer,
            "latency": output.latency,
            "token_usage": output.token_usage,
            "workbook": _workbook_name(Path(str(metadata.get("workbook_path", ""))), normalized),
            "sheets": metadata.get("workbook_sheets") or [],
            "verification": metadata.get("verification") or {},
            "retrieval": metadata.get("retrieval_info") or {},
            "qa": qa,
        }

    @staticmethod
    def _artifact_paths(job_dir: Path) -> list[str]:
        return sorted(path.relative_to(job_dir).as_posix() for path in job_dir.rglob("*") if path.is_file())


def _validate_stage(stage: str) -> Stage:
    value = str(stage).strip().lower()
    if value not in {"structure", "qa", "all"}:
        raise ValueError("stage must be one of: structure, qa, all")
    return value  # type: ignore[return-value]


def _validate_queries(queries: Iterable[str], *, required: bool) -> list[str]:
    result = [str(query).strip() for query in queries if str(query).strip()]
    if required and not result:
        raise ValueError("At least one non-empty query is required for qa and all stages")
    return result


def _normalize_sheet_filters(values: Iterable[str]) -> tuple[str, ...]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        for part in str(value).split(","):
            name = part.strip()
            if name and name not in seen:
                seen.add(name)
                result.append(name)
    return tuple(result)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _workbook_name(path: Path, normalized: list[dict[str, Any]]) -> str:
    try:
        resolved = path.resolve()
    except OSError:
        resolved = path
    for item in normalized:
        if resolved == item["path"].resolve():
            return str(item["name"])
    return path.name


def new_job_id() -> str:
    """Return a readable, filesystem-safe UTC timestamp for a generated job ID."""
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%S.%fZ")


__all__ = ["SUPPORTED_WORKBOOK_EXTENSIONS", "Stage", "TableAgentService", "new_job_id"]
